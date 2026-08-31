#!/usr/bin/env python3
"""Build an auditable three-day RTP observation report for Hilo and Plinko."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE = Path("/Users/robin/Desktop/waje data/新用户生命周期数据-含联运0825new.xlsx")
OUT = Path(__file__).resolve().parent
ASSETS = OUT / "assets"
KNOWLEDGE_REPORT = ROOT / "knowledge/02-数据/Waje-新上线游戏RTP与对照分析-2026-08-25.md"
WINDOW = (date(2026, 8, 21), date(2026, 8, 23))
TARGETS = ["Hilo", "Plinko"]
PEERS = ["limbo", "keno", "ColorGame", "Coin Flips"]

NAVY = "#173A63"
INK = "#18314F"
MUTED = "#5D7895"
GRID = "#D9E6F3"
BG = "#FFFFFF"
PANEL = "#F7FBFF"
HILO = "#2E7DCC"
PLINKO = "#249E8B"
EXPECTED = "#61758B"
AMBER = "#D99A16"
RED = "#C95757"
GREEN = "#1E8C69"


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def number(value: Any, default: float = 0.0) -> float:
    return float(value) if is_number(value) else default


def pct(value: float | None, digits: int = 2) -> str:
    return "N/A" if value is None else f"{value * 100:.{digits}f}%"


def pp(value: float | None, digits: int = 2) -> str:
    return "N/A" if value is None else f"{value * 100:+.{digits}f}pp"


def fmt_amount(value: float | None) -> str:
    if value is None:
        return "N/A"
    sign = "-" if value < 0 else ""
    value = abs(value)
    if value >= 1_000_000_000:
        return f"{sign}{value / 1_000_000_000:.2f}B"
    if value >= 1_000_000:
        return f"{sign}{value / 1_000_000:.2f}M"
    if value >= 1_000:
        return f"{sign}{value / 1_000:.1f}k"
    return f"{sign}{value:.0f}"


def fmt_exact(value: float | None) -> str:
    return "N/A" if value is None else f"{value:,.2f}"


def font(size: int, bold: bool = False) -> ImageFont.ImageFont:
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size, index=0)
    return ImageFont.load_default()


def text(draw: ImageDraw.ImageDraw, xy: tuple[float, float], value: str, size: int, fill: str = INK, bold: bool = False) -> None:
    draw.text((int(xy[0]), int(xy[1])), value, font=font(size, bold), fill=fill)


def line(draw: ImageDraw.ImageDraw, points: list[tuple[float, float]], fill: str, width: int = 4, dashed: bool = False) -> None:
    for start, end in zip(points, points[1:]):
        if not dashed:
            draw.line((start[0], start[1], end[0], end[1]), fill=fill, width=width)
            continue
        dx, dy = end[0] - start[0], end[1] - start[1]
        distance = max(1.0, (dx * dx + dy * dy) ** 0.5)
        step = 16.0
        for offset in range(0, int(distance), int(step * 2)):
            begin = min(1.0, offset / distance)
            finish = min(1.0, (offset + step) / distance)
            draw.line((start[0] + dx * begin, start[1] + dy * begin, start[0] + dx * finish, start[1] + dy * finish), fill=fill, width=width)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_values(path: Path) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    headers: dict[str, list[Any]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header = list(rows[0])
        headers[sheet_name] = header
        sheet_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            if not row or not any(value is not None for value in row):
                continue
            sheet_rows.append(dict(zip(header, row)))
        sheets[sheet_name] = sheet_rows
    return sheets, headers


def read_formula_summary(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    result: dict[str, Any] = {}
    for sheet_name in ["原始数据总数", "原始详细奖池"]:
        worksheet = workbook[sheet_name]
        formulas = []
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
                    if len(formulas) >= 8:
                        break
            if len(formulas) >= 8:
                break
        result[sheet_name] = formulas
    return result


def valid_game_row(row: dict[str, Any]) -> bool:
    required = ["完全下注额", "完全实际盈利", "完全预期盈利"]
    if not all(is_number(row.get(field)) for field in required):
        return False
    bet = number(row["完全下注额"])
    expected_rtp = 1 - number(row["完全预期盈利"]) / bet if bet else None
    return bool(bet > 0 and expected_rtp is not None and 0 < expected_rtp < 1.1)


def aggregate_game(rows: list[dict[str, Any]], game: str) -> dict[str, Any]:
    selected = [row for row in rows if str(row["游戏"]).lower() == game.lower() and valid_game_row(row)]
    if not selected:
        raise ValueError(f"No valid data for game={game}")
    full_bet = sum(number(row["完全下注额"]) for row in selected)
    actual_profit = sum(number(row["完全实际盈利"]) for row in selected)
    expected_profit = sum(number(row["完全预期盈利"]) for row in selected)
    base_bet = sum(number(row.get("基础下注额")) for row in selected)
    base_actual_profit = sum(number(row.get("基础实际盈利")) for row in selected)
    actual_rtp = 1 - actual_profit / full_bet
    expected_rtp = 1 - expected_profit / full_bet
    base_actual_rtp = 1 - base_actual_profit / base_bet if base_bet else None
    daily = []
    for row in sorted(selected, key=lambda item: item["日期"]):
        daily_bet = number(row["完全下注额"])
        daily_actual_profit = number(row["完全实际盈利"])
        daily_expected_profit = number(row["完全预期盈利"])
        daily_actual_rtp = 1 - daily_actual_profit / daily_bet
        daily_expected_rtp = 1 - daily_expected_profit / daily_bet
        daily.append(
            {
                "date": row["日期"].isoformat(),
                "complete_bet": daily_bet,
                "actual_profit": daily_actual_profit,
                "expected_profit": daily_expected_profit,
                "actual_rtp": daily_actual_rtp,
                "expected_rtp": daily_expected_rtp,
                "rtp_gap_pp": daily_actual_rtp - daily_expected_rtp,
                "relative_deviation": daily_actual_rtp / daily_expected_rtp - 1,
                "base_actual_rtp": number(row.get("基础真实回报比")) if is_number(row.get("基础真实回报比")) else None,
                "slippage": number(row.get("完全真实回报比")) - number(row.get("基础真实回报比")) if is_number(row.get("完全真实回报比")) and is_number(row.get("基础真实回报比")) else None,
            }
        )
    return {
        "game": game,
        "days": len({row["日期"] for row in selected}),
        "complete_bet": full_bet,
        "actual_profit": actual_profit,
        "expected_profit": expected_profit,
        "profit_delta": actual_profit - expected_profit,
        "actual_rtp": actual_rtp,
        "expected_rtp": expected_rtp,
        "rtp_gap_pp": actual_rtp - expected_rtp,
        "relative_deviation": actual_rtp / expected_rtp - 1,
        "base_actual_rtp": base_actual_rtp,
        "slippage": actual_rtp - base_actual_rtp if base_actual_rtp is not None else None,
        "daily": daily,
    }


def quality_checks(game_rows: list[dict[str, Any]], targets: dict[str, dict[str, Any]]) -> dict[str, Any]:
    checks: dict[str, Any] = {"status": "passed", "checks": [], "exclusions": []}
    date_set = sorted({row["日期"] for row in game_rows if isinstance(row.get("日期"), date)})
    expected_dates = [date(2026, 8, 21), date(2026, 8, 22), date(2026, 8, 23)]
    checks["checks"].append({"name": "分析日期覆盖", "expected": [item.isoformat() for item in expected_dates], "actual": [item.isoformat() for item in date_set], "status": "passed" if date_set == expected_dates else "failed"})

    actual_errors: list[float] = []
    expected_errors: list[float] = []
    relative_errors: list[float] = []
    excluded: list[str] = []
    for row in game_rows:
        if not is_number(row.get("完全下注额")) or number(row.get("完全下注额")) <= 0:
            excluded.append(str(row.get("游戏")))
            continue
        bet = number(row["完全下注额"])
        if is_number(row.get("完全实际盈利")) and is_number(row.get("完全真实回报比")):
            actual_errors.append(abs(number(row["完全真实回报比"]) - (1 - number(row["完全实际盈利"]) / bet)))
        if is_number(row.get("完全预期盈利")) and is_number(row.get("完全预期回报比")):
            expected_rtp = 1 - number(row["完全预期盈利"]) / bet
            expected_errors.append(abs(number(row["完全预期回报比"]) - expected_rtp))
            if expected_rtp > 0 and is_number(row.get("完全回报比差距")) and is_number(row.get("完全真实回报比")):
                relative_errors.append(abs(number(row["完全回报比差距"]) - (number(row["完全真实回报比"]) / expected_rtp - 1)))
        if not valid_game_row(row):
            excluded.append(str(row.get("游戏")))

    for label, errors, tolerance in [
        ("实际RTP复算", actual_errors, 0.0001),
        ("预期RTP复算", expected_errors, 0.0001),
        ("源表相对偏离率复算", relative_errors, 0.0002),
    ]:
        maximum = max(errors) if errors else None
        status = "passed" if maximum is not None and maximum <= tolerance else "failed"
        checks["checks"].append({"name": label, "samples": len(errors), "max_abs_error": maximum, "tolerance": tolerance, "status": status})

    for game, summary in targets.items():
        target_dates = [date.fromisoformat(row["date"]) for row in summary["daily"]]
        checks["checks"].append({"name": f"{game}三日覆盖", "expected_days": 3, "actual_days": len(target_dates), "dates": [item.isoformat() for item in target_dates], "status": "passed" if len(target_dates) == 3 else "failed"})

    checks["exclusions"] = sorted(set(excluded))
    if any(check["status"] == "failed" for check in checks["checks"]):
        checks["status"] = "failed"
    return checks


def platform_context(total_rows: list[dict[str, Any]], lifecycle_rows: list[dict[str, Any]]) -> dict[str, Any]:
    in_window = [row for row in total_rows if isinstance(row.get("日期"), date) and WINDOW[0] <= row["日期"] <= WINDOW[1]]
    if len(in_window) != 3:
        raise ValueError("Platform total sheet does not cover all three target dates")
    base_bet = sum(number(row.get("总基础下注额")) for row in in_window)
    complete_bet = sum(number(row.get("总完全下注额")) for row in in_window)
    base_actual = sum(number(row.get("总基础真实回报比")) * number(row.get("总基础下注额")) for row in in_window) / base_bet
    complete_actual = sum(number(row.get("总完全真实回报比")) * number(row.get("总完全下注额")) for row in in_window) / complete_bet

    lifecycle_window = [row for row in lifecycle_rows if isinstance(row.get("日期"), date) and WINDOW[0] <= row["日期"] <= WINDOW[1]]
    recharge = sum(number(row.get("当日充值总金额")) for row in lifecycle_window)
    withdrawal = sum(number(row.get("TX总金额")) for row in lifecycle_window)
    return {
        "base_bet": base_bet,
        "complete_bet": complete_bet,
        "base_actual_rtp": base_actual,
        "complete_actual_rtp": complete_actual,
        "slippage": complete_actual - base_actual,
        "tc_ratio": withdrawal / recharge if recharge else None,
        "recharge": recharge,
        "withdrawal": withdrawal,
        "daily": [
            {
                "date": row["日期"].isoformat(),
                "base_actual_rtp": number(row.get("总基础真实回报比")),
                "complete_actual_rtp": number(row.get("总完全真实回报比")),
                "slippage": number(row.get("贴水率")),
            }
            for row in sorted(in_window, key=lambda item: item["日期"])
        ],
    }


def draw_header(draw: ImageDraw.ImageDraw, title: str, subtitle: str) -> None:
    text(draw, (64, 44), title, 36, NAVY, True)
    text(draw, (64, 98), subtitle, 20, MUTED)


def save_daily_rtp(targets: dict[str, dict[str, Any]]) -> str:
    image = Image.new("RGB", (1800, 930), BG)
    draw = ImageDraw.Draw(image)
    draw_header(draw, "Hilo 与 Plinko：每日实际RTP与预期RTP", "2026年8月21—23日｜完全下注口径；实际与预期按同日下注额计算。")
    panels = [(85, 175, 865, 760), (935, 175, 1715, 760)]
    for (game, summary), (left, top, right, bottom), color in zip(targets.items(), panels, [HILO, PLINKO]):
        draw.rounded_rectangle((left, top, right, bottom), radius=18, fill=PANEL, outline=GRID, width=2)
        text(draw, (left + 28, top + 24), game, 28, NAVY, True)
        text(draw, (left + 28, top + 64), f"三日加权实际RTP {pct(summary['actual_rtp'])}｜预期 {pct(summary['expected_rtp'])}｜差异 {pp(summary['rtp_gap_pp'])}", 18, MUTED)
        chart_left, chart_right, chart_top, chart_bottom = left + 86, right - 42, top + 140, bottom - 75
        min_y, max_y = 0.90, 1.06
        for value in [0.90, 0.94, 0.98, 1.02, 1.06]:
            y = chart_bottom - (value - min_y) / (max_y - min_y) * (chart_bottom - chart_top)
            draw.line((chart_left, y, chart_right, y), fill=GRID, width=2)
            text(draw, (chart_left - 70, y - 12), f"{value * 100:.0f}%", 16, MUTED)
        x_values = [chart_left + (chart_right - chart_left) * index / 2 for index in range(3)]
        actual_points = []
        expected_points = []
        for x, row in zip(x_values, summary["daily"]):
            actual_y = chart_bottom - (row["actual_rtp"] - min_y) / (max_y - min_y) * (chart_bottom - chart_top)
            expected_y = chart_bottom - (row["expected_rtp"] - min_y) / (max_y - min_y) * (chart_bottom - chart_top)
            actual_points.append((x, actual_y))
            expected_points.append((x, expected_y))
            text(draw, (x - 35, chart_bottom + 18), row["date"][5:].replace("-", "/"), 18, MUTED)
            text(draw, (x - 35, actual_y - 34), pct(row["actual_rtp"], 1), 17, color, True)
        line(draw, expected_points, EXPECTED, 3, dashed=True)
        line(draw, actual_points, color, 5)
        for point in expected_points:
            draw.ellipse((point[0] - 6, point[1] - 6, point[0] + 6, point[1] + 6), fill=EXPECTED)
        for point in actual_points:
            draw.ellipse((point[0] - 9, point[1] - 9, point[0] + 9, point[1] + 9), fill=color)
        legend_y = bottom - 28
        draw.line((chart_left, legend_y, chart_left + 38, legend_y), fill=color, width=5)
        text(draw, (chart_left + 48, legend_y - 12), "实际RTP", 17, MUTED)
        line(draw, [(chart_left + 180, legend_y), (chart_left + 218, legend_y)], EXPECTED, 3, dashed=True)
        text(draw, (chart_left + 230, legend_y - 12), "预期RTP", 17, MUTED)
    text(draw, (66, 852), "解读：单日RTP超过100%仅表示当日派奖大于下注，不自动等同于系统故障；须结合有效局数、赔率/配置版本和大额派奖进一步复核。", 19, MUTED)
    output = ASSETS / "01_每日实际与预期RTP.png"
    image.save(output)
    return str(output.relative_to(OUT))


def save_target_summary(targets: dict[str, dict[str, Any]]) -> str:
    image = Image.new("RGB", (1800, 930), BG)
    draw = ImageDraw.Draw(image)
    draw_header(draw, "新游戏三日汇总：盈利差额与贴水影响", "2026年8月21—23日｜金额为来源统计单位；贴水率＝完全实际RTP－基础实际RTP。")
    left_panel = (80, 180, 850, 760)
    right_panel = (945, 180, 1720, 760)
    for panel, title in [(left_panel, "实际盈利与预期盈利"), (right_panel, "RTP差异与贴水率")]:
        draw.rounded_rectangle(panel, radius=18, fill=PANEL, outline=GRID, width=2)
        text(draw, (panel[0] + 28, panel[1] + 24), title, 26, NAVY, True)
    values = [(game, item["actual_profit"], item["expected_profit"]) for game, item in targets.items()]
    max_profit = max(max(actual, expected) for _, actual, expected in values) * 1.25
    for index, (game, actual, expected) in enumerate(values):
        base_x = left_panel[0] + 185 + index * 335
        bottom = left_panel[3] - 170
        top = left_panel[1] + 170
        text(draw, (base_x - 42, top - 58), game, 24, NAVY, True)
        for label, amount, offset, color in [("实际", actual, -48, HILO if game == "Hilo" else PLINKO), ("预期", expected, 48, EXPECTED)]:
            height = amount / max_profit * (bottom - top)
            x0 = base_x + offset - 36
            draw.rounded_rectangle((x0, bottom - height, x0 + 72, bottom), radius=8, fill=color)
            text(draw, (x0 - 12, bottom - height - 35), fmt_amount(amount), 18, color, True)
            text(draw, (x0 + 2, bottom + 16), label, 17, MUTED)
        text(draw, (base_x - 105, bottom + 54), f"实际较预期 {fmt_amount(actual - expected)}", 17, RED if actual - expected < 0 else GREEN)
    text(draw, (left_panel[0] + 40, left_panel[3] - 25), "正盈利对应平台收入；负的“实际较预期”表示三日实际盈利低于预期。", 16, MUTED)

    max_abs = max(abs(item["rtp_gap_pp"]) for item in targets.values()) * 1.35
    right_left, right_right = right_panel[0] + 110, right_panel[2] - 60
    midpoint = (right_left + right_right) / 2
    for tick in [-max_abs, -max_abs / 2, 0, max_abs / 2, max_abs]:
        x = midpoint + tick / (2 * max_abs) * (right_right - right_left)
        draw.line((x, right_panel[1] + 125, x, right_panel[3] - 110), fill=GRID if abs(tick) > 1e-9 else NAVY, width=2 if abs(tick) > 1e-9 else 3)
        text(draw, (x - 28, right_panel[3] - 82), f"{tick * 100:+.1f}", 16, MUTED)
    for index, (game, item) in enumerate(targets.items()):
        y = right_panel[1] + 230 + index * 185
        text(draw, (right_panel[0] + 32, y - 13), game, 23, NAVY, True)
        for label, value, color, y_offset in [("RTP差异", item["rtp_gap_pp"], HILO if game == "Hilo" else PLINKO, -18), ("贴水", item["slippage"], AMBER, 28)]:
            x = midpoint + value / (2 * max_abs) * (right_right - right_left)
            draw.line((midpoint, y + y_offset, x, y + y_offset), fill=color, width=10)
            draw.ellipse((x - 10, y + y_offset - 10, x + 10, y + y_offset + 10), fill=color)
            text(draw, (right_panel[0] + 145, y + y_offset - 14), label, 17, MUTED)
            text(draw, (x + 14 if x >= midpoint else x - 102, y + y_offset - 15), pp(value), 18, color, True)
    text(draw, (right_panel[0] + 32, right_panel[3] - 46), "RTP差异＝实际RTP－预期RTP；贴水率单独观察Bonus后实际RTP的变化。", 16, MUTED)
    output = ASSETS / "02_三日盈利差额与贴水率.png"
    image.save(output)
    return str(output.relative_to(OUT))


def save_peer_comparison(targets: dict[str, dict[str, Any]], peers: dict[str, dict[str, Any]]) -> str:
    ordered = [targets["Hilo"], targets["Plinko"], peers["limbo"], peers["keno"], peers["ColorGame"], peers["Coin Flips"]]
    image = Image.new("RGB", (1800, 970), BG)
    draw = ImageDraw.Draw(image)
    draw_header(draw, "同窗口轻量化游戏对照：RTP偏离与下注规模", "2026年8月21—23日｜六款游戏均以完全下注额加权；只纳入预期RTP有效的游戏。")
    left_panel = (72, 175, 1010, 810)
    right_panel = (1070, 175, 1728, 810)
    for panel, title in [(left_panel, "实际RTP相对预期的偏离（pp）"), (right_panel, "完全下注额")]:
        draw.rounded_rectangle(panel, radius=18, fill=PANEL, outline=GRID, width=2)
        text(draw, (panel[0] + 26, panel[1] + 24), title, 26, NAVY, True)
    gaps = [item["rtp_gap_pp"] for item in ordered]
    min_gap, max_gap = min(gaps) - 0.005, max(gaps) + 0.005
    chart_left, chart_right = left_panel[0] + 210, left_panel[2] - 50
    midpoint = chart_left + (0 - min_gap) / (max_gap - min_gap) * (chart_right - chart_left)
    for tick in [-0.015, -0.005, 0, 0.005, 0.015, 0.025]:
        if tick < min_gap or tick > max_gap:
            continue
        x = chart_left + (tick - min_gap) / (max_gap - min_gap) * (chart_right - chart_left)
        draw.line((x, left_panel[1] + 105, x, left_panel[3] - 70), fill=NAVY if tick == 0 else GRID, width=3 if tick == 0 else 2)
        text(draw, (x - 25, left_panel[3] - 46), f"{tick * 100:+.1f}", 15, MUTED)
    max_stake = max(item["complete_bet"] for item in ordered)
    for index, item in enumerate(ordered):
        y = left_panel[1] + 145 + index * 82
        color = HILO if item["game"] == "Hilo" else PLINKO if item["game"] == "Plinko" else EXPECTED
        text(draw, (left_panel[0] + 28, y - 12), item["game"], 18, NAVY, item["game"] in TARGETS)
        x = chart_left + (item["rtp_gap_pp"] - min_gap) / (max_gap - min_gap) * (chart_right - chart_left)
        draw.line((midpoint, y, x, y), fill=color, width=8)
        draw.ellipse((x - 10, y - 10, x + 10, y + 10), fill=color)
        text(draw, (x + 16 if x >= midpoint else x - 95, y - 12), pp(item["rtp_gap_pp"]), 17, color, True)

        y_bar = right_panel[1] + 145 + index * 82
        bar_left, bar_right = right_panel[0] + 170, right_panel[2] - 85
        bar_width = item["complete_bet"] / max_stake * (bar_right - bar_left)
        draw.rounded_rectangle((bar_left, y_bar - 17, bar_left + bar_width, y_bar + 17), radius=8, fill=color)
        text(draw, (right_panel[0] + 28, y_bar - 12), item["game"], 18, NAVY, item["game"] in TARGETS)
        text(draw, (bar_left + bar_width + 12, y_bar - 12), fmt_amount(item["complete_bet"]), 17, color, True)
    text(draw, (72, 858), "解读：Hilo、Plinko的三日累计RTP偏离落在同期轻量化对照游戏观察范围内；但Hilo投注规模远低于成熟游戏，三日结果仅可作为早期观察。", 19, MUTED)
    output = ASSETS / "03_同窗口轻量化游戏对照.png"
    image.save(output)
    return str(output.relative_to(OUT))


def save_platform_context(context: dict[str, Any]) -> str:
    image = Image.new("RGB", (1800, 900), BG)
    draw = ImageDraw.Draw(image)
    draw_header(draw, "三日全盘RTP与资金背景", "2026年8月21—23日｜全盘指标仅作背景，不用于归因Hilo或Plinko。")
    panel = (85, 178, 1715, 725)
    draw.rounded_rectangle(panel, radius=18, fill=PANEL, outline=GRID, width=2)
    chart_left, chart_right, chart_top, chart_bottom = 185, 1630, 285, 610
    min_y, max_y = 0.962, 0.970
    for value in [0.962, 0.964, 0.966, 0.968, 0.970]:
        y = chart_bottom - (value - min_y) / (max_y - min_y) * (chart_bottom - chart_top)
        draw.line((chart_left, y, chart_right, y), fill=GRID, width=2)
        text(draw, (82, y - 10), f"{value * 100:.1f}%", 16, MUTED)
    xs = [chart_left + (chart_right - chart_left) * index / 2 for index in range(3)]
    basic = []
    complete = []
    for x, item in zip(xs, context["daily"]):
        y_base = chart_bottom - (item["base_actual_rtp"] - min_y) / (max_y - min_y) * (chart_bottom - chart_top)
        y_complete = chart_bottom - (item["complete_actual_rtp"] - min_y) / (max_y - min_y) * (chart_bottom - chart_top)
        basic.append((x, y_base))
        complete.append((x, y_complete))
        text(draw, (x - 38, chart_bottom + 20), item["date"][5:].replace("-", "/"), 18, MUTED)
    line(draw, basic, EXPECTED, 4)
    line(draw, complete, HILO, 5)
    for point in basic:
        draw.ellipse((point[0] - 7, point[1] - 7, point[0] + 7, point[1] + 7), fill=EXPECTED)
    for point in complete:
        draw.ellipse((point[0] - 9, point[1] - 9, point[0] + 9, point[1] + 9), fill=HILO)
    text(draw, (185, 675), f"三日基础实际RTP {pct(context['base_actual_rtp'])}｜完全实际RTP {pct(context['complete_actual_rtp'])}｜整体贴水率 {pp(context['slippage'])}", 21, NAVY, True)
    text(draw, (185, 730), f"三日全盘TC比（提款÷充值） {pct(context['tc_ratio'])}；该指标没有游戏维度，不能解释单个游戏的RTP变化。", 19, MUTED)
    text(draw, (85, 804), "备注：TC比用于同期资金风险背景；新游戏RTP判断仍以游戏级完全下注、实际盈利和预期盈利为控制口径。", 17, MUTED)
    output = ASSETS / "04_全盘RTP与TC背景.png"
    image.save(output)
    return str(output.relative_to(OUT))


def markdown_report(data: dict[str, Any], charts: list[str]) -> str:
    hilo = data["targets"]["Hilo"]
    plinko = data["targets"]["Plinko"]
    peers = data["peers"]
    context = data["platform_context"]
    peer_gaps = [item["rtp_gap_pp"] for item in peers.values()]
    return f"""---
type: game-rtp-audit
status: provisional
window: 2026-08-21—2026-08-23
source: 新用户生命周期数据-含联运0825new.xlsx
updated: 2026-08-25
---

# Waje 新上线游戏 RTP 与对照分析｜2026年8月21—23日

## 核心结论

本次只观察 **Hilo** 与 **Plinko** 上线后的三个完整自然日。两款游戏的三日加权实际RTP均略高于预期，但尚未出现累计口径持续偏离的证据；结论为 **早期观察，需继续复核**，不能写成“已确认正常”或“无故障”。

- **Hilo：样本很小，日波动明显。** 完全下注额 **{fmt_exact(hilo['complete_bet'])}**，占同期全盘完全下注额 **{pct(hilo['platform_stake_share'], 3)}**；实际RTP **{pct(hilo['actual_rtp'])}**，较预期 **{pct(hilo['expected_rtp'])}** 高 **{pp(hilo['rtp_gap_pp'])}**。8月21日实际RTP为 **{pct(hilo['daily'][0]['actual_rtp'])}**，8月23日为 **{pct(hilo['daily'][-1]['actual_rtp'])}**，三天不足以排除高波动影响。
- **Plinko：累计RTP接近预期，但8月22日应重点复核。** 完全下注额 **{fmt_exact(plinko['complete_bet'])}**，实际RTP **{pct(plinko['actual_rtp'])}**，较预期高 **{pp(plinko['rtp_gap_pp'])}**；8月22日实际RTP达到 **{pct(plinko['daily'][1]['actual_rtp'])}**，高于当日预期 **{pp(plinko['daily'][1]['rtp_gap_pp'])}**。三日汇总尚未显示持续异常，但需核查大倍率、ROW、难度与封顶触发。
- **横向对照没有出现新游戏独有的累计偏离。** 同期四款轻量化对照游戏的RTP差异范围为 **{pp(min(peer_gaps))} 至 {pp(max(peer_gaps))}**；Hilo为 **{pp(hilo['rtp_gap_pp'])}**，Plinko为 **{pp(plinko['rtp_gap_pp'])}**。该范围仅说明“短窗波动可比”，不是故障阈值。
- **贴水对Plinko的RTP抬升更明显。** Hilo三日贴水率为 **{pp(hilo['slippage'])}**，Plinko为 **{pp(plinko['slippage'])}**；两项必须与RTP偏离分开阅读。

## 1. 新游戏三日RTP拆解

| 游戏 | 完全下注额 | 实际RTP | 预期RTP | RTP差异 | 相对偏离率 | 状态 |
|---|---:|---:|---:|---:|---:|---|
| Hilo | {fmt_exact(hilo['complete_bet'])} | {pct(hilo['actual_rtp'])} | {pct(hilo['expected_rtp'])} | {pp(hilo['rtp_gap_pp'])} | {pct(hilo['relative_deviation'])} | 早期观察·样本小 |
| Plinko | {fmt_exact(plinko['complete_bet'])} | {pct(plinko['actual_rtp'])} | {pct(plinko['expected_rtp'])} | {pp(plinko['rtp_gap_pp'])} | {pct(plinko['relative_deviation'])} | 早期观察·波动需复核 |

实际RTP与预期RTP均按三日的下注额和盈利先汇总后计算。源表“完全回报比差距”是相对偏离率，不是百分点差，因此报告将两者分列。

![Hilo与Plinko每日实际与预期RTP](../../analysis/new_game_rtp_audit_2026_08_25/assets/{charts[0].split('/')[-1]})

**解读：** Hilo与Plinko都出现了单日上下波动，尤其Plinko在8月22日实际RTP超过100%。单日玩家回报高于下注不自动代表故障；只有在有效局数、最终派奖、配置版本和大额派奖分布齐备后，才能判断是否存在配置或结算问题。

## 2. 下注、盈利与贴水影响

| 游戏 | 实际盈利 | 预期盈利 | 实际较预期 | 基础实际RTP | 完全实际RTP | 贴水率 |
|---|---:|---:|---:|---:|---:|---:|
| Hilo | {fmt_exact(hilo['actual_profit'])} | {fmt_exact(hilo['expected_profit'])} | {fmt_exact(hilo['profit_delta'])} | {pct(hilo['base_actual_rtp'])} | {pct(hilo['actual_rtp'])} | {pp(hilo['slippage'])} |
| Plinko | {fmt_exact(plinko['actual_profit'])} | {fmt_exact(plinko['expected_profit'])} | {fmt_exact(plinko['profit_delta'])} | {pct(plinko['base_actual_rtp'])} | {pct(plinko['actual_rtp'])} | {pp(plinko['slippage'])} |

两款游戏的实际盈利仍为正，但均低于对应预期：Hilo少 **{fmt_amount(abs(hilo['profit_delta']))}**，Plinko少 **{fmt_amount(abs(plinko['profit_delta']))}**。在本表口径下，这与实际RTP高于预期一致；不应将其直接解释为收入故障。

![新游戏三日盈利差额与贴水率](../../analysis/new_game_rtp_audit_2026_08_25/assets/{charts[1].split('/')[-1]})

## 3. 与轻量化游戏的同窗口对照

| 游戏 | 完全下注额 | 实际RTP | 预期RTP | RTP差异 | 实际较预期盈利 | 贴水率 |
|---|---:|---:|---:|---:|---:|---:|
""" + "\n".join(
        f"| {item['game']} | {fmt_exact(item['complete_bet'])} | {pct(item['actual_rtp'])} | {pct(item['expected_rtp'])} | {pp(item['rtp_gap_pp'])} | {fmt_exact(item['profit_delta'])} | {pp(item['slippage'])} |"
        for item in [hilo, plinko, peers["limbo"], peers["keno"], peers["ColorGame"], peers["Coin Flips"]]
    ) + f"""

![同窗口轻量化游戏对照](../../analysis/new_game_rtp_audit_2026_08_25/assets/{charts[2].split('/')[-1]})

**解读：** Hilo的RTP差异与Keno、ColorGame接近，但其投注额远低于成熟游戏，不能据此得出稳定结论。Plinko的累计偏离更接近对照组中位水平，但8月22日的单日高RTP仍需要按玩法参数继续下钻。

## 4. 同期全盘背景

三日全盘完全实际RTP为 **{pct(context['complete_actual_rtp'])}**，基础实际RTP为 **{pct(context['base_actual_rtp'])}**，整体贴水率为 **{pp(context['slippage'])}**。全盘TC比为 **{pct(context['tc_ratio'])}**，其定义为提款总金额÷充值总金额，仅反映同期资金背景，**不用于解释单个游戏的RTP变化**。

![三日全盘RTP与TC背景](../../analysis/new_game_rtp_audit_2026_08_25/assets/{charts[3].split('/')[-1]})

## 5. 问题定位与下一步

1. **P0：先补正式RTP审计所需事实。** 按 `game_id × 日期 × 配置版本 × 币种` 提供有效下注、最终派奖、有效局数、取消/退款/未结算、免费注和Bonus的聚合结果。
2. **P0：Plinko优先拆玩法参数。** 补齐难度、ROW、球数、下注档位、命中倍率和封顶触发，重点核查8月22日高RTP对应的具体组合。
3. **P0：Hilo补齐连续玩法链路。** 提供猜测、Skip、Cash Out、连续轮次、赔率分布及失败轮次聚合，判断短窗波动是否由少量连续高赔率结果驱动。
4. **P1：满7个完整自然日后复测。** 使用同一加权公式复算RTP、贴水和实际/预期盈利差；若累计偏离仍持续，再进入配置和结算排查。

## 附录：统计口径与数据限制

- 数据窗口：2026年8月21日—8月23日；8月17日处于隐藏模式，未进入本报告。
- 实际RTP＝1－完全实际盈利÷完全下注额；预期RTP＝1－完全预期盈利÷完全下注额。
- RTP差异以百分点展示；相对偏离率＝实际RTP÷预期RTP－1。
- 贴水率＝完全实际RTP－基础实际RTP，观察Bonus后玩家实际回报的变化。
- TC比＝提款总金额÷充值总金额，仅有生命周期维度，无法映射到单款游戏。
- 当前来源不含有效局数、最终派奖、配置版本、玩法参数和取消/退款状态；因此本报告属于**早期观察**，不构成最终故障鉴定。
"""


def xml_table(headers: list[str], rows: list[list[str]]) -> str:
    head = "".join(f"<th background-color=\"light-gray\"><p><b>{escape(item)}</b></p></th>" for item in headers)
    body = "".join("<tr>" + "".join(f"<td><p>{escape(item)}</p></td>" for item in row) + "</tr>" for row in rows)
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def feishu_xml(data: dict[str, Any], charts: list[str]) -> str:
    hilo = data["targets"]["Hilo"]
    plinko = data["targets"]["Plinko"]
    peers = data["peers"]
    context = data["platform_context"]
    target_rows = [
        ["Hilo", fmt_exact(hilo["complete_bet"]), pct(hilo["actual_rtp"]), pct(hilo["expected_rtp"]), pp(hilo["rtp_gap_pp"])],
        ["Plinko", fmt_exact(plinko["complete_bet"]), pct(plinko["actual_rtp"]), pct(plinko["expected_rtp"]), pp(plinko["rtp_gap_pp"])],
    ]
    peer_rows = [[item["game"], fmt_exact(item["complete_bet"]), pct(item["actual_rtp"]), pct(item["expected_rtp"]), pp(item["rtp_gap_pp"])] for item in [hilo, plinko, peers["limbo"], peers["keno"], peers["ColorGame"], peers["Coin Flips"]]]
    return f"""<title>Waje 新上线游戏 RTP 与对照分析｜2026年8月21—23日</title>
<p>统计窗口：2026年8月21日—8月23日｜数据来源：新用户生命周期数据-含联运0825new.xlsx｜报告性质：早期观察</p>
<h1 seq=\"auto\">核心结论</h1>
<callout emoji=\"🔎\" background-color=\"light-blue\" border-color=\"blue\"><p><b>总体判断：</b>Hilo与Plinko均只积累3个完整自然日。两款游戏的三日加权实际RTP略高于预期，但尚无累计口径持续偏离的证据；当前结论为<b>早期观察、持续复核</b>，不能写成“已确认正常”或“无故障”。</p></callout>
<callout emoji=\"🎯\" background-color=\"light-yellow\" border-color=\"yellow\"><p><b>Plinko优先复核：</b>8月22日实际RTP为<b>{pct(plinko['daily'][1]['actual_rtp'])}</b>，高于当日预期<b>{pp(plinko['daily'][1]['rtp_gap_pp'])}</b>；三日累计仅高<b>{pp(plinko['rtp_gap_pp'])}</b>，需结合难度、ROW、球数、倍率和封顶触发继续核查。</p></callout>
<callout emoji=\"📊\" background-color=\"light-gray\" border-color=\"gray\"><p><b>样本限制：</b>Hilo完全下注额为<b>{fmt_exact(hilo['complete_bet'])}</b>，仅占同期全盘完全下注额<b>{pct(hilo['platform_stake_share'], 3)}</b>；日波动不适合作为配置异常的最终依据。</p></callout>
<h1 seq=\"auto\">两款新游戏的三日RTP拆解</h1>
<p>实际与预期均按三日的完全下注和盈利汇总后计算；RTP差异以百分点展示，源表“完全回报比差距”单列为相对偏离率。</p>
{xml_table(["游戏", "完全下注额", "实际RTP", "预期RTP", "RTP差异"], target_rows)}
<img path=\"@./{charts[0]}\" width=\"880\" caption=\"Hilo与Plinko每日实际RTP与预期RTP\"/>
<p><b>解读：</b>Hilo从8月21日的{pct(hilo['daily'][0]['actual_rtp'])}波动至8月23日的{pct(hilo['daily'][-1]['actual_rtp'])}；Plinko在8月22日出现单日RTP高于100%。单日派奖高于下注不自动等同于故障，需结合有效局数、最终派奖和配置版本复核。</p>
<h1 seq=\"auto\">下注、盈利与贴水影响</h1>
{xml_table(["游戏", "实际盈利", "预期盈利", "实际较预期", "贴水率"], [["Hilo", fmt_exact(hilo['actual_profit']), fmt_exact(hilo['expected_profit']), fmt_exact(hilo['profit_delta']), pp(hilo['slippage'])], ["Plinko", fmt_exact(plinko['actual_profit']), fmt_exact(plinko['expected_profit']), fmt_exact(plinko['profit_delta']), pp(plinko['slippage'])]])}
<img path=\"@./{charts[1]}\" width=\"880\" caption=\"实际盈利、预期盈利与贴水率\"/>
<p><b>解读：</b>两款游戏实际盈利均为正，但分别低于预期{fmt_amount(abs(hilo['profit_delta']))}和{fmt_amount(abs(plinko['profit_delta']))}，与三日实际RTP高于预期一致。贴水率需要与RTP偏离分开判断：Hilo为{pp(hilo['slippage'])}，Plinko为{pp(plinko['slippage'])}。</p>
<h1 seq=\"auto\">与轻量化游戏的同窗口对照</h1>
{xml_table(["游戏", "完全下注额", "实际RTP", "预期RTP", "RTP差异"], peer_rows)}
<img path=\"@./{charts[2]}\" width=\"880\" caption=\"六款轻量化游戏的RTP偏离与下注规模对照\"/>
<p><b>解读：</b>Hilo和Plinko的累计RTP偏离没有表现出与同期对照游戏显著不同的单向特征；但Hilo投注规模很小，Plinko存在明显单日波动，因此两者都需要在满7个完整自然日后复测。</p>
<h1 seq=\"auto\">同期全盘背景</h1>
<p>三日全盘完全实际RTP为<b>{pct(context['complete_actual_rtp'])}</b>，基础实际RTP为<b>{pct(context['base_actual_rtp'])}</b>，整体贴水率为<b>{pp(context['slippage'])}</b>。同期TC比为<b>{pct(context['tc_ratio'])}</b>，定义为提款÷充值，仅作资金背景，不能归因到单款游戏。</p>
<img path=\"@./{charts[3]}\" width=\"880\" caption=\"三日全盘RTP与TC背景\"/>
<h1 seq=\"auto\">下一步</h1>
<ol><li><b>P0：</b>补齐game_id、配置版本、有效局数、最终派奖、有效下注、取消/退款/未结算、免费注和Bonus聚合字段。</li><li><b>P0：</b>Plinko补齐难度、ROW、球数、下注档位、命中倍率和封顶触发；Hilo补齐猜测、Skip、Cash Out、连续轮次和赔率分布。</li><li><b>P1：</b>满7个完整自然日后，按同一公式复测加权RTP、贴水率和实际/预期盈利差；若累计偏离持续，再进入配置和结算排查。</li></ol>
<h1 seq=\"auto\">口径说明</h1>
<p>实际RTP＝1－完全实际盈利÷完全下注额；预期RTP＝1－完全预期盈利÷完全下注额；贴水率＝完全实际RTP－基础实际RTP；TC比＝提款总金额÷充值总金额。当前来源不含有效局数、最终派奖、配置版本和玩法参数，因此本报告属于观察性分析，不构成最终故障鉴定。</p>
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_SOURCE))
    args = parser.parse_args()
    source = Path(args.source).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(source)
    ASSETS.mkdir(parents=True, exist_ok=True)

    sheets, headers = load_values(source)
    game_rows = []
    for row in sheets["生命周期奖池分游戏汇总"]:
        item = dict(row)
        item["日期"] = as_date(item.get("日期"))
        if item["日期"] and WINDOW[0] <= item["日期"] <= WINDOW[1]:
            game_rows.append(item)
    total_rows = []
    for row in sheets["原始数据总数"]:
        item = dict(row)
        item["日期"] = as_date(item.get("日期"))
        total_rows.append(item)
    lifecycle_rows = []
    for row in sheets["原始数据活跃周期"]:
        item = dict(row)
        item["日期"] = as_date(item.get("日期"))
        lifecycle_rows.append(item)

    target_metrics = {game: aggregate_game(game_rows, game) for game in TARGETS}
    peer_metrics = {game: aggregate_game(game_rows, game) for game in PEERS}
    context = platform_context(total_rows, lifecycle_rows)
    for item in target_metrics.values():
        item["platform_stake_share"] = item["complete_bet"] / context["complete_bet"]
        item["daily_rtp_range_pp"] = max(row["actual_rtp"] for row in item["daily"]) - min(row["actual_rtp"] for row in item["daily"])

    checks = quality_checks(game_rows, target_metrics)
    charts = [
        save_daily_rtp(target_metrics),
        save_target_summary(target_metrics),
        save_peer_comparison(target_metrics, peer_metrics),
        save_platform_context(context),
    ]
    data = {
        "title": "Waje 新上线游戏 RTP 与对照分析｜2026年8月21—23日",
        "window": {"start": WINDOW[0].isoformat(), "end": WINDOW[1].isoformat()},
        "source": {"path": str(source), "file_name": source.name, "sha256": sha256_file(source), "modified_at": datetime.fromtimestamp(source.stat().st_mtime).astimezone().isoformat(timespec="seconds")},
        "source_headers": {name: [str(value) for value in values if value is not None] for name, values in headers.items()},
        "targets": target_metrics,
        "peers": peer_metrics,
        "platform_context": context,
        "quality_status": checks["status"],
        "limitations": [
            "仅覆盖Hilo、Plinko上线后的3个完整自然日。",
            "来源未提供有效局数、最终派奖、配置版本、取消/退款/未结算及玩法参数，不能形成最终故障鉴定。",
            "TC比仅有生命周期维度，不能归因至单个游戏。",
        ],
    }
    manifest = {
        "source_file": source.name,
        "window": data["window"],
        "target_games": TARGETS,
        "peer_games": PEERS,
        "charts": [
            {"path": chart, "window": "2026-08-21—2026-08-23", "source": "新用户生命周期数据-含联运0825new.xlsx"}
            for chart in charts
        ],
    }
    (OUT / "report_data.json").write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT / "quality_checks.json").write_text(json.dumps(checks, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT / "source_manifest.json").write_text(json.dumps({"source": data["source"], "formula_samples": read_formula_summary(source), "window": data["window"]}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT / "chart_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report = markdown_report(data, charts)
    (OUT / "report.md").write_text(report, encoding="utf-8")
    KNOWLEDGE_REPORT.write_text(report, encoding="utf-8")
    (OUT / "feishu_release.xml").write_text(feishu_xml(data, charts), encoding="utf-8")
    receipt = {
        "status": "passed" if checks["status"] == "passed" else "failed",
        "window": data["window"],
        "target_games": TARGETS,
        "charts": charts,
        "markdown": str(KNOWLEDGE_REPORT),
        "feishu_xml": str(OUT / "feishu_release.xml"),
    }
    (OUT / "run_receipt.json").write_text(json.dumps(receipt, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(receipt, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
