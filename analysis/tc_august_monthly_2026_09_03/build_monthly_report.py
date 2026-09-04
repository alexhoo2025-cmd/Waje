#!/usr/bin/env python3
"""Build the August TC + game/lifecycle monthly report.

The input is limited to aggregate lifecycle exports and an aggregate-only
Metabase result.  No user, order-detail, account, device or credential data
is read.  The script keeps a complete machine-readable snapshot and applies
the user's display rule: remove only columns that are entirely zero/blank in
the displayed scope.
"""

from __future__ import annotations

import base64
import csv
import hashlib
import html
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent
ASSETS = OUT / "assets"
HTML_OUT = ROOT / "output/html/Waje-8月全产品TC比拆解分析与审计-2026-09-03.html"
MD_OUT = ROOT / "knowledge/02-数据/Waje-8月全产品TC比拆解分析与审计-2026-09-03.md"
TC_FILE = OUT / "tc_query_result.json"
EXISTING_TC = ROOT / "analysis/tc_game_rtp_tracking_2026_09_02/metabase_tc_daily_2026_08_19_09_01.csv"
NEW_USER_FILE = ROOT / "data/outputs/origin_new_user/2026-09-01-26d/source-data.json"

MONTH_START = date(2026, 8, 1)
MONTH_END = date(2026, 8, 31)
LIFECYCLES = (1, 2, 3, 4)
PERIODS = [
    {"key": "week_1", "label": "第1周（8/1—8/7）", "start": date(2026, 8, 1), "end": date(2026, 8, 7)},
    {"key": "week_2", "label": "第2周（8/8—8/14）", "start": date(2026, 8, 8), "end": date(2026, 8, 14)},
    {"key": "week_3", "label": "第3周（8/15—8/21）", "start": date(2026, 8, 15), "end": date(2026, 8, 21)},
    {"key": "week_4", "label": "第4周（8/22—8/31，月末10天）", "start": date(2026, 8, 22), "end": date(2026, 8, 31)},
]

INK = "#17324D"
MUTED = "#61758B"
GRID = "#D9E6F3"
PANEL = "#F7FBFF"
WHITE = "#FFFFFF"
BLUE = "#2563EB"
NAVY = "#173A63"
ORANGE = "#EA580C"
GOLD = "#B7791F"
TEAL = "#0F8A70"
RED = "#B4535B"
GREEN = "#15803D"


def dump(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def clean(value: Any) -> str:
    return "" if value is None else "".join(str(value).split())


def num(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "—", "N/A", "NULL", "null"}:
        return None
    try:
        result = float(text.rstrip("%"))
        if text.endswith("%"):
            result /= 100
        return result if math.isfinite(result) else None
    except ValueError:
        return None


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def pct(value: float | None, digits: int = 2) -> str:
    return "N/A" if value is None else f"{value * 100:.{digits}f}%"


def pp(value: float | None, digits: int = 2) -> str:
    return "N/A" if value is None else f"{value * 100:+.{digits}f}个百分点"


def money(value: float | None) -> str:
    if value is None:
        return "N/A"
    sign = "-" if value < 0 else ""
    v = abs(value)
    if v >= 100_000_000:
        return f"{sign}{v / 100_000_000:.2f}亿"
    if v >= 10_000:
        return f"{sign}{v / 10_000:.2f}万"
    return f"{value:,.0f}"


def raw_money(value: float | None) -> str:
    return "N/A" if value is None else f"{value:,.2f}"


def valid_expected(value: float | None) -> bool:
    return value is not None and 0 < value <= 1.1


def font(size: int, bold: bool = False) -> ImageFont.ImageFont:
    candidates = [
        ("/System/Library/Fonts/PingFang.ttc", 1 if bold else 0),
        ("/System/Library/Fonts/STHeiti Medium.ttc", 0),
        ("/System/Library/Fonts/Supplemental/Arial Unicode.ttf", 0),
    ]
    for path, index in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size, index=index)
    return ImageFont.load_default()


def draw_text(draw: ImageDraw.ImageDraw, xy: tuple[int, int], value: str, size: int, fill: str = INK, bold: bool = False, anchor: str | None = None) -> None:
    draw.text(xy, value, font=font(size, bold), fill=fill, anchor=anchor)


def line_chart(path: Path, title: str, labels: list[str], series: list[tuple[str, list[float | None], str]], y_min: float, y_max: float, unit: str, annotate: bool = False, annotation_digits: int = 2, x_label_step: int = 1) -> None:
    width, height = 1600, 760
    image = Image.new("RGB", (width, height), WHITE)
    draw = ImageDraw.Draw(image)
    draw_text(draw, (54, 34), title, 32, bold=True)
    left, right, top, bottom = 120, 80, 112, 110
    plot_w, plot_h = width - left - right, height - top - bottom
    for i in range(6):
        y = top + plot_h * i / 5
        value = y_max - (y_max - y_min) * i / 5
        draw.line((left, y, width - right, y), fill=GRID, width=2)
        draw_text(draw, (left - 16, int(y)), f"{value:.0f}{unit}", 18, fill=MUTED, anchor="ra")
    if len(labels) == 1:
        xs = [left + plot_w / 2]
    else:
        xs = [left + plot_w * i / (len(labels) - 1) for i in range(len(labels))]
    for label_index, (x, label) in enumerate(zip(xs, labels)):
        if label_index == 0 or label_index == len(labels) - 1 or label_index % max(1, x_label_step) == 0:
            draw_text(draw, (int(x), height - 72), label, 18, fill=MUTED, anchor="ma")
    for series_index, (name, values, color) in enumerate(series):
        points = []
        for x, value in zip(xs, values):
            if value is None:
                points.append(None)
            else:
                y = top + (y_max - value) / (y_max - y_min) * plot_h
                points.append((int(x), int(y)))
        last = None
        for point_index, point in enumerate(points):
            if point is not None and last is not None:
                draw.line((last[0], last[1], point[0], point[1]), fill=color, width=5)
            if point is not None:
                draw.ellipse((point[0] - 8, point[1] - 8, point[0] + 8, point[1] + 8), fill=color, outline=WHITE, width=2)
                if annotate:
                    label_y = point[1] - 24 if series_index == 0 else point[1] + 24
                    draw_text(draw, (point[0], label_y), f"{values[point_index]:.{annotation_digits}f}{unit}", 16, fill=color, bold=True, anchor="mm")
                last = point
            else:
                last = None
        lx = width - right - 260
        ly = 45 + series.index((name, values, color)) * 34
        draw.line((lx, ly, lx + 28, ly), fill=color, width=5)
        draw_text(draw, (lx + 40, ly), name, 18, fill=INK, anchor="lm")
    image.save(path)


def bar_chart(path: Path, title: str, labels: list[str], values: list[float], colors: list[str], value_labels: list[str], subtitle: str = "") -> None:
    width, height = 1600, max(620, 150 + len(labels) * 34)
    image = Image.new("RGB", (width, height), WHITE)
    draw = ImageDraw.Draw(image)
    draw_text(draw, (54, 34), title, 32, bold=True)
    if subtitle:
        draw_text(draw, (54, 82), subtitle, 18, fill=MUTED)
    left, right = 330, 180
    top = 130 if subtitle else 105
    max_value = max(values) if values else 1
    bar_h = 22
    for i, (label, value, color, val_label) in enumerate(zip(labels, values, colors, value_labels)):
        y = top + i * 34
        draw_text(draw, (left - 18, y + bar_h / 2), label, 18, fill=INK, anchor="ra")
        draw.rounded_rectangle((left, y, left + (width - left - right) * value / max_value, y + bar_h), radius=8, fill=color)
        draw_text(draw, (left + (width - left - right) * value / max_value + 12, y + bar_h / 2), val_label, 17, fill=MUTED, anchor="lm")
    image.save(path)


def lifecycle_combo_chart(path: Path, rows: list[dict[str, Any]], title: str = "8月生命周期1—4：下注额（柱）与RTP（折线）", category_getter: Callable[[dict[str, Any]], str] | None = None, show_expected: bool = True, bar_label_getter: Callable[[dict[str, Any]], str] | None = None) -> None:
    """Draw a dual-axis combo chart: bet amount bars + RTP lines."""
    width, height = 1800, 860
    image = Image.new("RGB", (width, height), WHITE)
    draw = ImageDraw.Draw(image)
    draw_text(draw, (54, 32), title, 32, bold=True)
    subtitle = "左轴：完全下注额（亿）；右轴：实际RTP与配置预期RTP（%）" if show_expected else "左轴：完全下注额（亿）；右轴：实际RTP（%）；柱顶标注：下注额｜月度下注额占比；配置预期RTP见表格"
    draw_text(draw, (54, 82), subtitle, 18, fill=MUTED)
    left, right, top, bottom = 145, 160, 135, 110
    plot_w, plot_h = width - left - right, height - top - bottom
    max_bet = max((r["full_bet"] for r in rows), default=1.0)
    max_bet_billion = max_bet / 100_000_000
    for i in range(6):
        y = top + plot_h * i / 5
        value = max_bet_billion * (1 - i / 5)
        draw.line((left, y, width - right, y), fill=GRID, width=2)
        draw_text(draw, (left - 16, int(y)), f"{value:.0f}亿", 18, fill=MUTED, anchor="ra")
    rtp_min, rtp_max = 90.0, 100.0
    for i in range(6):
        y = top + plot_h * i / 5
        value = rtp_max - (rtp_max - rtp_min) * i / 5
        draw_text(draw, (width - right + 16, int(y)), f"{value:.0f}%", 18, fill=MUTED, anchor="la")
    xs = [left + plot_w * (i + 0.5) / len(rows) for i in range(len(rows))]
    bar_width = min(150, plot_w / len(rows) * 0.48)
    for x, row in zip(xs, rows):
        bet_billion = row["full_bet"] / 100_000_000
        bar_h = plot_h * bet_billion / max_bet_billion if max_bet_billion else 0
        y0 = top + plot_h
        y1 = y0 - bar_h
        draw.rounded_rectangle((x - bar_width / 2, y1, x + bar_width / 2, y0), radius=10, fill="#8DB5F4", outline=BLUE, width=2)
        draw_text(draw, (int(x), int(y1 - 18)), bar_label_getter(row) if bar_label_getter else money(row["full_bet"]), 17, fill=NAVY, bold=True, anchor="ms")
        category = category_getter(row) if category_getter else f"生命周期{row['lifecycle']}"
        draw_text(draw, (int(x), height - 86), category, 16 if category_getter else 18, fill=INK, anchor="ma")

    def point(value: float | None) -> tuple[int, int] | None:
        if value is None:
            return None
        y = top + (rtp_max - value * 100) / (rtp_max - rtp_min) * plot_h
        return None if y < top - 2 or y > top + plot_h + 2 else (0, int(y))

    actual_points = []
    expected_points = []
    for x, row in zip(xs, rows):
        actual = point(row.get("actual_rtp"))
        expected = point(row.get("expected_rtp")) if show_expected else None
        actual_points.append((int(x), actual[1]) if actual else None)
        expected_points.append((int(x), expected[1]) if expected else None)

    def draw_series(points: list[tuple[int, int] | None], color: str, dashed: bool, label: str, label_offset: int) -> None:
        last = None
        for idx, pnt in enumerate(points):
            if pnt is not None and last is not None:
                if dashed:
                    segments = 10
                    for j in range(segments):
                        if j % 2 == 0:
                            x1 = last[0] + (pnt[0] - last[0]) * j / segments
                            y1 = last[1] + (pnt[1] - last[1]) * j / segments
                            x2 = last[0] + (pnt[0] - last[0]) * (j + 1) / segments
                            y2 = last[1] + (pnt[1] - last[1]) * (j + 1) / segments
                            draw.line((x1, y1, x2, y2), fill=color, width=4)
                else:
                    draw.line((last[0], last[1], pnt[0], pnt[1]), fill=color, width=5)
            if pnt is not None:
                draw.ellipse((pnt[0] - 8, pnt[1] - 8, pnt[0] + 8, pnt[1] + 8), fill=color, outline=WHITE, width=2)
                value = rows[idx].get("actual_rtp" if label == "实际RTP" else "expected_rtp")
                if value is not None:
                    draw_text(draw, (pnt[0], pnt[1] + label_offset), pct(value), 16, fill=color, bold=True, anchor="mm")
                last = pnt
            else:
                last = None
        legend_x = width - right - 290
        legend_y = 45 + (0 if label == "实际RTP" else 34)
        if dashed:
            for j in range(0, 30, 8):
                draw.line((legend_x + j, legend_y, legend_x + j + 5, legend_y), fill=color, width=4)
        else:
            draw.line((legend_x, legend_y, legend_x + 30, legend_y), fill=color, width=5)
        draw_text(draw, (legend_x + 42, legend_y), label, 18, fill=INK, anchor="lm")

    draw_series(actual_points, ORANGE, False, "实际RTP", -24)
    if show_expected:
        draw_series(expected_points, BLUE, True, "配置预期RTP", 24)
    image.save(path)


def heatmap(path: Path, title: str, rows: list[dict[str, Any]]) -> None:
    width, height = 1600, max(820, 185 + len(rows) * 25)
    image = Image.new("RGB", (width, height), WHITE)
    draw = ImageDraw.Draw(image)
    draw_text(draw, (48, 34), title, 32, bold=True)
    draw_text(draw, (48, 80), "颜色参考：以97.0%为视觉中线；红色表示RTP较低，青色表示较高；绿/红边框表示偏离平均，绝对值越大颜色越深", 18, fill=MUTED)
    left, top, cell_w, cell_h = 320, 132, 460, 25
    draw_text(draw, (left + cell_w / 2, top - 16), "生命周期3", 18, fill=MUTED, anchor="mm")
    draw_text(draw, (left + cell_w + 250, top - 16), "生命周期4", 18, fill=MUTED, anchor="mm")
    for i, row in enumerate(rows):
        y = top + i * cell_h
        gaps = [row.get("life3_gap"), row.get("life4_gap")]
        max_gap = max((abs(gap) for gap in gaps if gap is not None), default=0.0)
        dominant_gap = max((gap for gap in gaps if gap is not None), key=abs, default=None)
        label_color = (
            ("#15803d" if dominant_gap > 0 else "#b91c1c")
            if dominant_gap is not None and max_gap >= 0.01
            else INK
        )
        if max_gap >= 0.01:
            label_fill = (
                ("#dcfce7" if dominant_gap > 0 else "#fee2e2")
                if max_gap < 0.02
                else ("#86efac" if dominant_gap > 0 else "#fca5a5")
            )
            draw.rounded_rectangle((42, y + 2, left - 12, y + cell_h - 4), radius=5, fill=label_fill)
        draw_text(draw, (left - 14, y + cell_h / 2), row["game"], 16, fill=label_color, bold=max_gap >= 0.01, anchor="ra")
        for j, key in enumerate(("life3_actual_rtp", "life4_actual_rtp")):
            value = row.get(key)
            x = left + j * cell_w
            if value is None:
                fill = "#F1F5F9"
                label = "N/A"
            else:
                if value < 0.97:
                    ratio = max(0.0, min(1.0, (value - 0.90) / 0.07))
                    low = (220, 38, 38)
                    near = (254, 226, 226)
                    rgb = tuple(int(low[k] * (1 - ratio) + near[k] * ratio) for k in range(3))
                else:
                    ratio = max(0.0, min(1.0, (value - 0.97) / 0.04))
                    near = (219, 234, 254)
                    high = (13, 148, 136)
                    rgb = tuple(int(near[k] * (1 - ratio) + high[k] * ratio) for k in range(3))
                fill = f"#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"
                label = pct(value, 1)
            gap_key = "life3_gap" if key == "life3_actual_rtp" else "life4_gap"
            gap = row.get(gap_key)
            border = None
            if gap is not None and abs(gap) >= 0.01:
                if gap > 0:
                    border = "#15803d" if abs(gap) >= 0.02 else "#4ade80"
                else:
                    border = "#b91c1c" if abs(gap) >= 0.02 else "#f87171"
            draw.rectangle((x, y, x + cell_w - 8, y + cell_h - 3), fill=fill, outline=border or fill, width=3 if border else 1)
            dark_text = value is None or (value is not None and 0.945 <= value <= 0.98)
            draw_text(draw, (x + (cell_w - 8) / 2, y + (cell_h - 3) / 2), label, 15, fill=INK if dark_text else WHITE, anchor="mm")
    legend_y = height - 38
    for idx, (label, color) in enumerate((("<95%", "#dc2626"), ("97%基线", "#dbeafe"), (">99%", "#0d9488"), ("N/A", "#f1f5f9"))):
        x = 430 + idx * 170
        draw.rectangle((x, legend_y - 9, x + 24, legend_y + 9), fill=color, outline=GRID)
        draw_text(draw, (x + 34, legend_y), label, 16, fill=MUTED, anchor="lm")
    draw_text(draw, (1180, legend_y), "绿框：正偏离；红框：负偏离；深色：绝对偏离更大", 15, fill=MUTED, anchor="lm")
    image.save(path)


def header_map(header: Iterable[Any]) -> dict[str, int]:
    aliases = {
        "生命周期": "lifecycle",
        "游戏类型": "game",
        "完全预期盈利": "expected_profit",
        "完全实际盈利": "actual_profit",
        "完全下注额": "full_bet",
        "完全真实回报比": "source_actual_rtp",
        "预期回报比": "source_expected_rtp",
        "完全预期回报比": "source_expected_rtp",
        "基础下注额": "base_bet",
        "基础实际盈利": "base_actual_profit",
        "基础真实回报比": "source_base_rtp",
    }
    result = {}
    for i, item in enumerate(header):
        key = aliases.get(clean(item))
        if key:
            result[key] = i
    required = {"lifecycle", "game", "expected_profit", "actual_profit", "full_bet"}
    missing = required - result.keys()
    if missing:
        raise ValueError(f"缺少明细字段: {sorted(missing)}")
    return result


def read_detail(path: Path, business_date: date) -> tuple[list[dict[str, Any]], list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        fields = header_map(header)
        records = []
        for row in rows:
            if not row or not any(v is not None for v in row):
                continue
            lifecycle = num(row[fields["lifecycle"]])
            game = str(row[fields["game"]] or "").strip()
            if lifecycle is None or not game:
                continue
            def get(field: str) -> float | None:
                return num(row[fields[field]]) if field in fields else None
            records.append({
                "business_date": business_date.isoformat(),
                "game": game,
                "lifecycle": int(lifecycle),
                "full_bet": get("full_bet"),
                "actual_profit": get("actual_profit"),
                "expected_profit": get("expected_profit"),
                "source_expected_rtp": get("source_expected_rtp"),
                "source_actual_rtp": get("source_actual_rtp"),
            })
        return records, list(header)
    finally:
        wb.close()


def canonical_dir(day: date) -> Path:
    if day <= date(2026, 8, 27):
        return ROOT / "data/raw/lifecycle_joint/2026-08-28-30d" / day.isoformat()
    if day <= date(2026, 8, 30):
        return ROOT / "data/raw/lifecycle_joint/2026-08-31" / day.isoformat()
    return ROOT / "data/raw/lifecycle_joint/2026-09-02" / day.isoformat()


def aggregate(rows: list[dict[str, Any]]) -> dict[str, Any]:
    bet = sum((r["full_bet"] or 0.0) for r in rows)
    actual = sum((r["actual_profit"] or 0.0) for r in rows)
    expected_rows = [r for r in rows if (r["full_bet"] or 0) > 0 and valid_expected(r.get("source_expected_rtp"))]
    expected_bet = sum(r["full_bet"] or 0.0 for r in expected_rows)
    expected_rtp = sum((r["full_bet"] or 0.0) * r["source_expected_rtp"] for r in expected_rows) / expected_bet if expected_bet else None
    actual_on_coverage = 1 - sum((r["actual_profit"] or 0.0) for r in expected_rows) / expected_bet if expected_bet else None
    actual_rtp = 1 - actual / bet if bet else None
    return {
        "full_bet": bet,
        "actual_profit": actual,
        "actual_rtp": actual_rtp,
        "expected_rtp": expected_rtp,
        "expected_bet": expected_bet,
        "config_coverage": expected_bet / bet if expected_bet else None,
        "actual_rtp_on_coverage": actual_on_coverage,
        "rtp_gap": actual_on_coverage - expected_rtp if actual_on_coverage is not None and expected_rtp is not None else None,
        "row_count": len(rows),
        "active_days": len({r["business_date"] for r in rows if (r["full_bet"] or 0) > 0}),
    }


def period_rows(day_rows: list[dict[str, Any]], start: date, end: date) -> list[dict[str, Any]]:
    return [r for r in day_rows if start.isoformat() <= r["business_date"] <= end.isoformat()]


def parse_tc() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    source = json.loads(TC_FILE.read_text(encoding="utf-8"))
    result = []
    for row in source.get("rows", []):
        d = as_date(row["business_date"])
        if d is None:
            continue
        recharge = num(row.get("success_recharge_amount")) or 0.0
        withdraw = num(row.get("success_withdraw_amount")) or 0.0
        result.append({
            "business_date": d.isoformat(),
            "success_recharge_amount": recharge,
            "success_withdraw_amount": withdraw,
            "tc_rate_display": num(row.get("tc_rate")),
            "tc_rate": withdraw / recharge if recharge else None,
            "data_state": "actual_aggregate",
        })
    return sorted(result, key=lambda r: r["business_date"]), source


def compare_tc_overlap(tc_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not EXISTING_TC.exists():
        return [{"status": "existing_snapshot_missing", "path": str(EXISTING_TC.relative_to(ROOT))}]
    old = {}
    with EXISTING_TC.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            old[row["date"]] = row
    mismatches = []
    for row in tc_rows:
        d = row["business_date"]
        if d not in old:
            continue
        old_recharge = num(old[d].get("success_recharge_amount"))
        old_withdraw = num(old[d].get("success_withdraw_amount"))
        if abs((old_recharge or 0) - row["success_recharge_amount"]) > 0.01 or abs((old_withdraw or 0) - row["success_withdraw_amount"]) > 0.01:
            mismatches.append({"business_date": d, "status": "source_mismatch"})
    return mismatches


def discover_duplicate_checks(day: date) -> dict[str, Any]:
    candidates = []
    for receipt in (ROOT / "data/raw/lifecycle_joint").glob(f"*/{day.isoformat()}/query-receipt.json"):
        root = receipt.parent
        detail_path = root / "detail.xlsx"
        if not detail_path.exists():
            continue
        try:
            rows, _ = read_detail(detail_path, day)
            signature = hashlib.sha256(json.dumps(sorted([
                [r["game"], r["lifecycle"], r["full_bet"], r["actual_profit"], r["source_expected_rtp"]]
                for r in rows
            ]), ensure_ascii=False, separators=(",", ":")).encode("utf-8")).hexdigest()
            candidates.append({"root": str(root.relative_to(ROOT)), "signature": signature, "receipt_hash": sha256(receipt)})
        except Exception as exc:
            candidates.append({"root": str(root.relative_to(ROOT)), "status": "read_error", "error": str(exc)})
    sigs = {c["signature"] for c in candidates if "signature" in c}
    return {
        "date": day.isoformat(),
        "candidate_count": len(candidates),
        "status": "consistent" if len(sigs) <= 1 else "conflict_review_required",
        "candidates": candidates,
    }


def new_user_context() -> dict[str, Any]:
    if not NEW_USER_FILE.exists():
        return {"status": "missing", "rows": [], "accepted_date_start": None, "accepted_date_end": None}
    source = json.loads(NEW_USER_FILE.read_text(encoding="utf-8"))
    sheet_rows = []
    for sheet_name, sheet in source.get("sheets", {}).items():
        headers = sheet.get("headers", [])
        index = {str(h): i for i, h in enumerate(headers)}
        rows = []
        for raw in sheet.get("rows", []):
            d = as_date(raw[index.get("日期", 0)])
            if d is None or not (MONTH_START <= d <= MONTH_END):
                continue
            def value(key: str) -> float | None:
                i = index.get(key)
                return num(raw[i]) if i is not None and i < len(raw) else None
            new_users = value("新增人数") or 0.0
            row = {
                "date": d.isoformat(),
                "new_users": new_users,
                "new_payers": value("新增付费人数") or 0.0,
                "first_payers": value("首充付费人数") or 0.0,
                "d1_retention": value("次留"),
                "d3_retention": value("3日留"),
                "d7_retention": value("7日留"),
            }
            rows.append(row)
        mapping = sheet.get("mapping", {})
        def weighted(key: str) -> float | None:
            denominator = sum(r["new_users"] for r in rows)
            if not denominator:
                return None
            vals = [(r["new_users"], r[key]) for r in rows if r[key] is not None]
            if not vals:
                return None
            return sum(weight * value for weight, value in vals) / sum(weight for weight, _ in vals)
        total_users = sum(r["new_users"] for r in rows)
        total_payers = sum(r["new_payers"] for r in rows)
        sheet_rows.append({
            "segment": sheet_name,
            "package_channel": mapping.get("package_channel"),
            "media": mapping.get("attribution_media"),
            "accepted_days": len(rows),
            "new_users": total_users,
            "new_payers": total_payers,
            "first_payers": sum(r["first_payers"] for r in rows),
            "new_pay_rate": total_payers / total_users if total_users else None,
            "d1_retention": weighted("d1_retention"),
            "d3_retention": weighted("d3_retention"),
            "d7_retention": weighted("d7_retention"),
            "data_state": "actual_matured_context",
        })
    weighted_average = {}
    for key in ("new_pay_rate", "d1_retention", "d3_retention", "d7_retention"):
        valid_rows = [r for r in sheet_rows if r.get(key) is not None and r.get("new_users", 0) > 0]
        denominator = sum(r["new_users"] for r in valid_rows)
        weighted_average[key] = sum(r["new_users"] * r[key] for r in valid_rows) / denominator if denominator else None
    return {
        "status": "partial_maturity_window",
        "accepted_date_start": "2026-08-06",
        "accepted_date_end": "2026-08-30",
        "excluded_dates": ["2026-08-01", "2026-08-02", "2026-08-03", "2026-08-04", "2026-08-05", "2026-08-31"],
        "rows": sorted(sheet_rows, key=lambda r: r["segment"]),
        "weighted_average": weighted_average,
        "source_file": str(NEW_USER_FILE.relative_to(ROOT)),
    }


def column_has_information(values: list[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, (int, float)):
            if math.isfinite(float(value)) and abs(float(value)) > 1e-12:
                return True
            continue
        text = str(value).strip()
        if not text:
            continue
        if text in {"0", "0.0", "0.00", "0.00%", "0%"}:
            continue
        return True
    return False


def visible_fields(fields: list[tuple[str, str, Callable[[dict[str, Any]], Any]]], rows: list[dict[str, Any]], removed: list[dict[str, Any]], scope: str) -> list[tuple[str, str, Callable[[dict[str, Any]], Any]]]:
    result = []
    for key, label, formatter in fields:
        values = [formatter(row) for row in rows]
        if column_has_information(values):
            result.append((key, label, formatter))
        else:
            removed.append({"scope": scope, "field": key, "label": label, "reason": "all_zero_or_blank"})
    return result


def cell(value: Any) -> str:
    if value is None or value == "":
        return "N/A"
    return html.escape(str(value))


def table_html(fields: list[tuple[str, str, Callable[[dict[str, Any]], Any]]], rows: list[dict[str, Any]], removed: list[dict[str, Any]], scope: str, dense: bool = False, cell_class_for: Callable[[dict[str, Any], str, Any], str] | None = None, row_class_for: Callable[[dict[str, Any]], str] | None = None) -> str:
    active = visible_fields(fields, rows, removed, scope)
    head = "".join(f"<th>{html.escape(label)}</th>" for _, label, _ in active)
    body = []
    for row in rows:
        cells = []
        for key, _, formatter in active:
            value = formatter(row)
            class_name = cell_class_for(row, key, value) if cell_class_for else ""
            class_attr = f' class="{class_name}"' if class_name else ""
            cells.append(f"<td{class_attr}>{cell(value)}</td>")
        row_class = row_class_for(row) if row_class_for else ""
        row_attr = f' class="{row_class}"' if row_class else ""
        body.append(f"<tr{row_attr}>" + "".join(cells) + "</tr>")
    cls = " dense" if dense else ""
    return f'<div class="table-wrap{cls}"><table><thead><tr>{head}</tr></thead><tbody>{"".join(body)}</tbody></table></div>'


def signed_gap_class(value: float | None, prefix: str) -> str:
    """Return a signed, magnitude-aware class for percentage-point gaps."""
    if value is None or abs(value) < 0.0005:
        return ""
    direction = "positive" if value > 0 else "negative"
    magnitude = abs(value)
    level = (
        "extreme" if magnitude >= 0.05
        else "critical" if magnitude >= 0.02
        else "highlight" if magnitude >= 0.01
        else "soft" if magnitude >= 0.005
        else "faint"
    )
    return f"{prefix}-{direction}-{level}"


def deviation_class(value: float | None) -> str:
    return signed_gap_class(value, "deviation")


def rtp_deviation_class(row: dict[str, Any], key: str, _: Any) -> str:
    if key == "profit_bet_share_gap":
        return deviation_class(row.get("profit_bet_share_gap"))
    if key in {"rtp_gap", "game"}:
        return deviation_class(row.get("rtp_gap"))
    return ""


def average_deviation_class(row: dict[str, Any], key: str, _: Any) -> str:
    if key not in {"average_gap", "game"}:
        return ""
    return deviation_class(row.get("average_gap"))


def detail_average_row_class(row: dict[str, Any]) -> str:
    gap = row.get("average_gap")
    if gap is None:
        return ""
    if abs(gap) >= 0.02:
        return "detail-row-critical"
    if abs(gap) >= 0.01:
        return "detail-row-highlight"
    return ""


def tc_baseline_class(row: dict[str, Any], key: str, _: Any) -> str:
    return signed_gap_class(row.get("baseline_delta"), "tc")


def context_deviation_class(row: dict[str, Any], key: str, _: Any) -> str:
    if key not in {"new_pay_rate", "d1_retention", "d3_retention", "d7_retention"}:
        return ""
    average = row.get("_weighted_average", {}).get(key)
    value = row.get(key)
    if average is None or value is None:
        return ""
    delta = value - average
    if delta >= 0.05:
        return "context-high"
    if delta <= -0.05:
        return "context-low"
    return ""


def context_row_class(row: dict[str, Any]) -> str:
    average = row.get("_weighted_average", {})
    deltas = []
    for key in ("new_pay_rate", "d1_retention", "d3_retention", "d7_retention"):
        value = row.get(key)
        baseline = average.get(key)
        if value is not None and baseline is not None:
            deltas.append(value - baseline)
    if any(delta <= -0.05 for delta in deltas):
        return "context-row-low"
    if any(delta >= 0.05 for delta in deltas):
        return "context-row-high"
    return ""


def data_uri(path: Path) -> str:
    return "data:image/png;base64," + base64.b64encode(path.read_bytes()).decode("ascii")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    ASSETS.mkdir(parents=True, exist_ok=True)
    days = [MONTH_START.fromordinal(MONTH_START.toordinal() + i) for i in range((MONTH_END - MONTH_START).days + 1)]
    day_records: list[dict[str, Any]] = []
    source_days = []
    all_headers = set()
    for day in days:
        root = canonical_dir(day)
        detail_path = root / "detail.xlsx"
        receipt_path = root / "query-receipt.json"
        records, headers = read_detail(detail_path, day)
        all_headers.update(clean(h) for h in headers if h is not None)
        receipt = json.loads(receipt_path.read_text(encoding="utf-8")) if receipt_path.exists() else {}
        expected = receipt.get("expected_rows", {})
        returned = receipt.get("returned_rows", {})
        source_days.append({
            "date": day.isoformat(),
            "canonical_root": str(root.relative_to(ROOT)),
            "files": {name: sha256(root / name) for name in ("summary.xlsx", "detail.xlsx", "game.xlsx", "active.xlsx", "query-receipt.json") if (root / name).exists()},
            "detail_rows_parsed": len(records),
            "receipt_status": receipt.get("status", "not_observed"),
            "returned_rows": returned,
            "expected_rows": expected,
            "row_count_match": not expected or returned == expected,
        })
        day_records.extend(records)
    duplicate_checks = [discover_duplicate_checks(day) for day in days]
    tc_rows, tc_source = parse_tc()
    tc_overlap = compare_tc_overlap(tc_rows)
    tc_dates = {r["business_date"] for r in tc_rows}
    expected_dates = {d.isoformat() for d in days}
    tc_missing_dates = sorted(expected_dates - tc_dates)
    new_user = new_user_context()

    # Monthly and weekly lifecycle aggregates.
    target = [r for r in day_records if r["lifecycle"] in LIFECYCLES]
    games = sorted({r["game"] for r in target})
    monthly_detail = []
    for game in games:
        for life in LIFECYCLES:
            rows = [r for r in target if r["game"] == game and r["lifecycle"] == life]
            monthly_detail.append({"game": game, "lifecycle": life, **aggregate(rows), "data_state": "actual_aggregate"})
    display_detail_all = [r for r in monthly_detail if r["lifecycle"] in (3, 4)]
    hidden_detail_rows = [r for r in display_detail_all if (r["full_bet"] or 0) <= 0]
    display_detail = [r for r in display_detail_all if (r["full_bet"] or 0) > 0]
    lifecycle_summary = [{"lifecycle": life, **aggregate([r for r in target if r["lifecycle"] == life]), "data_state": "actual_aggregate"} for life in LIFECYCLES]
    lifecycle_average = {r["lifecycle"]: r["actual_rtp"] for r in lifecycle_summary}
    for row in display_detail:
        row["lifecycle_average_rtp"] = lifecycle_average.get(row["lifecycle"])
        row["average_gap"] = row["actual_rtp"] - row["lifecycle_average_rtp"] if row["actual_rtp"] is not None and row["lifecycle_average_rtp"] is not None else None
        row["average_status"] = "显著偏离" if row["average_gap"] is not None and abs(row["average_gap"]) >= 0.02 else ("轻度偏离" if row["average_gap"] is not None and abs(row["average_gap"]) >= 0.01 else "接近平均")
    game_summary = [{"game": game, **aggregate([r for r in target if r["game"] == game]), "data_state": "actual_aggregate"} for game in games]
    game_summary.sort(key=lambda r: r["full_bet"], reverse=True)
    monthly = aggregate(target)
    monthly["period_key"] = "august_2026"
    monthly["period_start"] = MONTH_START.isoformat()
    monthly["period_end"] = MONTH_END.isoformat()
    monthly["expected_days"] = 31
    monthly["observed_days"] = len({r["business_date"] for r in day_records})
    monthly["observed_game_count"] = len(games)
    profit_total = monthly["actual_profit"] or 0.0
    for game_row in game_summary:
        game_row["bet_share"] = game_row["full_bet"] / monthly["full_bet"] if monthly["full_bet"] else None
        game_row["profit_share"] = game_row["actual_profit"] / profit_total if profit_total else None
        game_row["profit_bet_share_gap"] = (
            game_row["profit_share"] - game_row["bet_share"]
            if game_row["profit_share"] is not None and game_row["bet_share"] is not None
            else None
        )
    tada_row = next((row for row in game_summary if row["game"].lower() == "tada"), None)
    pp_row = next((row for row in game_summary if row["game"].lower() == "pp"), None)
    third_party_tada_pp = {
        "data_state": "actual_aggregate",
        "business_definition": "第三方 TaDa（源数据名 Tada）与 PP 的生命周期1—4月度下注额合计占比",
        "tada_bet": tada_row["full_bet"] if tada_row else None,
        "tada_share": tada_row["bet_share"] if tada_row else None,
        "pp_bet": pp_row["full_bet"] if pp_row else None,
        "pp_share": pp_row["bet_share"] if pp_row else None,
        "tada_profit": tada_row["actual_profit"] if tada_row else None,
        "tada_profit_share": tada_row["profit_share"] if tada_row else None,
        "pp_profit": pp_row["actual_profit"] if pp_row else None,
        "pp_profit_share": pp_row["profit_share"] if pp_row else None,
        "combined_bet": (tada_row["full_bet"] if tada_row else 0.0) + (pp_row["full_bet"] if pp_row else 0.0),
        "combined_share": ((tada_row["full_bet"] if tada_row else 0.0) + (pp_row["full_bet"] if pp_row else 0.0)) / monthly["full_bet"] if monthly["full_bet"] else None,
        "combined_profit": (tada_row["actual_profit"] if tada_row else 0.0) + (pp_row["actual_profit"] if pp_row else 0.0),
        "combined_profit_share": ((tada_row["actual_profit"] if tada_row else 0.0) + (pp_row["actual_profit"] if pp_row else 0.0)) / profit_total if profit_total else None,
        "scope": "Tada + PP; lifecycle 1—4; August total full bet",
    }
    monthly["third_party_tada_pp"] = third_party_tada_pp
    profit_rank = sorted(game_summary, key=lambda r: r["actual_profit"], reverse=True)
    bet_positive = [r for r in game_summary if (r["full_bet"] or 0) > 0]
    distribution_analysis = {
        "data_state": "actual_aggregate",
        "scope": "2026-08-01—2026-08-31; lifecycle 1—4; game aggregate",
        "total_bet": monthly["full_bet"],
        "total_profit": monthly["actual_profit"],
        "top5_bet_share": sum(r["bet_share"] for r in game_summary[:5]),
        "top8_bet_share": sum(r["bet_share"] for r in game_summary[:8]),
        "top5_profit_share": sum(r["profit_share"] for r in profit_rank[:5]),
        "top_bet_games": [{"game": r["game"], "bet_share": r["bet_share"], "actual_rtp": r["actual_rtp"]} for r in game_summary[:5]],
        "top_profit_games": [{"game": r["game"], "profit_share": r["profit_share"], "bet_share": r["bet_share"], "actual_rtp": r["actual_rtp"]} for r in profit_rank[:5]],
        "profit_bet_share_gap_rank": [{"game": r["game"], "gap": r["profit_bet_share_gap"], "bet_share": r["bet_share"], "profit_share": r["profit_share"]} for r in sorted(bet_positive, key=lambda r: r["profit_bet_share_gap"], reverse=True)[:5]],
        "recommendation_boundary": "仅依据下注额、实际盈利和实际RTP分布；没有用户留存、体验、包体和版本拆分时，不单独据此决定推荐曝光。",
    }
    package_split_plan = {
        "status": "not_available_in_current_lifecycle_snapshot",
        "current_report_grain": "business_date × game × lifecycle",
        "required_dimensions": ["platform", "package_name", "app_version_or_web_version", "release_id", "game", "lifecycle"],
        "required_platform_values": ["Android", "H5"],
        "recommended_fields": [
            "platform", "package_name", "app_version", "web_version", "release_id", "game", "lifecycle",
            "full_bet", "bet_share_within_package", "bet_share_all_product", "actual_profit", "profit_share_within_package",
            "actual_rtp", "expected_rtp", "data_state",
        ],
        "validation": [
            "平台×包体×版本分层合计与全产品总额对账",
            "包体内下注额占比合计为100%",
            "缺少platform或package_name时显示待补字段，不归入Android或H5",
        ],
    }
    monthly["package_split_status"] = package_split_plan["status"]
    monthly["tc"] = None
    if not tc_missing_dates:
        recharge = sum(r["success_recharge_amount"] for r in tc_rows)
        withdraw = sum(r["success_withdraw_amount"] for r in tc_rows)
        monthly["tc_recharge"] = recharge
        monthly["tc_withdraw"] = withdraw
        monthly["tc"] = withdraw / recharge if recharge else None
        monthly["tc_data_state"] = "actual_aggregate_verified_overlap" if not tc_overlap else "actual_aggregate_source_mismatch"
    else:
        monthly["tc_data_state"] = "missing_dates"
    # Use the month-level weighted TC as the daily baseline.  A difference of
    # at least 2 percentage points is highlighted for investigation; it is not
    # a causal conclusion.
    for tc_row in tc_rows:
        if monthly.get("tc") is not None and tc_row.get("tc_rate") is not None:
            tc_row["baseline_delta"] = tc_row["tc_rate"] - monthly["tc"]
            tc_row["baseline_status"] = "重点偏离" if abs(tc_row["baseline_delta"]) >= 0.02 else "正常波动"
        else:
            tc_row["baseline_delta"] = None
            tc_row["baseline_status"] = "待补数据"
    weekly = []
    for period in PERIODS:
        selected = period_rows(target, period["start"], period["end"])
        row = {
            "period_key": period["key"],
            "label": period["label"],
            "period_start": period["start"].isoformat(),
            "period_end": period["end"].isoformat(),
            **aggregate(selected),
        }
        row["expected_days"] = (period["end"] - period["start"]).days + 1
        row["observed_days"] = len({r["business_date"] for r in selected})
        row["observed_game_count"] = len({r["game"] for r in selected})
        row["daily_average_bet"] = row["full_bet"] / row["observed_days"] if row["observed_days"] else None
        tc_selected = [r for r in tc_rows if period["start"].isoformat() <= r["business_date"] <= period["end"].isoformat()]
        row["tc_observed_days"] = len(tc_selected)
        row["tc_missing_dates"] = sorted({d.isoformat() for d in days if period["start"] <= d <= period["end"]} - {r["business_date"] for r in tc_selected})
        recharge = sum(r["success_recharge_amount"] for r in tc_selected)
        withdraw = sum(r["success_withdraw_amount"] for r in tc_selected)
        row["tc_recharge"] = recharge
        row["tc_withdraw"] = withdraw
        row["tc"] = withdraw / recharge if recharge and not row["tc_missing_dates"] else None
        row["tc_data_state"] = "actual_aggregate_verified_overlap" if row["tc"] is not None and not tc_overlap else ("missing_dates" if row["tc_missing_dates"] else "actual_aggregate_source_mismatch")
        weekly.append(row)

    # Per-game lifecycle 3/4 heatmap inputs.
    heat_rows = []
    for game in games:
        by_life = {r["lifecycle"]: r for r in monthly_detail if r["game"] == game}
        heat_rows.append({
            "game": game,
            "life3_actual_rtp": by_life[3]["actual_rtp"],
            "life4_actual_rtp": by_life[4]["actual_rtp"],
            "life3_gap": by_life[3].get("average_gap"),
            "life4_gap": by_life[4].get("average_gap"),
        })

    removed_columns: list[dict[str, Any]] = []
    # Charts are generated from actual values.  No simulated data is used.
    labels = [p["label"].replace("第", "").replace("（", " ").split(" ")[0] for p in weekly]
    line_chart(ASSETS / "01_四周TC与生命周期RTP.png", "8月四周：TC比与生命周期1—4实际RTP", labels, [
        ("TC比", [r["tc"] * 100 if r["tc"] is not None else None for r in weekly], BLUE),
        ("生命周期1—4实际RTP", [r["actual_rtp"] * 100 if r["actual_rtp"] is not None else None for r in weekly], ORANGE),
    ], 70, 100, "%", annotate=True)
    life_labels = [f"生命周期{r['lifecycle']}" for r in lifecycle_summary]
    lifecycle_combo_chart(ASSETS / "02_生命周期RTP与下注额.png", lifecycle_summary)
    top_games = game_summary[:12]
    lifecycle_combo_chart(ASSETS / "03_主要游戏下注规模.png", top_games, "8月主要游戏：下注额（柱）与实际RTP（折线）", lambda r: r["game"], show_expected=False, bar_label_getter=lambda r: f"{money(r['full_bet'])}｜{pct(r.get('bet_share'))}")
    heatmap(ASSETS / "04_生命周期3与4RTP热力图.png", "8月游戏 × 生命周期3/4实际RTP", heat_rows)
    line_chart(ASSETS / "05_每日TC趋势.png", "8月每日 TC 比（成功提现 ÷ 成功充值）", [r["business_date"][5:] for r in tc_rows], [("每日TC比", [r["tc_rate"] * 100 if r["tc_rate"] is not None else None for r in tc_rows], TEAL)], 65, 90, "%", annotate=True, annotation_digits=1, x_label_step=3)
    profit_top_games = profit_rank[:12]
    bar_chart(
        ASSETS / "06_主要游戏盈利贡献.png",
        "8月主要游戏：实际盈利贡献分布",
        [r["game"] for r in profit_top_games],
        [r["actual_profit"] for r in profit_top_games],
        ["#2f855a" if r["profit_bet_share_gap"] >= 0 else "#c2410c" for r in profit_top_games],
        [f"{money(r['actual_profit'])}｜{pct(r.get('profit_share'))}" for r in profit_top_games],
        "按实际盈利排序；标签为实际盈利｜月度盈利占比；颜色仅表示分布，不代表推荐结论",
    )

    # Display field definitions.  The raw snapshots retain everything; this only controls presentation columns.
    fmt_money = lambda key: lambda r: money(r.get(key))
    fmt_pct = lambda key: lambda r: pct(r.get(key))
    fmt_pp = lambda key: lambda r: pp(r.get(key))
    weekly_fields = [
        ("label", "周期", lambda r: r["label"]),
        ("observed_days", "有效日期", lambda r: f"{r['observed_days']}/{r['expected_days']}"),
        ("observed_game_count", "游戏数", lambda r: r["observed_game_count"]),
        ("full_bet", "下注额", fmt_money("full_bet")),
        ("daily_average_bet", "日均下注额", fmt_money("daily_average_bet")),
        ("actual_rtp", "实际RTP", fmt_pct("actual_rtp")),
        ("tc", "TC比", fmt_pct("tc")),
        ("config_coverage", "配置覆盖率", fmt_pct("config_coverage")),
        ("data_state", "数据状态", lambda r: "实际聚合｜已核对" if r["tc_data_state"] == "actual_aggregate_verified_overlap" else r["tc_data_state"]),
    ]
    life_fields = [
        ("lifecycle", "生命周期", lambda r: f"生命周期{r['lifecycle']}"),
        ("full_bet", "完全下注额", fmt_money("full_bet")),
        ("actual_profit", "完全实际盈利", fmt_money("actual_profit")),
        ("actual_rtp", "实际RTP", fmt_pct("actual_rtp")),
        ("expected_rtp", "配置预期RTP", fmt_pct("expected_rtp")),
        ("config_coverage", "配置覆盖率", fmt_pct("config_coverage")),
        ("rtp_gap", "实际-预期", fmt_pp("rtp_gap")),
    ]
    game_fields = [
        ("game", "游戏", lambda r: r["game"]),
        ("full_bet", "完全下注额", fmt_money("full_bet")),
        ("bet_share", "下注额占比", fmt_pct("bet_share")),
        ("actual_profit", "实际盈利", fmt_money("actual_profit")),
        ("profit_share", "盈利占比", fmt_pct("profit_share")),
        ("profit_bet_share_gap", "盈利占比−下注额占比", fmt_pp("profit_bet_share_gap")),
        ("actual_rtp", "实际RTP", fmt_pct("actual_rtp")),
        ("expected_rtp", "配置预期RTP", fmt_pct("expected_rtp")),
        ("config_coverage", "配置覆盖率", fmt_pct("config_coverage")),
        ("rtp_gap", "实际-预期", fmt_pp("rtp_gap")),
        ("active_days", "有效日期", lambda r: r["active_days"]),
    ]
    detail_fields = [
        ("game", "游戏", lambda r: r["game"]),
        ("lifecycle", "生命周期", lambda r: f"生命周期{r['lifecycle']}"),
        ("full_bet", "下注额", fmt_money("full_bet")),
        ("actual_rtp", "实际RTP", fmt_pct("actual_rtp")),
        ("lifecycle_average_rtp", "该生命周期平均RTP", fmt_pct("lifecycle_average_rtp")),
        ("average_gap", "偏离平均", fmt_pp("average_gap")),
        ("average_status", "偏离状态", lambda r: r["average_status"]),
        ("expected_rtp", "配置预期RTP", fmt_pct("expected_rtp")),
        ("rtp_gap", "实际-预期", fmt_pp("rtp_gap")),
    ]
    tc_fields = [
        ("business_date", "业务日", lambda r: r["business_date"]),
        ("success_recharge_amount", "成功充值金额", fmt_money("success_recharge_amount")),
        ("success_withdraw_amount", "成功提现金额", fmt_money("success_withdraw_amount")),
        ("tc_rate", "TC比", fmt_pct("tc_rate")),
        ("baseline_delta", "偏离月度基线", fmt_pp("baseline_delta")),
        ("baseline_status", "偏离状态", lambda r: r["baseline_status"]),
    ]
    new_user_fields = [
        ("segment", "包体/渠道", lambda r: r["segment"]),
        ("accepted_days", "成熟日期数", lambda r: r["accepted_days"]),
        ("new_users", "新增人数", lambda r: f"{r['new_users']:,.0f}"),
        ("new_payers", "新增付费人数", lambda r: f"{r['new_payers']:,.0f}"),
        ("new_pay_rate", "新增付费率", fmt_pct("new_pay_rate")),
        ("d1_retention", "次日留存", fmt_pct("d1_retention")),
        ("d3_retention", "3日留存", fmt_pct("d3_retention")),
        ("d7_retention", "7日留存", fmt_pct("d7_retention")),
    ]

    weekly_table = table_html(weekly_fields, weekly, removed_columns, "weekly_overview")
    life_table = table_html(life_fields, lifecycle_summary, removed_columns, "lifecycle_summary")
    game_table = table_html(game_fields, game_summary[:15], removed_columns, "game_summary", cell_class_for=rtp_deviation_class)
    detail_table = table_html(detail_fields, sorted(display_detail, key=lambda r: (r["game"].lower(), r["lifecycle"])), removed_columns, "lifecycle_3_4_detail", dense=True, cell_class_for=lambda row, key, value: average_deviation_class(row, key, value) or rtp_deviation_class(row, key, value), row_class_for=detail_average_row_class)
    tc_table = table_html(tc_fields, tc_rows, removed_columns, "tc_daily", dense=True, cell_class_for=tc_baseline_class)
    new_user_display_rows = [{**row, "_weighted_average": new_user.get("weighted_average", {})} for row in new_user.get("rows", [])]
    new_user_table = table_html(new_user_fields, new_user_display_rows, removed_columns, "new_user_context", dense=True, cell_class_for=context_deviation_class, row_class_for=context_row_class)
    context_average = new_user.get("weighted_average", {})
    context_average_note = (
        f"<div class=\"definition\"><strong>总体加权平均：</strong>新增付费率 {pct(context_average.get('new_pay_rate'))}，次日留存 {pct(context_average.get('d1_retention'))}，"
        f"3日留存 {pct(context_average.get('d3_retention'))}，7日留存 {pct(context_average.get('d7_retention'))}。加权分母为各包体/渠道的新增人数。相对平均值偏离至少5个百分点的包体/渠道整行突出显示，单项指标同步保留更强颜色。</div>"
    )
    if new_user.get("rows"):
        highest_context = max(new_user["rows"], key=lambda r: r.get("new_pay_rate") or -1)
        lowest_context = min(new_user["rows"], key=lambda r: r.get("new_pay_rate") if r.get("new_pay_rate") is not None else 999)
        context_summary = (
            f"<p><strong>简要分析：</strong>{html.escape(highest_context['segment'])}的新增付费率和留存指标整体高于加权平均，"
            f"{html.escape(lowest_context['segment'])}整体低于加权平均；高亮项表示相对平均值偏离至少5个百分点。"
            "这说明包体/渠道对应的用户结构或投放质量存在差异，仍需结合归因链路、样本成熟度和上报完整性继续核查，不能仅凭本表判定因果。</p>"
        )
    else:
        context_summary = "<p>当前没有可用于比较的成熟包体/渠道数据。</p>"
    context_md_note = (
        f"总体加权平均（按新增人数加权）：新增付费率 **{pct(context_average.get('new_pay_rate'))}**，次日留存 **{pct(context_average.get('d1_retention'))}**，"
        f"3日留存 **{pct(context_average.get('d3_retention'))}**，7日留存 **{pct(context_average.get('d7_retention'))}**。\\n\\n"
        f"简要分析：{highest_context['segment'] if new_user.get('rows') else '暂无'}在新增付费率和留存上整体高于平均，"
        f"{lowest_context['segment'] if new_user.get('rows') else '暂无'}整体低于平均；高亮项表示相对平均值偏离至少5个百分点。该差异需要结合归因、用户结构和上报完整性继续核查，不能单凭本表判定因果。"
    )
    game_share_note = (
        f"<div class=\"definition\"><strong>下注额占比与第三方汇总：</strong>单款游戏下注额占比 = 该游戏生命周期1—4完全下注额 ÷ 月度生命周期1—4完全下注总额。"
        f"TaDa（源数据名 Tada）下注额为 {money(third_party_tada_pp['tada_bet'])}，占 {pct(third_party_tada_pp['tada_share'])}；"
        f"PP下注额为 {money(third_party_tada_pp['pp_bet'])}，占 {pct(third_party_tada_pp['pp_share'])}；"
        f"两者合计 {money(third_party_tada_pp['combined_bet'])}，占月度总下注额 {pct(third_party_tada_pp['combined_share'])}。"
        "该占比只表示下注规模，不代表RTP、付费率或盈利贡献。</div>"
    )
    game_share_md_note = (
        f"下注额占比与第三方汇总：单款游戏下注额占比 = 该游戏生命周期1—4完全下注额 ÷ 月度生命周期1—4完全下注总额。"
        f"TaDa（源数据名 Tada）下注额为 **{money(third_party_tada_pp['tada_bet'])}**，占 **{pct(third_party_tada_pp['tada_share'])}**；"
        f"PP下注额为 **{money(third_party_tada_pp['pp_bet'])}**，占 **{pct(third_party_tada_pp['pp_share'])}**；"
        f"两者合计 **{money(third_party_tada_pp['combined_bet'])}**，占月度总下注额 **{pct(third_party_tada_pp['combined_share'])}**。"
        "该占比只表示下注规模，不代表RTP、付费率或盈利贡献。"
    )
    deviation_color_note = (
        "<div class=\"definition\"><strong>偏离颜色说明：</strong>绿色系表示正偏离（实际RTP高于对应基线），红色系表示负偏离（实际RTP低于对应基线）；非零偏离从浅色开始，绝对偏离达到约0.5、1、2、5个百分点时逐级加深。颜色只用于定位偏离，不直接代表经营好坏或因果。</div>"
    )
    deviation_color_md_note = (
        "偏离颜色说明：绿色系表示正偏离（实际RTP高于对应基线），红色系表示负偏离（实际RTP低于对应基线）；非零偏离从浅色开始，绝对偏离达到约0.5、1、2、5个百分点时逐级加深。颜色只用于定位偏离，不直接代表经营好坏或因果。"
    )
    tc_color_note = (
        "<div class=\"definition\"><strong>TC偏离颜色说明：</strong>绿色表示当日TC比高于8月加权TC基线，红色表示低于基线；非零偏离从浅色开始，绝对偏离越大颜色越深。颜色只表示相对基线的位置，仅用于优先核查，不代表归因结论。</div>"
    )
    tc_color_md_note = (
        "TC偏离颜色说明：绿色表示当日TC比高于8月加权TC基线，红色表示低于基线；非零偏离从浅色开始，绝对偏离越大颜色越深。颜色只表示相对基线的位置，仅用于优先核查，不代表归因结论。"
    )
    game_index = {row["game"]: row for row in game_summary}
    whot_row = game_index.get("Whot")
    omg_row = game_index.get("OMG")
    easywin_row = game_index.get("EasyWin")
    distribution_analysis_note = (
        f"<div class=\"callout recommendation\"><strong>分布汇总与推荐策略：</strong>下注额前5款游戏合计占月度下注额 {pct(distribution_analysis['top5_bet_share'])}，前8款合计占 {pct(distribution_analysis['top8_bet_share'])}；盈利贡献前5款合计占 {pct(distribution_analysis['top5_profit_share'])}，说明月度结果高度集中在少数游戏。"
        f"Tada占下注额 {pct(tada_row['bet_share'])}、盈利占比 {pct(tada_row['profit_share'])}；Fish占下注额 {pct(game_index['Fish']['bet_share'])}、盈利占比 {pct(game_index['Fish']['profit_share'])}。"
        f"Whot的下注额占比为 {pct(whot_row['bet_share'])}、盈利占比为 {pct(whot_row['profit_share'])}，实际RTP {pct(whot_row['actual_rtp'])}；OMG的下注额占比为 {pct(omg_row['bet_share'])}、盈利占比为 {pct(omg_row['profit_share'])}，实际RTP {pct(omg_row['actual_rtp'])}。"
        f"推荐方案：1）将高下注规模且数据完整的游戏作为主推荐池，但不只按盈利排序；2）Whot、OMG等盈利占比明显高于下注占比的游戏进入风险观察池，先核查留存、体验、退出/放弃和业务成功率，再决定是否扩量；3）EasyWin实际RTP {pct(easywin_row['actual_rtp'])}但下注额占比仅 {pct(easywin_row['bet_share'])}，只进入小流量测试和数据复核，不作为全局推荐依据；4）第三方游戏没有可比预期RTP，推荐排序使用实际业务结果和体验指标，不用配置偏离做结论。</div>"
    )
    distribution_analysis_md_note = (
        f"分布汇总与推荐策略：下注额前5款游戏合计占月度下注额 **{pct(distribution_analysis['top5_bet_share'])}**，前8款合计占 **{pct(distribution_analysis['top8_bet_share'])}**；盈利贡献前5款合计占 **{pct(distribution_analysis['top5_profit_share'])}**，月度结果高度集中在少数游戏。"
        f"Tada占下注额 **{pct(tada_row['bet_share'])}**、盈利占比 **{pct(tada_row['profit_share'])}**；Fish占下注额 **{pct(game_index['Fish']['bet_share'])}**、盈利占比 **{pct(game_index['Fish']['profit_share'])}**。"
        f"Whot的下注额占比为 **{pct(whot_row['bet_share'])}**、盈利占比为 **{pct(whot_row['profit_share'])}**，实际RTP **{pct(whot_row['actual_rtp'])}**；OMG的下注额占比为 **{pct(omg_row['bet_share'])}**、盈利占比为 **{pct(omg_row['profit_share'])}**，实际RTP **{pct(omg_row['actual_rtp'])}**。\n\n"
        f"推荐方案：高下注规模且数据完整的游戏作为主推荐池，但不只按盈利排序；Whot、OMG等盈利占比明显高于下注占比的游戏进入风险观察池，先核查留存、体验、退出/放弃和业务成功率，再决定是否扩量；EasyWin实际RTP **{pct(easywin_row['actual_rtp'])}**但下注额占比仅 **{pct(easywin_row['bet_share'])}**，只进入小流量测试和数据复核，不作为全局推荐依据；第三方游戏没有可比预期RTP，推荐排序使用实际业务结果和体验指标，不用配置偏离做结论。"
    )
    package_split_note = (
        "<div class=\"callout\"><strong>后续包体拆分方案：</strong>本期生命周期源快照没有平台/包体字段，因此本月只能出全产品汇总，不能反推 Android 或 H5 的下注额、盈利和RTP。后续报告必须按 <code>platform × package_name × version × game × lifecycle</code> 分层；至少增加平台（Android/H5）、包体名、版本（app_version/web_version）、release_id、下注额、包体内下注额占比、全产品下注额占比、实际盈利、盈利占比、实际RTP、数据状态等栏位，并提供平台、包体、版本、游戏、生命周期筛选。包体分层合计必须与全产品总额对账；缺少包体字段时显示“待补字段”，不归入 Android 或 H5。</div>"
    )
    package_split_md_note = (
        "后续包体拆分方案：本期生命周期源快照没有平台/包体字段，因此本月只能出全产品汇总，不能反推 Android 或 H5 的下注额、盈利和RTP。后续报告必须按 `platform × package_name × version × game × lifecycle` 分层；至少增加平台（Android/H5）、包体名、版本（app_version/web_version）、release_id、下注额、包体内下注额占比、全产品下注额占比、实际盈利、盈利占比、实际RTP、数据状态等栏位，并提供平台、包体、版本、游戏、生命周期筛选。包体分层合计必须与全产品总额对账；缺少包体字段时显示“待补字段”，不归入 Android 或 H5。"
    )
    valid_detail_rows = [row for row in display_detail if row.get("average_gap") is not None]
    largest_detail_outlier = max(valid_detail_rows, key=lambda row: abs(row["average_gap"])) if valid_detail_rows else None
    negative_detail_impact = sorted(
        [row for row in valid_detail_rows if row["average_gap"] < 0],
        key=lambda row: row["full_bet"] * abs(row["average_gap"]),
        reverse=True,
    )
    if valid_detail_rows:
        life3_avg = lifecycle_average.get(3)
        life4_avg = lifecycle_average.get(4)
        impact_refs = "、".join(f"{row['game']}-生命周期{row['lifecycle']}（{pp(row['average_gap'])}，下注{money(row['full_bet'])}）" for row in negative_detail_impact[:2])
        detail_analysis_html = (
            f"<div class=\"definition\"><strong>明细简析：</strong>生命周期3加权平均RTP为 {pct(life3_avg)}，生命周期4为 {pct(life4_avg)}。"
            f"绝对偏离最大的项目是 {html.escape(largest_detail_outlier['game'])}-生命周期{largest_detail_outlier['lifecycle']}（{pp(largest_detail_outlier['average_gap'])}），但需结合下注规模判断。"
            f"按偏离幅度与下注规模共同看，优先关注 {html.escape(impact_refs)}；正负偏离使用不同色调，绝对偏离达到约1、2、5个百分点时颜色逐级加深。高亮仅表示相对同生命周期平均值偏离，不代表因果或系统故障。</div>"
        )
        detail_analysis_md = (
            f"明细简析：生命周期3加权平均RTP为 **{pct(life3_avg)}**，生命周期4为 **{pct(life4_avg)}**。"
            f"绝对偏离最大的项目是 {largest_detail_outlier['game']}-生命周期{largest_detail_outlier['lifecycle']}（{pp(largest_detail_outlier['average_gap'])}），需结合下注规模判断。"
            f"按偏离幅度与下注规模共同看，优先关注 {impact_refs}；正负偏离使用不同色调，绝对偏离达到约1、2、5个百分点时颜色逐级加深。高亮仅表示相对同生命周期平均值偏离，不代表因果或系统故障。"
        )
    else:
        detail_analysis_html = "<div class=\"definition\"><strong>明细简析：</strong>当前没有可用于计算生命周期平均的有效 RTP 行。</div>"
        detail_analysis_md = "明细简析：当前没有可用于计算生命周期平均的有效 RTP 行。"

    source_days_complete = sum(1 for r in source_days if r["row_count_match"] and r["receipt_status"] in {"ok", "success", "complete", ""})
    duplicate_conflicts = [r["date"] for r in duplicate_checks if r["status"] == "conflict_review_required"]
    overall_status = "ok" if len(source_days) == 31 and source_days_complete == 31 and not duplicate_conflicts and len(tc_rows) == 31 and not tc_overlap else "partial"
    query_sql = """SELECT\n  DATE(FROM_UNIXTIME(time)) AS business_date,\n  SUM(CASE WHEN type = 1 AND status = 3 THEN amount ELSE 0 END) / 100.0 AS success_recharge_amount,\n  SUM(CASE WHEN type = 2 AND status = 103 THEN amount ELSE 0 END) / 100.0 AS success_withdraw_amount,\n  SUM(CASE WHEN type = 2 AND status = 103 THEN amount ELSE 0 END) / NULLIF(SUM(CASE WHEN type = 1 AND status = 3 THEN amount ELSE 0 END), 0) AS tc_rate\nFROM whot_center.order_log\nWHERE time >= UNIX_TIMESTAMP('2026-08-01 00:00:00')\n  AND time < UNIX_TIMESTAMP('2026-09-01 00:00:00')\nGROUP BY 1\nORDER BY business_date;"""
    (OUT / "tc_query.sql").write_text(query_sql + "\n", encoding="utf-8")

    chart_data = {
        "weekly": weekly,
        "lifecycle": lifecycle_summary,
        "top_games": top_games,
        "profit_top_games": profit_top_games,
        "heatmap_lifecycle_3_4": heat_rows,
        "tc_daily": tc_rows,
        "third_party_tada_pp": third_party_tada_pp,
        "distribution_analysis": distribution_analysis,
        "data_state": "actual",
    }
    dump(OUT / "source_manifest.json", {
        "status": overall_status,
        "scope": {"start": MONTH_START.isoformat(), "end": MONTH_END.isoformat(), "timezone": "Asia/Hong_Kong", "lifecycle": "1—4"},
        "canonical_selection": "8/1—8/27 from 2026-08-28-30d; 8/28—8/30 from 2026-08-31; 8/31 from 2026-09-02",
        "dates": source_days,
        "duplicate_checks": duplicate_checks,
        "source_headers_observed": sorted(all_headers),
    })
    dump(OUT / "daily_lifecycle.json", day_records)
    dump(OUT / "monthly_detail_124.json", monthly_detail)
    dump(OUT / "display_detail_lifecycle_3_4_50.json", display_detail)
    dump(OUT / "lifecycle_summary.json", lifecycle_summary)
    dump(OUT / "game_summary.json", game_summary)
    dump(OUT / "third_party_tada_pp.json", third_party_tada_pp)
    dump(OUT / "distribution_analysis.json", distribution_analysis)
    dump(OUT / "package_split_plan.json", package_split_plan)
    dump(OUT / "weekly_overview.json", weekly)
    dump(OUT / "monthly_overview.json", monthly)
    dump(OUT / "tc_daily.json", {"source": tc_source, "rows": tc_rows, "overlap_mismatches": tc_overlap, "missing_dates": tc_missing_dates, "data_state": "actual_aggregate"})
    dump(OUT / "new_user_context.json", new_user)
    dump(OUT / "chart_data.json", chart_data)
    dump(OUT / "quality_checks.json", {
        "status": overall_status,
        "lifecycle_date_coverage": {"expected": 31, "observed": len(source_days), "receipt_complete": source_days_complete},
        "monthly_detail_rows": len(monthly_detail),
        "display_detail_rows": len(display_detail),
        "expected_display_detail_rows": len(display_detail),
        "hidden_empty_detail_rows": len(hidden_detail_rows),
        "hidden_empty_detail_keys": [[r["game"], r["lifecycle"]] for r in hidden_detail_rows],
        "visual_highlights": {
            "game_name": "同步高亮实际-预期或偏离生命周期平均值达到阈值的游戏名称",
            "heatmap": "游戏名称和对应格子按同生命周期平均RTP偏离同步高亮；97%绝对RTP配色保留",
            "average_gap_threshold_pp": {"highlight": 1, "critical": 2},
            "signed_gap_color_scale": {"positive": "绿色系", "negative": "红色系", "levels_pp": [0.5, 1, 2, 5], "small_nonzero_visible": True},
            "tc_baseline_color_scale": {"positive": "绿色系", "negative": "红色系", "small_nonzero_visible": True},
        },
        "tc_result_rows": len(tc_rows),
        "tc_missing_dates": tc_missing_dates,
        "tc_overlap_mismatches": tc_overlap,
        "duplicate_conflicts": duplicate_conflicts,
        "removed_display_columns": removed_columns,
        "formulas": {
            "actual_rtp": "1 - SUM(完全实际盈利) / SUM(完全下注额)",
            "expected_rtp": "自研或有有效配置的游戏采用下注额加权预期RTP；第三方游戏不适用该口径，显示N/A",
            "third_party_expected_rtp": "Tada、OMG、PP等第三方游戏没有可比的预期RTP，不视为数据缺失",
            "game_bet_share": "单款游戏生命周期1—4完全下注额 / 月度生命周期1—4完全下注总额",
            "third_party_tada_pp_bet_share": "(Tada完全下注额 + PP完全下注额) / 月度生命周期1—4完全下注总额",
            "game_profit_share": "单款游戏完全实际盈利 / 月度完全实际盈利总额",
            "profit_bet_share_gap": "盈利占比 - 下注额占比",
            "tc": "SUM(成功提现金额) / SUM(成功充值金额)",
        },
        "third_party_tada_pp": third_party_tada_pp,
        "distribution_analysis": distribution_analysis,
        "package_split_plan": package_split_plan,
        "no_zero_fill": True,
    })
    artifact = {
        "surface": "standalone_html_and_feishu_docx",
        "title": "Waje 8月全产品TC比与游戏×生命周期分析月报｜2026年8月1—31日",
        "status": overall_status,
        "data_state": "actual_aggregate_only",
        "scope": {"start": MONTH_START.isoformat(), "end": MONTH_END.isoformat(), "periods": [p["key"] for p in PERIODS]},
        "sources": [
            {"id": "lifecycle_joint", "path": "data/raw/lifecycle_joint/<canonical-date-folders>", "hash_manifest": "source_manifest.json", "grain": "business_date × game × lifecycle", "rows": len(day_records)},
            {"id": "metabase_tc", "path": "analysis/tc_august_monthly_2026_09_03/tc_query_result.json", "query": "tc_query.sql", "grain": "business_date", "rows": len(tc_rows), "query_result_metadata": {"bytes_processed": "not_exposed_by_page", "elapsed_ms": "not_exposed_by_page"}},
            {"id": "origin_new_user_context", "path": str(NEW_USER_FILE.relative_to(ROOT)) if NEW_USER_FILE.exists() else None, "grain": "package/channel × accepted cohort dates", "status": new_user.get("status")},
        ],
        "metrics": {"monthly_lifecycle_1_4": monthly, "game_count": len(games), "monthly_detail_rows": 124, "visible_detail_rows": len(display_detail), "hidden_empty_detail_rows": len(hidden_detail_rows), "third_party_tada_pp": third_party_tada_pp, "distribution_analysis": distribution_analysis, "package_split_plan": package_split_plan},
        "presentation_rule": "remove only columns entirely zero or blank in the displayed table; retain N/A and missing-status columns",
        "privacy": "aggregate only; no user, order-detail, account, device or credential data",
    }
    dump(OUT / "artifact.json", artifact)

    # Markdown archive for local review and Feishu content drafting.
    md = f"""# Waje 8月全产品TC比与游戏×生命周期分析月报\n\n> 统计周期：2026-08-01—2026-08-31｜时区：Asia/Hong_Kong｜数据状态：{overall_status}｜本报告只使用真实聚合结果。\n\n## 结论摘要\n\n- 生命周期1—4月度完全下注额：**{money(monthly['full_bet'])}**；实际RTP：**{pct(monthly['actual_rtp'])}**。\n- 四周按 **8/1—7、8/8—14、8/15—21、8/22—31（月末10天）** 拆分；第4周同时展示日均下注额。\n- TC查询返回 **{len(tc_rows)}/31** 个业务日；与本地已有 8/19—8/31 快照重叠核对：**{'通过' if not tc_overlap else '存在差异'}**。\n- 月度游戏×生命周期1—4共 **124 行**；最后详细区只展示生命周期3、4，共 **62 行**。\n- 预期RTP仅在来源配置字段有效时计算；0或空值显示为 `N/A`，不从预期盈利反推。\n\n## 一、四周整体对比\n\n{weekly_table}\n\nTC比按成功提现金额除以成功充值金额计算；RTP按累计金额加权，不对每日百分比做简单平均。\n\n## 二、生命周期1—4回报结构\n\n{life_table}\n\n## 三、分游戏回报与生命周期深钻\n\n{game_table}\n\n## 四、TC日级核验\n\n{tc_table}\n\n查询来源：Metabase `whot_center.order_log`；成功充值为 `type=1,status=3`，成功提现为 `type=2,status=103`。页面结果的 TC 展示值为四舍五入后的显示值，报告以金额重新计算精确比例。\n\n## 五、新增用户付费背景\n\n当前可核验的成熟背景数据覆盖 **2026-08-06—2026-08-30**；8/1—8/5及8/31不纳入成熟留存结论。该背景没有游戏维度，不与单款游戏RTP直接关联。\n\n{new_user_table}\n\n## 六、数据边界与后续动作\n\n1. 生命周期来源覆盖 31/31 个日期，月度明细按真实观察到的游戏集合形成，不将未观察到补成0。\n2. 第4周是10天周期，比较总量时必须同时阅读日均下注额。\n3. 配置覆盖率不足的游戏只能做实际RTP观察，不能视为与配置偏离已被认证。\n4. TC完整月度结论依赖31天查询结果与重叠日期核对；如果后续刷新出现差异，应冻结TC区域并重新核对状态映射。\n\n## 七、最后明细\n\n完整明细规模为 124 行；本报告展示生命周期3/4的 62 行。完整124行保存在 `analysis/tc_august_monthly_2026_09_03/monthly_detail_124.json`。\n\n{detail_table}\n\n## 来源与回执\n\n- 生命周期：`analysis/tc_august_monthly_2026_09_03/source_manifest.json`。\n- TC：`analysis/tc_august_monthly_2026_09_03/tc_query.sql`、`tc_daily.json`。\n- 质量检查：`analysis/tc_august_monthly_2026_09_03/quality_checks.json`。\n- 所有结果为聚合数据，不含用户、订单明细、支付账户、设备唯一标识或凭据。\n"""
    md = md.replace(
        f"{tc_table}\n\n查询来源",
        f"{tc_table}\n\n{tc_color_md_note}\n\n查询来源",
        1,
    )
    md = md.replace(
        f"{game_table}\n\n## 四、TC日级核验",
        f"{game_table}\n\n{game_share_md_note}\n\n{deviation_color_md_note}\n\n{distribution_analysis_md_note}\n\n{package_split_md_note}\n\n## 四、TC日级核验",
        1,
    )
    md = md.replace(
        "- 月度游戏×生命周期1—4共 **124 行**；最后详细区只展示生命周期3、4，共 **62 行**。",
        f"- 月度游戏×生命周期1—4共 **124 行**；最后详细区只展示有有效下注的生命周期3、4，共 **{len(display_detail)} 行**；无有效数据行隐藏 **{len(hidden_detail_rows)} 行**。\n- TaDa（源数据名 Tada）与PP合计下注额：**{money(third_party_tada_pp['combined_bet'])}**，占月度总下注额 **{pct(third_party_tada_pp['combined_share'])}**；下注额前5款占 **{pct(distribution_analysis['top5_bet_share'])}**，盈利贡献前5款占 **{pct(distribution_analysis['top5_profit_share'])}**。",
    )
    md = md.replace(
        "TC比按成功提现金额除以成功充值金额计算；RTP按累计金额加权，不对每日百分比做简单平均。",
        "配置覆盖率说明：有效预期RTP配置对应的完全下注额 ÷ 生命周期1—4的完全下注总额。它反映有多少下注额可以进行实际RTP与配置预期RTP对比，不代表游戏数覆盖率或实际RTP表现。配置为空、为0或无效时不纳入分子；比例不足100%时，预期RTP对比只适用于已覆盖的下注额。\n\nTC比按成功提现金额除以成功充值金额计算；RTP按累计金额加权，不对每日百分比做简单平均。",
    )
    md = md.replace(
        "## 三、分游戏回报与生命周期深钻",
        "生命周期组合图使用柱状图表示完全下注额、折线表示实际RTP，配置预期RTP及偏离程度在表格中查看；左右坐标轴分别表示金额和百分比。\n\n## 三、分游戏回报与生命周期深钻",
    )
    md = md.replace(
        "先按生命周期1—4完全下注额列出主要游戏；小额游戏的极端比例不直接判定为经营问题。",
        "先按生命周期1—4完全下注额列出主要游戏。图表采用组合展示：柱状图表示各游戏完全下注额，折线表示实际RTP，柱体和折线点均标注数值；配置预期RTP及偏离程度在表格中查看。小额游戏的极端比例不直接判定为经营问题。",
    )
    md = md.replace(
        "1. 生命周期来源覆盖 31/31 个日期，月度明细按真实观察到的游戏集合形成，不将未观察到补成0。\n2. 第4周是10天周期，比较总量时必须同时阅读日均下注额。\n3. 配置覆盖率不足的游戏只能做实际RTP观察，不能视为与配置偏离已被认证。\n4. TC完整月度结论依赖31天查询结果与重叠日期核对；如果后续刷新出现差异，应冻结TC区域并重新核对状态映射。",
        "1. 第三方游戏不提供可比的预期RTP，该情况属于口径不适用，不是数据缺失；真正的配置异常另行核验。\n2. TC后续如出现重叠日期差异，先核对时间范围、金额单位和状态映射。",
    )
    md = md.replace(
        "预期RTP仅在来源配置字段有效时计算；0或空值显示为 `N/A`，不从预期盈利反推。",
        "自研或有有效配置的游戏使用预期RTP；Tada、OMG、PP等第三方游戏没有可比预期RTP，显示 `N/A` 表示口径不适用，不是数据缺失。",
    )
    md = md.replace(
        "月度游戏×生命周期1—4共 **124 行**；最后详细区只展示生命周期3、4，共 **62 行**。",
        f"月度游戏×生命周期1—4共 **124 行**；最后详细区只展示有有效下注的生命周期3、4，共 **{len(display_detail)} 行**；无有效数据行隐藏 **{len(hidden_detail_rows)} 行**。",
    )
    md = md.replace(
        "完整明细规模为 124 行；本报告展示生命周期3/4的 62 行。",
        f"完整明细规模为 124 行；本报告展示有有效下注的生命周期3/4，共 {len(display_detail)} 行；另有 {len(hidden_detail_rows)} 行因下注额为0而隐藏。",
    )
    md = md.replace(
        "## 来源与回执",
        f"{detail_analysis_md}\n\n## 来源与回执",
    )
    MD_OUT.parent.mkdir(parents=True, exist_ok=True)
    MD_OUT.write_text(md, encoding="utf-8")

    figures = [
        ("01_四周TC与生命周期RTP.png", "四周 TC 比与生命周期 1—4 实际 RTP"),
        ("02_生命周期RTP与下注额.png", "生命周期1—4：下注额（柱）与RTP（折线）"),
        ("03_主要游戏下注规模.png", "主要游戏：下注额（柱）与实际RTP（折线）"),
        ("04_生命周期3与4RTP热力图.png", "游戏 × 生命周期 3/4 实际 RTP"),
        ("05_每日TC趋势.png", "每日 TC 比"),
        ("06_主要游戏盈利贡献.png", "主要游戏实际盈利贡献"),
    ]
    figure_parts = [f'<figure><img src="{data_uri(ASSETS / filename)}" alt="{html.escape(title)}"/><figcaption><strong>{html.escape(title)}</strong><br/>实际聚合数据；金额加权；具体窗口和公式见表格及来源说明。</figcaption></figure>' for filename, title in figures]
    summary_status = "已核验实际聚合" if overall_status == "ok" else "部分数据已核验"
    weekly_status = "已完成31天TC聚合并通过重叠日期核对" if not tc_overlap and not tc_missing_dates else "TC部分待补证"
    removed_note = "；".join(f"{r['scope']}：{r['label']}" for r in removed_columns) or "本次展示表没有整列 0/空字段"
    html_doc = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/><title>Waje 8月全产品TC比与游戏生命周期分析月报</title>
<style>
:root {{ color-scheme:light; --ink:#17324d; --muted:#61758b; --line:#d9e6f3; --panel:#f7fbff; --blue:#2563eb; --gold:#b7791f; --orange:#ea580c; --green:#15803d; }}
* {{ box-sizing:border-box; }} body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif; color:var(--ink); background:#f1f5f9; line-height:1.65; }}
main {{ width:min(1280px,calc(100% - 36px)); margin:28px auto 54px; background:#fff; padding:54px 62px; border-radius:18px; box-shadow:0 10px 30px #17324d12; }}
h1 {{ font-size:34px; line-height:1.2; margin:0 0 12px; }} h2 {{ font-size:24px; margin:48px 0 16px; padding-top:4px; }} h3 {{ font-size:18px; margin:26px 0 10px; }} p {{ margin:10px 0; }} .meta {{ color:var(--muted); margin:0; }}
.tag {{ display:inline-block; padding:4px 10px; background:#eff6ff; color:#1d4ed8; border-radius:999px; font-size:13px; margin-right:8px; }} .tag.actual {{ background:#ecfdf5; color:#166534; }} .tag.warn {{ background:#fff7e7; color:#9a5b00; }}
.callout {{ margin:28px 0; padding:18px 22px; border-left:5px solid var(--blue); background:#eff6ff; border-radius:10px; }} .callout.warn {{ border-left-color:var(--gold); background:#fff7e7; }} .definition {{ margin:16px 0 24px; padding:14px 18px; border:1px solid var(--line); background:#fafcff; border-radius:10px; color:#40566e; }}
.grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:14px; margin:22px 0; }} .kpi {{ border:1px solid var(--line); background:var(--panel); padding:18px; border-radius:12px; }} .kpi .label {{ color:var(--muted); font-size:14px; }} .kpi strong {{ display:block; font-size:27px; margin-top:5px; }} .kpi small {{ color:var(--muted); }}
.table-wrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:12px; margin:16px 0 24px; }} table {{ width:100%; border-collapse:collapse; font-size:14px; }} th {{ background:#eef5fb; color:#25445f; white-space:nowrap; }} th,td {{ padding:10px 12px; border-bottom:1px solid var(--line); text-align:right; vertical-align:top; }} th:first-child,td:first-child {{ text-align:left; }} tr:last-child td {{ border-bottom:0; }} .dense table {{ font-size:12px; }} .dense th,.dense td {{ padding:8px 9px; }} .deviation-highlight {{ background:#fff7ed; color:#9a3412; font-weight:700; }} .deviation-critical {{ background:#fee2e2; color:#991b1b; font-weight:750; }} .baseline-alert {{ background:#ffedd5; color:#9a3412; font-weight:750; }} .context-high {{ background:#e0f2fe; color:#075985; font-weight:750; }} .context-low {{ background:#fee2e2; color:#991b1b; font-weight:750; }} tr.context-row-high td {{ background:#f0f9ff; }} tr.context-row-low td {{ background:#fff1f2; }} tr.context-row-high td.context-high {{ background:#e0f2fe; }} tr.context-row-low td.context-low {{ background:#fee2e2; }} tr.detail-row-highlight td {{ background:#fff7ed; }} tr.detail-row-critical td {{ background:#fee2e2; }} tr.detail-row-highlight td.deviation-highlight {{ background:#fed7aa; }} tr.detail-row-critical td.deviation-critical {{ background:#fca5a5; }}
figure {{ margin:30px 0; border:1px solid var(--line); border-radius:14px; padding:12px; background:#fff; }} figure img {{ display:block; width:100%; height:auto; border-radius:8px; }} figcaption {{ color:var(--muted); padding:10px 5px 2px; font-size:14px; }} figcaption strong {{ color:var(--ink); }}
details {{ margin-top:20px; border:1px solid var(--line); border-radius:12px; padding:0 16px; }} summary {{ cursor:pointer; padding:14px 0; color:#1d4ed8; font-weight:650; }} code {{ background:#f1f5f9; padding:2px 5px; border-radius:4px; }} li {{ margin:8px 0; }} .sources {{ color:var(--muted); font-size:13px; border-top:1px solid var(--line); margin-top:44px; padding-top:20px; }}
@media (max-width:780px) {{ main {{ width:100%; margin:0; border-radius:0; padding:30px 20px; }} .grid {{ grid-template-columns:1fr 1fr; }} h1 {{ font-size:28px; }} h2 {{ font-size:22px; }} }} @media (max-width:520px) {{ .grid {{ grid-template-columns:1fr; }} th,td {{ padding:8px 9px; }} }} @media print {{ body {{ background:#fff; }} main {{ width:100%; box-shadow:none; margin:0; padding:22px; }} figure {{ break-inside:avoid; }} details {{ break-before:page; }} }}
</style></head><body><main>
<div><span class="tag actual">{html.escape(summary_status)}</span><span class="tag">8月月报</span><span class="tag">生命周期1—4</span><span class="tag">4个周期</span></div>
<h1>Waje 8月全产品 TC 比与游戏 × 生命周期分析月报</h1>
<p class="meta">业务日期：2026-08-01—2026-08-31 ｜ 数据时区：Asia/Hong_Kong ｜ 第4周期为月末10天 ｜ 报告生成：2026-09-03</p>
<div class="callout"><strong>结论先行：</strong>8月生命周期1—4累计完全下注额为 <strong>{money(monthly['full_bet'])}</strong>，实际RTP为 <strong>{pct(monthly['actual_rtp'])}</strong>；TC查询返回 <strong>{len(tc_rows)}/31</strong> 个业务日，{html.escape(weekly_status)}。本月游戏×生命周期明细共124行，最后展开区只展示生命周期3/4的62行。</div>
<div class="grid"><div class="kpi"><div class="label">月度完全下注额</div><strong>{money(monthly['full_bet'])}</strong><small>生命周期1—4｜31日</small></div><div class="kpi"><div class="label">月度实际RTP</div><strong>{pct(monthly['actual_rtp'])}</strong><small>1 − 实际盈利 ÷ 下注额</small></div><div class="kpi"><div class="label">月度TC比</div><strong>{pct(monthly.get('tc'))}</strong><small>成功提现 ÷ 成功充值</small></div><div class="kpi"><div class="label">观察到的游戏数</div><strong>{len(games)}</strong><small>真实观察集合，不补造</small></div></div>
<section><h2>1. 月度与四周整体对比</h2><p>四个周期按 7/7/7/10 天划分。第4周期不只看总量，还看日均下注额；TC和RTP均按累计金额加权。</p>{weekly_table}<div class="definition"><strong>配置覆盖率说明：</strong>有效预期RTP配置对应的完全下注额 ÷ 生命周期1—4的完全下注总额。它反映“有多少下注额可以进行实际RTP与配置预期RTP对比”，不代表游戏数覆盖率或实际RTP表现。配置为空、为0或无效时不纳入分子；比例不足100%时，预期RTP对比只适用于已覆盖的下注额。</div>{figure_parts[0]}</section>
<section><h2>2. 生命周期1—4回报结构</h2><p>预期RTP仅使用来源中有效的配置预期回报比，并按下注额加权。配置字段为0或空时显示 <code>N/A</code>，不把缺失配置当作0。图表采用组合展示：柱状图表示完全下注额，折线表示实际RTP；配置预期RTP以虚线作为对照，左右坐标轴分别表示金额和百分比。</p>{life_table}{figure_parts[1]}</section>
<section><h2>3. 分游戏回报与生命周期深钻</h2><p>先看下注规模，再看RTP偏离和有效观察日。游戏组合图采用柱状图表示各游戏完全下注额、折线表示实际RTP，并直接标注数值；配置预期RTP和偏离程度在表格中查看。表中实际-预期偏离达到 1 个百分点时浅橙高亮，达到 2 个百分点时红色高亮；热力图中的游戏名称和对应格子也按同生命周期平均 RTP 的偏离程度同步高亮。只有配置有效的项目参与实际-预期判断。小额游戏的极端比例不直接判定为经营问题。</p>{game_table}{figure_parts[2]}{figure_parts[3]}</section>
<section><h2>4. TC日级核验</h2><p>来源为 Metabase <code>whot_center.order_log</code> 聚合查询。成功充值：<code>type=1,status=3</code>；成功提现：<code>type=2,status=103</code>。查询页面的TC展示值为两位小数，报告用金额重新计算精确比例。日级基线为 8 月加权 TC 比；相对基线绝对偏离达到 2 个百分点的日期在表格中橙色高亮，仅表示优先核查，不代表归因结论。</p>{tc_table}{figure_parts[4]}</section>
<section><h2>5. 新增用户付费背景</h2><p>当前可核验成熟日期为 2026-08-06—2026-08-30；8/1—8/5和8/31不纳入成熟留存结论。该来源没有游戏维度，只作为同期包体/渠道背景。</p>{context_average_note}{new_user_table}{context_summary}</section>
<section><h2>6. 数据限制与行动建议</h2><div class="callout warn"><strong>展示列清理：</strong>{html.escape(removed_note)}。局部单元格为0不会删除整列；<code>N/A</code>和“待补数据”用于表达不可用状态，保留。</div><ol><li>生命周期来源覆盖 {len(source_days)}/31 个日期，月度明细按真实观察到的游戏集合形成。</li><li>第4周期为10天，横向比较时使用日均下注额辅助判断。</li><li>配置覆盖率不足的游戏只能做实际RTP观察，不能视为预期偏离已认证。</li><li>后续刷新如果TC重叠日期出现差异，应冻结TC结论并复核时间边界、金额单位和状态映射。</li></ol></section>
<details><summary>展开：全部游戏 × 生命周期1—4明细（124行；正文展示生命周期3/4，共62行）</summary>{detail_table}{detail_analysis_html}<p class="meta">完整124行结构化明细保存在本地分析快照，不在正文全部展开。</p></details>
<div class="sources"><strong>来源与回执</strong><br/>生命周期：本地 GM Lifecycle Pool V2（Joint）8月1—31日原始快照；TC：Metabase 只读聚合查询；新增用户背景：Origin BQ新增付费用户分析的成熟日期快照。完整源文件哈希、查询SQL、质量检查和图表数据位于本地分析目录。所有输出为聚合数据，不含用户、订单明细、支付账户、设备唯一标识或凭据。</div>
</main></body></html>"""
    deviation_css = """
.deviation-positive-highlight { background:#dbeafe !important; color:#1d4ed8 !important; font-weight:700; }
.deviation-positive-critical { background:#93c5fd !important; color:#1e3a8a !important; font-weight:800; }
.deviation-positive-extreme { background:#2563eb !important; color:#ffffff !important; font-weight:800; }
.deviation-positive-soft { background:#eff6ff !important; color:#2563eb !important; font-weight:650; }
.deviation-positive-faint { background:#f8fbff !important; color:#315b91 !important; font-weight:600; }
.deviation-negative-highlight { background:#ffedd5 !important; color:#c2410c !important; font-weight:700; }
.deviation-negative-critical { background:#fdba74 !important; color:#9a3412 !important; font-weight:800; }
.deviation-negative-extreme { background:#dc2626 !important; color:#ffffff !important; font-weight:800; }
.deviation-negative-soft { background:#fff7ed !important; color:#c2410c !important; font-weight:650; }
.deviation-negative-faint { background:#fffaf5 !important; color:#99532f !important; font-weight:600; }
.tc-positive-highlight { background:#bbf7d0 !important; color:#166534 !important; font-weight:750; }
.tc-positive-critical { background:#4ade80 !important; color:#14532d !important; font-weight:800; }
.tc-positive-extreme { background:#16a34a !important; color:#ffffff !important; font-weight:800; }
.tc-positive-soft { background:#dcfce7 !important; color:#15803d !important; font-weight:650; }
.tc-positive-faint { background:#f0fdf4 !important; color:#26734d !important; font-weight:600; }
.tc-negative-highlight { background:#fca5a5 !important; color:#991b1b !important; font-weight:750; }
.tc-negative-critical { background:#ef4444 !important; color:#ffffff !important; font-weight:800; }
.tc-negative-extreme { background:#b91c1c !important; color:#ffffff !important; font-weight:800; }
.tc-negative-soft { background:#fee2e2 !important; color:#b91c1c !important; font-weight:650; }
.tc-negative-faint { background:#fff5f5 !important; color:#b45353 !important; font-weight:600; }
"""
    html_doc = html_doc.replace("</style></head>", deviation_css + "</style></head>", 1)
    html_doc = html_doc.replace(
        "。本月游戏×生命周期明细共124行，最后展开区只展示生命周期3/4的62行。",
        f"。TaDa（源数据名 Tada）与PP合计下注额为 {money(third_party_tada_pp['combined_bet'])}，占月度总下注额 {pct(third_party_tada_pp['combined_share'])}；下注额前5款占 {pct(distribution_analysis['top5_bet_share'])}，盈利贡献前5款占 {pct(distribution_analysis['top5_profit_share'])}。本月游戏×生命周期明细共124行，最后展开区只展示有有效下注的生命周期3/4，共{len(display_detail)}行。",
        1,
    )
    html_doc = html_doc.replace(
        "先看下注规模，再看RTP偏离和有效观察日。游戏组合图采用柱状图表示各游戏完全下注额、折线表示实际RTP，并直接标注数值；配置预期RTP和偏离程度在表格中查看。",
        "先看下注规模，再看RTP偏离和有效观察日。游戏组合图采用柱状图表示各游戏完全下注额、折线表示实际RTP；柱顶同时标注下注额和月度下注额占比，折线点标注实际RTP。配置预期RTP和偏离程度在表格中查看。",
        1,
    )
    html_doc = html_doc.replace(
        "表中实际-预期偏离达到 1 个百分点时浅橙高亮，达到 2 个百分点时红色高亮；",
        "表中实际-预期偏离按正负使用不同色调：绿色系表示正偏离，红色系表示负偏离；非零偏离从浅色开始，绝对偏离达到约0.5、1、2、5个百分点时颜色逐级加深；",
        1,
    )
    html_doc = html_doc.replace(
        f"{game_table}{figure_parts[2]}",
        f"{game_table}{game_share_note}{deviation_color_note}{distribution_analysis_note}{package_split_note}{figure_parts[2]}{figure_parts[5]}",
        1,
    )
    html_doc = html_doc.replace(
        "日级基线为 8 月加权 TC 比；相对基线绝对偏离达到 2 个百分点的日期在表格中橙色高亮，仅表示优先核查，不代表归因结论。",
        "日级基线为 8 月加权 TC 比；偏离值按正负使用不同色调，绿色表示高于基线，红色表示低于基线，绝对偏离越大颜色越深，仅表示优先核查，不代表归因结论。",
        1,
    )
    html_doc = html_doc.replace(
        f"{tc_table}{figure_parts[4]}",
        f"{tc_table}{tc_color_note}{figure_parts[4]}",
        1,
    )
    html_doc = re.sub(
        r"<section><h2>6\. 数据限制与行动建议</h2>.*?</section>",
        "<section><h2>6. 数据限制与行动建议</h2><div class=\"callout warn\"><strong>需要注意：</strong>第三方游戏没有可比的预期RTP，显示 N/A 属于口径不适用，不是数据缺失；实际-预期对比只适用于有有效配置的项目。</div><ol><li>TC后续如出现重叠日期差异，先核对时间范围、金额单位和状态映射。</li></ol></section>",
        html_doc,
        count=1,
        flags=re.DOTALL,
    )
    html_doc = html_doc.replace(
        "配置为空、为0或无效时不纳入分子；比例不足100%时，预期RTP对比只适用于已覆盖的下注额。",
        "第三方游戏不提供可比的预期RTP，该情况属于口径不适用，不是数据缺失；真正的配置异常另行核验。",
    )
    html_doc = html_doc.replace(
        "配置字段为0或空时显示 <code>N/A</code>，不把缺失配置当作0。",
        "Tada、OMG、PP等第三方游戏没有可比的预期RTP，显示 <code>N/A</code> 表示口径不适用，不是数据缺失。",
    )
    html_doc = html_doc.replace(
        "本月游戏×生命周期明细共124行，最后展开区只展示生命周期3/4的62行。",
        f"本月游戏×生命周期明细共124行，最后展开区只展示有有效下注的生命周期3/4，共{len(display_detail)}行；无有效数据的{len(hidden_detail_rows)}行已隐藏。",
    )
    html_doc = html_doc.replace(
        "展开：全部游戏 × 生命周期1—4明细（124行；正文展示生命周期3/4，共62行）",
        f"展开：全部游戏 × 生命周期1—4明细（124行；正文展示有有效下注的生命周期3/4，共{len(display_detail)}行）",
    )
    html_doc = html_doc.replace(
        "完整124行结构化明细保存在本地分析快照，不在正文全部展开。",
        f"完整124行结构化明细保存在本地分析快照；其中{len(hidden_detail_rows)}行因下注额为0而不在正文展示。",
    )
    HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
    HTML_OUT.write_text(html_doc, encoding="utf-8")
    dump(OUT / "run_receipt.json", {
        "status": overall_status,
        "completed_at": datetime.now().astimezone().isoformat(),
        "html": str(HTML_OUT.relative_to(ROOT)),
        "markdown": str(MD_OUT.relative_to(ROOT)),
        "source_manifest": "analysis/tc_august_monthly_2026_09_03/source_manifest.json",
        "monthly_detail_rows": len(monthly_detail),
        "visible_detail_rows": len(display_detail),
        "lifecycle_dates": len(source_days),
        "tc_result_rows": len(tc_rows),
        "tc_overlap_mismatches": len(tc_overlap),
        "duplicate_conflict_dates": duplicate_conflicts,
        "game_bet_share_rows": len([row for row in game_summary if row.get("bet_share") is not None]),
        "third_party_tada_pp": third_party_tada_pp,
        "metabase_query": {"status": "success_visible", "result_rows": len(tc_rows), "bytes_processed": "not_exposed_by_page", "elapsed_ms": "not_exposed_by_page", "saved_raw_result": False, "aggregate_snapshot": "tc_query_result.json"},
        "display_column_cleanup_count": len(removed_columns),
        "privacy": "aggregate_only",
    })
    print(json.dumps({"status": overall_status, "html": str(HTML_OUT), "markdown": str(MD_OUT), "monthly_detail_rows": len(monthly_detail), "visible_detail_rows": len(display_detail), "tc_rows": len(tc_rows), "tc_overlap_mismatches": len(tc_overlap), "removed_columns": removed_columns}, ensure_ascii=False))


if __name__ == "__main__":
    main()
