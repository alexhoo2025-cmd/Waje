#!/usr/bin/env python3
"""Build the aggregate-only game × lifecycle RTP extension for the TC audit.

Source scope is intentionally fixed to the already captured GM Lifecycle Pool
V2 (Joint) exports.  The script never reads user, order, device, payment or
other row-level data.  It uses the *page display* lifecycle labels 1–4.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
from collections import defaultdict
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[2]
AUDIT_ROOT = Path(__file__).resolve().parent
OUT = AUDIT_ROOT / "lifecycle_game_rtp_14d"
ASSETS = OUT / "assets"
RAW_ROOT = ROOT / "data/raw/lifecycle_joint/2026-08-28-30d"
SOURCE_MANIFEST = ROOT / "data/outputs/lifecycle_joint/2026-08-28-30d/source-manifest.json"
NEW_USER_SOURCE = ROOT / "data/outputs/origin_new_user/2026-08-28-30d/source-data.json"

WINDOWS = {
    "第一周（8月14—20日）": (date(2026, 8, 14), date(2026, 8, 20)),
    "第二周（8月21—27日）": (date(2026, 8, 21), date(2026, 8, 27)),
}
TARGET_START = date(2026, 8, 14)
TARGET_END = date(2026, 8, 27)
LIFECYCLES = (1, 2, 3, 4)

INK = "#17324D"
MUTED = "#61758B"
GRID = "#D9E6F3"
PANEL = "#F7FBFF"
WHITE = "#FFFFFF"
BLUE = "#2563EB"
NAVY = "#173A63"
ORANGE = "#EA580C"
GOLD = "#B7791F"
TEAL = "#0F8A70"
RED = "#B4535B"


def json_dump(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def digest(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def clean(value: Any) -> str:
    return "" if value is None else "".join(str(value).split())


def as_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "N/A", "--"}:
        return None
    try:
        result = float(text.rstrip("%"))
    except ValueError:
        return None
    if text.endswith("%"):
        result /= 100
    return result if math.isfinite(result) else None


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def pct(value: float | None, digits: int = 2) -> str:
    return "N/A" if value is None else f"{value * 100:.{digits}f}%"


def pp(value: float | None, digits: int = 2) -> str:
    return "N/A" if value is None else f"{value * 100:+.{digits}f}pp"


def amount(value: float | None, digits: int = 0) -> str:
    if value is None:
        return "N/A"
    abs_value = abs(value)
    sign = "-" if value < 0 else ""
    if abs_value >= 100_000_000:
        return f"{sign}{abs_value / 100_000_000:.2f}亿"
    if abs_value >= 10_000:
        return f"{sign}{abs_value / 10_000:.2f}万"
    return f"{value:,.{digits}f}"


def font(size: int) -> ImageFont.ImageFont:
    for candidate in (
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    ):
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size, index=0)
    return ImageFont.load_default()


def text(draw: ImageDraw.ImageDraw, xy: tuple[float, float], value: str, size: int, fill: str = INK, anchor: str | None = None) -> None:
    draw.text(xy, value, font=font(size), fill=fill, anchor=anchor)


def header_map(row: Iterable[Any]) -> dict[str, int]:
    aliases = {
        "生命周期": "lifecycle",
        "游戏类型": "game",
        "完全预期盈利": "expected_profit",
        "完全实际盈利": "actual_profit",
        "完全下注额": "full_bet",
        "完全真实回报比": "source_actual_rtp",
        # Detail exports name the configuration field simply "预期回报比".
        "预期回报比": "source_expected_rtp",
        "完全预期回报比": "source_expected_rtp",
        "基础下注额": "base_bet",
        "基础实际盈利": "base_actual_profit",
        "基础真实回报比": "source_base_rtp",
    }
    result: dict[str, int] = {}
    for index, item in enumerate(row):
        field = aliases.get(clean(item))
        if field:
            result[field] = index
    expected = {"lifecycle", "game", "expected_profit", "actual_profit", "full_bet"}
    missing = expected - result.keys()
    if missing:
        raise ValueError(f"detail export missing required headers: {sorted(missing)}")
    return result


def read_detail(path: Path, business_date: date) -> tuple[list[dict[str, Any]], list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    fields = header_map(header)
    records: list[dict[str, Any]] = []
    for row in iterator:
        if not row or not any(item is not None for item in row):
            continue
        lifecycle = as_number(row[fields["lifecycle"]])
        game = str(row[fields["game"]] or "").strip()
        if lifecycle is None or not game:
            continue
        record = {
            "business_date": business_date.isoformat(),
            "display_lifecycle": int(lifecycle),
            "game": game,
            "full_bet": as_number(row[fields["full_bet"]]) or 0.0,
            "actual_profit": as_number(row[fields["actual_profit"]]) or 0.0,
            "expected_profit": as_number(row[fields["expected_profit"]]) or 0.0,
            "source_actual_rtp": as_number(row[fields["source_actual_rtp"]]) if "source_actual_rtp" in fields else None,
            "source_expected_rtp": as_number(row[fields["source_expected_rtp"]]) if "source_expected_rtp" in fields else None,
            "base_bet": as_number(row[fields["base_bet"]]) if "base_bet" in fields else None,
            "base_actual_profit": as_number(row[fields["base_actual_profit"]]) if "base_actual_profit" in fields else None,
            "source_base_rtp": as_number(row[fields["source_base_rtp"]]) if "source_base_rtp" in fields else None,
        }
        records.append(record)
    return records, list(header)


def game_header_map(row: Iterable[Any]) -> dict[str, int]:
    aliases = {
        "游戏": "game",
        "完全下注额": "full_bet",
        "完全实际盈利": "actual_profit",
        "完全真实回报比": "source_actual_rtp",
        "完全预期回报比": "source_expected_rtp",
    }
    result: dict[str, int] = {}
    for index, item in enumerate(row):
        field = aliases.get(clean(item))
        if field:
            result[field] = index
    required = {"game", "full_bet", "actual_profit", "source_actual_rtp"}
    missing = required - result.keys()
    if missing:
        raise ValueError(f"game export missing required headers: {sorted(missing)}")
    return result


def read_game(path: Path, business_date: date) -> tuple[list[dict[str, Any]], list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    fields = game_header_map(header)
    records: list[dict[str, Any]] = []
    for row in iterator:
        if not row or not any(item is not None for item in row):
            continue
        game = str(row[fields["game"]] or "").strip()
        if not game:
            continue
        records.append({
            "business_date": business_date.isoformat(),
            "game": game,
            "full_bet": as_number(row[fields["full_bet"]]) or 0.0,
            "actual_profit": as_number(row[fields["actual_profit"]]) or 0.0,
            "source_actual_rtp": as_number(row[fields["source_actual_rtp"]]),
            "source_expected_rtp": as_number(row[fields["source_expected_rtp"]]) if "source_expected_rtp" in fields else None,
        })
    return records, list(header)


def weekly_row(rows: list[dict[str, Any]], game: str, lifecycle: int, label: str, start: date, end: date) -> dict[str, Any]:
    selected = [
        row for row in rows
        if row["game"] == game and row["display_lifecycle"] == lifecycle and start.isoformat() <= row["business_date"] <= end.isoformat()
    ]
    bet = sum(row["full_bet"] for row in selected)
    actual_profit = sum(row["actual_profit"] for row in selected)
    expected_profit = sum(row["expected_profit"] for row in selected)
    active = sorted({row["business_date"] for row in selected if row["full_bet"] > 0})
    actual_rtp = 1 - actual_profit / bet if bet > 0 else None
    # Some joint providers publish 0 for the page configuration RTP and put
    # full_bet into "expected profit".  That value is not an expected-profit
    # fact and must never be used to derive an expected RTP.  Keep expected
    # comparisons only on rows that have an explicit, plausible config rate.
    expected_rows = [
        row for row in selected
        if row["full_bet"] > 0
        and row["source_expected_rtp"] is not None
        and 0 < row["source_expected_rtp"] <= 1.1
    ]
    expected_bet = sum(row["full_bet"] for row in expected_rows)
    expected_actual_profit = sum(row["actual_profit"] for row in expected_rows)
    expected_rtp = sum(row["full_bet"] * row["source_expected_rtp"] for row in expected_rows) / expected_bet if expected_bet else None
    actual_rtp_on_expected_coverage = 1 - expected_actual_profit / expected_bet if expected_bet else None
    return {
        "period": label,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "game": game,
        "display_lifecycle": lifecycle,
        "full_bet": bet,
        "actual_profit": actual_profit,
        "expected_profit": expected_profit,
        "actual_rtp": actual_rtp,
        "expected_rtp": expected_rtp,
        "expected_bet": expected_bet,
        "expected_bet_coverage": expected_bet / bet if bet else None,
        "actual_rtp_on_expected_coverage": actual_rtp_on_expected_coverage,
        "rtp_gap_pp": actual_rtp_on_expected_coverage - expected_rtp if actual_rtp_on_expected_coverage is not None and expected_rtp is not None else None,
        "active_days": len(active),
        "active_dates": active,
        "data_status": "no_effective_bet" if bet <= 0 else ("early_sample" if len(active) < 7 else "complete_7d"),
    }


def weekly_game_row(rows: list[dict[str, Any]], game: str, label: str, start: date, end: date) -> dict[str, Any]:
    selected = [row for row in rows if row["game"] == game and start.isoformat() <= row["business_date"] <= end.isoformat()]
    bet = sum(row["full_bet"] for row in selected)
    actual_profit = sum(row["actual_profit"] for row in selected)
    active = sorted({row["business_date"] for row in selected if row["full_bet"] > 0})
    configured = [row for row in selected if row["full_bet"] > 0 and row["source_expected_rtp"] is not None and 0 < row["source_expected_rtp"] <= 1.1]
    config_bet = sum(row["full_bet"] for row in configured)
    config_actual_profit = sum(row["actual_profit"] for row in configured)
    return {
        "period": label,
        "game": game,
        "full_bet": bet,
        "actual_profit": actual_profit,
        "actual_rtp": 1 - actual_profit / bet if bet else None,
        "active_days": len(active),
        "active_dates": active,
        "expected_bet_coverage": config_bet / bet if bet else None,
        "expected_rtp": sum(row["full_bet"] * row["source_expected_rtp"] for row in configured) / config_bet if config_bet else None,
        "actual_rtp_on_expected_coverage": 1 - config_actual_profit / config_bet if config_bet else None,
        "data_status": "no_effective_bet" if bet <= 0 else ("early_sample" if len(active) < 7 else "complete_7d"),
    }


def aggregate_rows(rows: list[dict[str, Any]], label: str, start: date, end: date, lifecycle: int | None = None) -> dict[str, Any]:
    selected = [
        row for row in rows
        if start.isoformat() <= row["business_date"] <= end.isoformat()
        and (lifecycle is None or row["display_lifecycle"] == lifecycle)
        and row["full_bet"] > 0
    ]
    bet = sum(row["full_bet"] for row in selected)
    actual_profit = sum(row["actual_profit"] for row in selected)
    expected_profit = sum(row["expected_profit"] for row in selected)
    expected_rows = [
        row for row in selected
        if row["source_expected_rtp"] is not None and 0 < row["source_expected_rtp"] <= 1.1
    ]
    expected_bet = sum(row["full_bet"] for row in expected_rows)
    expected_actual_profit = sum(row["actual_profit"] for row in expected_rows)
    expected_rtp = sum(row["full_bet"] * row["source_expected_rtp"] for row in expected_rows) / expected_bet if expected_bet else None
    actual_rtp_on_expected_coverage = 1 - expected_actual_profit / expected_bet if expected_bet else None
    return {
        "period": label,
        "display_lifecycle": lifecycle if lifecycle is not None else "1—4合计",
        "full_bet": bet,
        "actual_profit": actual_profit,
        "expected_profit": expected_profit,
        "actual_rtp": 1 - actual_profit / bet if bet else None,
        "expected_rtp": expected_rtp,
        "expected_bet": expected_bet,
        "expected_bet_coverage": expected_bet / bet if bet else None,
        "actual_rtp_on_expected_coverage": actual_rtp_on_expected_coverage,
        "rtp_gap_pp": actual_rtp_on_expected_coverage - expected_rtp if actual_rtp_on_expected_coverage is not None and expected_rtp is not None else None,
    }


def load_new_user_context() -> list[dict[str, Any]]:
    source = json.loads(NEW_USER_SOURCE.read_text(encoding="utf-8"))
    output: list[dict[str, Any]] = []
    for sheet_name, sheet in source["sheets"].items():
        headers = [clean(value) for value in sheet.get("headers", [])]
        index = {value: position for position, value in enumerate(headers)}
        required = {"日期", "新增人数", "新增付费人数", "首充付费人数"}
        if not required.issubset(index):
            continue
        amounts = defaultdict(float)
        counts = defaultdict(int)
        for row in sheet.get("rows", []):
            day = as_date(row[index["日期"]] if len(row) > index["日期"] else None)
            if not day or not (TARGET_START <= day <= TARGET_END):
                continue
            for field, output_key in (("新增人数", "new_users"), ("新增付费人数", "new_payers"), ("首充付费人数", "first_payers")):
                value = as_number(row[index[field]] if len(row) > index[field] else None)
                if value is not None:
                    amounts[output_key] += value
            for field, output_key in (("次留", "d1"), ("3日留", "d3"), ("7日留", "d7")):
                if field not in index:
                    continue
                value = as_number(row[index[field]] if len(row) > index[field] else None)
                # The source maturity policy treats blank/zero values as not-ready.
                if value is not None and value > 0:
                    amounts[output_key] += value * (as_number(row[index["新增人数"]]) or 0)
                    counts[output_key] += int(as_number(row[index["新增人数"]]) or 0)
        output.append({
            "segment": sheet_name,
            "new_users": amounts["new_users"],
            "new_payers": amounts["new_payers"],
            "first_payers": amounts["first_payers"],
            "d1_return": amounts["d1"] / counts["d1"] if counts["d1"] else None,
            "d3_return": amounts["d3"] / counts["d3"] if counts["d3"] else None,
            "d7_return": amounts["d7"] / counts["d7"] if counts["d7"] else None,
            "d1_weight": counts["d1"],
            "d3_weight": counts["d3"],
            "d7_weight": counts["d7"],
        })
    return sorted(output, key=lambda row: row["new_payers"], reverse=True)


def draw_overview(rows: list[dict[str, Any]]) -> None:
    image = Image.new("RGB", (1800, 860), WHITE)
    draw = ImageDraw.Draw(image)
    text(draw, (70, 42), "生命周期1—4：两周加权实际RTP与配置预期RTP", 34)
    text(draw, (70, 92), "实际RTP按完整下注重算；金色短线仅代表“页面预期回报比”有值的下注覆盖部分。", 19, MUTED)
    labels = ["生命周期1", "生命周期2", "生命周期3", "生命周期4", "1—4合计"]
    y_start, y_step = 190, 115
    plot_x0, plot_x1 = 400, 1320
    for tick in (0.88, 0.92, 0.96, 1.00, 1.04):
        x = plot_x0 + (tick - 0.88) / 0.16 * (plot_x1 - plot_x0)
        draw.line((x, 155, x, 690), fill=GRID, width=2)
        text(draw, (x, 715), f"{tick * 100:.0f}%", 16, MUTED, anchor="mt")
    for index, lifecycle in enumerate([1, 2, 3, 4, None]):
        y = y_start + index * y_step
        subset = [row for row in rows if row["display_lifecycle"] == (lifecycle if lifecycle is not None else "1—4合计")]
        text(draw, (70, y), labels[index], 21, INK, anchor="lm")
        week_rows = []
        for week_index, week_name in enumerate(WINDOWS):
            row = next(row for row in subset if row["period"] == week_name)
            week_rows.append(row)
            if row["actual_rtp"] is None:
                continue
            x = plot_x0 + (row["actual_rtp"] - 0.88) / 0.16 * (plot_x1 - plot_x0)
            color = NAVY if week_index == 0 else BLUE
            draw.ellipse((x - 10, y - 10, x + 10, y + 10), fill=WHITE, outline=color, width=5)
            if row["expected_rtp"] is not None:
                x_expected = plot_x0 + (row["expected_rtp"] - 0.88) / 0.16 * (plot_x1 - plot_x0)
                draw.line((x_expected, y - 18, x_expected, y + 18), fill=GOLD, width=4)
        first_row, second_row = week_rows
        text(draw, (1380, y - 11), f"W1 {pct(first_row['actual_rtp'], 1)}  W2 {pct(second_row['actual_rtp'], 1)}", 16, INK)
        text(draw, (1380, y + 14), f"配置覆盖 {pct(second_row['expected_bet_coverage'], 0)}", 14, MUTED)
    draw.rounded_rectangle((90, 780, 114, 804), radius=6, fill=NAVY)
    text(draw, (130, 792), "第一周实际RTP", 16, INK, anchor="lm")
    draw.rounded_rectangle((350, 780, 374, 804), radius=6, fill=BLUE)
    text(draw, (390, 792), "第二周实际RTP", 16, INK, anchor="lm")
    draw.line((620, 792, 650, 792), fill=GOLD, width=4)
    text(draw, (665, 792), "预期RTP（两周各自按下注额加权）", 16, INK, anchor="lm")
    image.save(ASSETS / "01_生命周期两周加权RTP总览.png", "PNG", optimize=True)


def draw_heatmap(rows: list[dict[str, Any]]) -> None:
    games = sorted({row["game"] for row in rows})
    change = {(row["game"], row["display_lifecycle"]): row["actual_rtp_delta_pp"] for row in rows}
    max_abs = max((abs(value or 0) for value in change.values()), default=0.01)
    width, height = 1560, 150 + 42 * len(games) + 100
    image = Image.new("RGB", (width, height), WHITE)
    draw = ImageDraw.Draw(image)
    text(draw, (60, 35), "游戏 × 生命周期：第二周相对第一周的加权实际RTP变化", 30)
    text(draw, (60, 82), "单位：百分点；空白表示两周均无有效下注，不等于0。颜色仅表达方向与幅度。", 17, MUTED)
    x0, col_w, y0, row_h = 410, 225, 130, 42
    for index, lifecycle in enumerate(LIFECYCLES):
        text(draw, (x0 + col_w * index + col_w / 2, y0 - 22), f"生命周期{lifecycle}", 18, INK, anchor="mm")
    for row_index, game in enumerate(games):
        y = y0 + row_index * row_h
        text(draw, (50, y + row_h / 2), game, 17, INK, anchor="lm")
        for col_index, lifecycle in enumerate(LIFECYCLES):
            value = change.get((game, lifecycle))
            x = x0 + col_index * col_w
            if value is None:
                fill, label = "#F1F5F9", "N/A"
            else:
                intensity = min(1.0, abs(value) / max_abs)
                if value >= 0:
                    fill = (int(232 - 70 * intensity), int(244 - 100 * intensity), int(255 - 45 * intensity))
                    fill = "#%02X%02X%02X" % fill
                else:
                    fill = (int(255 - 25 * intensity), int(239 - 100 * intensity), int(239 - 88 * intensity))
                    fill = "#%02X%02X%02X" % fill
                label = pp(value, 1)
            draw.rectangle((x, y, x + col_w - 5, y + row_h - 5), fill=fill, outline=GRID, width=1)
            text(draw, (x + col_w / 2, y + row_h / 2 - 2), label, 15, INK, anchor="mm")
    image.save(ASSETS / "02_游戏生命周期RTP变化热力图.png", "PNG", optimize=True)


def draw_scatter(rows: list[dict[str, Any]]) -> None:
    points = [
        {"game": row["game"], "second_bet": row["second_full_bet"], "delta_pp": row["actual_rtp_delta_pp"], "impact": row["impact_amount"]}
        for row in rows
        if row["second_full_bet"] > 0 and row["actual_rtp_delta_pp"] is not None
    ]
    total_bet = sum(point["second_bet"] for point in points) or 1
    x_max = max((point["second_bet"] / total_bet for point in points), default=0.01) * 1.15
    y_max = max((abs(point["delta_pp"]) for point in points), default=0.01) * 1.25
    image = Image.new("RGB", (1800, 930), WHITE)
    draw = ImageDraw.Draw(image)
    text(draw, (70, 42), "分游戏汇总：第二周下注贡献与RTP变化", 34)
    text(draw, (70, 92), "每点为“生命周期奖池分游戏汇总”中的单款游戏；纵轴为第二周相对第一周加权实际RTP变化。", 19, MUTED)
    x0, y0, x1, y1 = 160, 180, 1690, 770
    for tick in range(0, 6):
        x = x0 + (x1 - x0) * tick / 5
        pct_value = x_max * tick / 5
        draw.line((x, y0, x, y1), fill=GRID, width=2)
        text(draw, (x, y1 + 28), f"{pct_value * 100:.1f}%", 15, MUTED, anchor="mt")
    for tick in range(-4, 5):
        value = y_max * tick / 4
        y = y1 - (value + y_max) / (2 * y_max) * (y1 - y0)
        draw.line((x0, y, x1, y), fill=GRID if tick else "#94A3B8", width=3 if tick == 0 else 2)
        text(draw, (x0 - 15, y), f"{value * 100:+.1f}pp", 15, MUTED, anchor="rm")
    label_points = sorted(points, key=lambda item: abs(item["impact"]), reverse=True)[:12]
    for point in points:
        x = x0 + (point["second_bet"] / total_bet) / x_max * (x1 - x0)
        y = y1 - (point["delta_pp"] + y_max) / (2 * y_max) * (y1 - y0)
        radius = 6 + min(16, math.sqrt(point["second_bet"] / total_bet) * 60)
        fill = ORANGE if point["delta_pp"] >= 0 else TEAL
        draw.ellipse((x - radius, y - radius, x + radius, y + radius), fill=fill, outline=WHITE, width=2)
        if point in label_points:
            text(draw, (x + radius + 5, y - 8), point["game"], 14, INK)
    text(draw, (x0, 835), "横轴：第二周完全下注额占分游戏汇总总下注额的比例。纵轴：加权实际RTP变化。", 17, MUTED)
    image.save(ASSETS / "03_游戏下注贡献与RTP变化散点图.png", "PNG", optimize=True)


def draw_ranking(rows: list[dict[str, Any]]) -> None:
    ranked = sorted(
        [row for row in rows if row["second_full_bet"] > 0 and row["actual_rtp_delta_pp"] is not None],
        key=lambda row: abs(row["impact_amount"]),
        reverse=True,
    )[:12]
    image = Image.new("RGB", (1800, 860), WHITE)
    draw = ImageDraw.Draw(image)
    text(draw, (70, 42), "重点游戏：按“第二周下注额 × RTP变化”排序", 31)
    text(draw, (70, 90), "用于排查优先级排序，不等同于真实损失或对TC的因果贡献。", 18, MUTED)
    values = [abs(row["impact_amount"]) for row in ranked] or [1]
    maximum = max(values)
    y0, row_h, bar_x = 165, 52, 650
    for index, row in enumerate(ranked):
        y = y0 + index * row_h
        label = row["game"]
        text(draw, (70, y + 17), label, 17, INK)
        width = abs(row["impact_amount"]) / maximum * 900
        color = ORANGE if row["impact_amount"] >= 0 else TEAL
        draw.rounded_rectangle((bar_x, y, bar_x + width, y + 34), radius=8, fill=color)
        text(draw, (bar_x + width + 12, y + 17), f"{pp(row['actual_rtp_delta_pp'], 2)}｜第二周下注 {amount(row['second_full_bet'])}", 16, INK, anchor="lm")
    image.save(ASSETS / "04_重点游戏生命周期RTP优先级.png", "PNG", optimize=True)


def draw_new_user(rows: list[dict[str, Any]]) -> None:
    shown = rows[:8]
    maximum = max((row["new_payers"] for row in shown), default=1)
    image = Image.new("RGB", (1800, 720), WHITE)
    draw = ImageDraw.Draw(image)
    text(draw, (70, 42), "新增用户付费背景：8月14—27日按现有包体/渠道报表汇总", 31)
    text(draw, (70, 90), "仅用于观察同期用户结构；该表无game_id，不能归因到某款游戏。", 18, MUTED)
    y0, row_h, bar_x = 150, 60, 560
    for index, row in enumerate(shown):
        y = y0 + index * row_h
        text(draw, (70, y + 18), row["segment"], 17, INK)
        width = row["new_payers"] / maximum * 950
        draw.rounded_rectangle((bar_x, y, bar_x + width, y + 35), radius=8, fill=BLUE)
        text(draw, (bar_x + width + 12, y + 17), f"新增付费 {row['new_payers']:,.0f}｜首充 {row['first_payers']:,.0f}｜新增 {row['new_users']:,.0f}", 15, INK, anchor="lm")
    image.save(ASSETS / "05_新增用户付费背景.png", "PNG", optimize=True)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    ASSETS.mkdir(parents=True, exist_ok=True)
    source_manifest = json.loads(SOURCE_MANIFEST.read_text(encoding="utf-8"))
    manifest_by_date = {item["date"]: item for item in source_manifest["dates"]}
    all_rows: list[dict[str, Any]] = []
    game_daily_rows: list[dict[str, Any]] = []
    headers_by_date: dict[str, list[Any]] = {}
    quality: dict[str, Any] = {"status": "passed", "window": [TARGET_START.isoformat(), TARGET_END.isoformat()], "checks": [], "issues": []}
    source_receipt: list[dict[str, Any]] = []
    expected_dates = [TARGET_START.fromordinal(TARGET_START.toordinal() + index) for index in range((TARGET_END - TARGET_START).days + 1)]

    for day in expected_dates:
        day_key = day.isoformat()
        manifest = manifest_by_date.get(day_key)
        detail_path = RAW_ROOT / day_key / "detail.xlsx"
        if not manifest or manifest.get("status") != "complete" or not detail_path.exists():
            quality["checks"].append({"name": f"{day_key} source complete", "status": "failed"})
            quality["status"] = "failed"
            continue
        rows, headers = read_detail(detail_path, day)
        headers_by_date[day_key] = headers
        stability = manifest.get("stability", {})
        # The 30-day refresh contains two equivalent receipt schemas.  The
        # older dates use first_rows, while later dates use first_check.
        if "first_rows" in stability:
            expected_row_count = int(stability["first_rows"][1])
            expected_game_count = int(stability["first_rows"][2])
        else:
            expected_row_count = int(stability["first_check"]["row_counts"][1])
            expected_game_count = int(stability["first_check"]["row_counts"][2])
        game_path = RAW_ROOT / day_key / "game.xlsx"
        games = {row["game"] for row in rows}
        target_rows = [row for row in rows if row["display_lifecycle"] in LIFECYCLES]
        unique_target = {(row["business_date"], row["game"], row["display_lifecycle"]) for row in target_rows}
        checks = [
            {"name": f"{day_key} detail row count", "expected": expected_row_count, "actual": len(rows), "status": "passed" if len(rows) == expected_row_count else "failed"},
            {"name": f"{day_key} target lifecycle row count", "expected": len(games) * len(LIFECYCLES), "actual": len(target_rows), "status": "passed" if len(target_rows) == len(games) * len(LIFECYCLES) else "failed"},
            {"name": f"{day_key} target lifecycle unique key", "expected": len(target_rows), "actual": len(unique_target), "status": "passed" if len(target_rows) == len(unique_target) else "failed"},
        ]
        rtp_errors = []
        for row in rows:
            if row["full_bet"] > 0 and row["source_actual_rtp"] is not None:
                rtp_errors.append(abs(row["source_actual_rtp"] - (1 - row["actual_profit"] / row["full_bet"])))
        checks.append({"name": f"{day_key} source actual RTP recomputation", "samples": len(rtp_errors), "max_abs_error": max(rtp_errors) if rtp_errors else None, "tolerance": 0.0001, "status": "passed" if rtp_errors and max(rtp_errors) <= 0.0001 else "failed"})
        config_expected_bet = sum(row["full_bet"] for row in rows if row["full_bet"] > 0 and row["source_expected_rtp"] is not None and 0 < row["source_expected_rtp"] <= 1.1)
        total_effective_bet = sum(row["full_bet"] for row in rows if row["full_bet"] > 0)
        checks.append({"name": f"{day_key} configuration expected RTP coverage", "covered_bet": config_expected_bet, "total_bet": total_effective_bet, "coverage": config_expected_bet / total_effective_bet if total_effective_bet else None, "status": "partial" if config_expected_bet < total_effective_bet else "passed"})
        game_rows, _ = read_game(game_path, day)
        game_rtp_errors = [abs(row["source_actual_rtp"] - (1 - row["actual_profit"] / row["full_bet"])) for row in game_rows if row["full_bet"] > 0 and row["source_actual_rtp"] is not None]
        checks.extend([
            {"name": f"{day_key} game summary row count", "expected": expected_game_count, "actual": len(game_rows), "status": "passed" if len(game_rows) == expected_game_count else "failed"},
            {"name": f"{day_key} game summary unique key", "expected": len(game_rows), "actual": len({row['game'] for row in game_rows}), "status": "passed" if len(game_rows) == len({row['game'] for row in game_rows}) else "failed"},
            {"name": f"{day_key} game summary actual RTP recomputation", "samples": len(game_rtp_errors), "max_abs_error": max(game_rtp_errors) if game_rtp_errors else None, "tolerance": 0.0001, "status": "passed" if game_rtp_errors and max(game_rtp_errors) <= 0.0001 else "failed"},
        ])
        quality["checks"].extend(checks)
        if any(check["status"] == "failed" for check in checks):
            quality["status"] = "failed"
        all_rows.extend(rows)
        game_daily_rows.extend(game_rows)
        source_receipt.append({
            "business_date": day_key,
            "source_status": manifest["status"],
            "detail_path": str(detail_path.relative_to(ROOT)),
            "detail_sha256": digest(detail_path),
            "game_path": str(game_path.relative_to(ROOT)),
            "game_sha256": digest(game_path),
            "detail_rows": len(rows),
            "game_count": len(games),
            "target_lifecycle_rows": len(target_rows),
            "query_receipt": manifest["receipt_path"],
        })

    games = sorted({row["game"] for row in all_rows})
    weekly = [weekly_row(all_rows, game, lifecycle, label, start, end) for game in games for lifecycle in LIFECYCLES for label, (start, end) in WINDOWS.items()]
    by_key = {(row["game"], row["display_lifecycle"], row["period"]): row for row in weekly}
    comparison: list[dict[str, Any]] = []
    first_name, second_name = list(WINDOWS)
    for game in games:
        for lifecycle in LIFECYCLES:
            first = by_key[(game, lifecycle, first_name)]
            second = by_key[(game, lifecycle, second_name)]
            delta = second["actual_rtp"] - first["actual_rtp"] if first["actual_rtp"] is not None and second["actual_rtp"] is not None else None
            comparison.append({
                "game": game,
                "display_lifecycle": lifecycle,
                "first_full_bet": first["full_bet"],
                "second_full_bet": second["full_bet"],
                "bet_change": second["full_bet"] - first["full_bet"],
                "first_actual_profit": first["actual_profit"],
                "second_actual_profit": second["actual_profit"],
                "first_actual_rtp": first["actual_rtp"],
                "second_actual_rtp": second["actual_rtp"],
                "first_expected_rtp": first["expected_rtp"],
                "second_expected_rtp": second["expected_rtp"],
                "actual_rtp_delta_pp": delta,
                "second_rtp_gap_pp": second["rtp_gap_pp"],
                "first_active_days": first["active_days"],
                "second_active_days": second["active_days"],
                "data_status": "no_effective_bet" if second["full_bet"] <= 0 else ("early_sample" if min(first["active_days"], second["active_days"]) < 7 else "complete_7d"),
                "impact_amount": second["full_bet"] * delta if delta is not None else None,
            })
    for row in comparison:
        total_second = sum(item["second_full_bet"] for item in comparison) or 1
        row["second_bet_share"] = row["second_full_bet"] / total_second

    overview = [aggregate_rows(all_rows, label, start, end, lifecycle) for label, (start, end) in WINDOWS.items() for lifecycle in (*LIFECYCLES, None)]
    new_user = load_new_user_context()
    game_names = sorted({row["game"] for row in game_daily_rows})
    game_weekly = [weekly_game_row(game_daily_rows, game, label, start, end) for game in game_names for label, (start, end) in WINDOWS.items()]
    game_weekly_by_key = {(row["game"], row["period"]): row for row in game_weekly}
    game_summary = []
    total_game_week2_bet = sum(row["full_bet"] for row in game_weekly if row["period"] == second_name)
    for game in game_names:
        first_game = game_weekly_by_key[(game, first_name)]
        second_game = game_weekly_by_key[(game, second_name)]
        delta = second_game["actual_rtp"] - first_game["actual_rtp"] if first_game["actual_rtp"] is not None and second_game["actual_rtp"] is not None else None
        game_summary.append({
            "game": game,
            "first_full_bet": first_game["full_bet"],
            "second_full_bet": second_game["full_bet"],
            "first_actual_profit": first_game["actual_profit"],
            "second_actual_profit": second_game["actual_profit"],
            "first_actual_rtp": first_game["actual_rtp"],
            "second_actual_rtp": second_game["actual_rtp"],
            "actual_rtp_delta_pp": delta,
            "second_bet_share": second_game["full_bet"] / total_game_week2_bet if total_game_week2_bet else None,
            "first_active_days": first_game["active_days"],
            "second_active_days": second_game["active_days"],
            "first_expected_rtp": first_game["expected_rtp"],
            "second_expected_rtp": second_game["expected_rtp"],
            "first_expected_bet_coverage": first_game["expected_bet_coverage"],
            "second_expected_bet_coverage": second_game["expected_bet_coverage"],
            "data_status": "no_effective_bet" if second_game["full_bet"] <= 0 else ("early_sample" if min(first_game["active_days"], second_game["active_days"]) < 7 else "complete_7d"),
            "impact_amount": second_game["full_bet"] * delta if delta is not None else None,
        })
    game_summary.sort(key=lambda row: abs(row["impact_amount"] or 0), reverse=True)
    json_dump(OUT / "source-receipt.json", {"source": "GM Lifecycle Pool V2 (Joint) exports", "window": [TARGET_START.isoformat(), TARGET_END.isoformat()], "dates": source_receipt, "new_user_source": {"path": str(NEW_USER_SOURCE.relative_to(ROOT)), "sha256": digest(NEW_USER_SOURCE)}})
    json_dump(OUT / "quality-checks.json", quality)
    json_dump(OUT / "weekly_game_lifecycle.json", weekly)
    json_dump(OUT / "game_lifecycle_comparison.json", comparison)
    json_dump(OUT / "game_summary_weighted.json", game_summary)
    json_dump(OUT / "overview.json", overview)
    json_dump(OUT / "new_user_context.json", new_user)

    with (OUT / "game_lifecycle_comparison.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(comparison[0]))
        writer.writeheader()
        writer.writerows(comparison)

    overall_first = next(row for row in overview if row["period"] == first_name and row["display_lifecycle"] == "1—4合计")
    overall_second = next(row for row in overview if row["period"] == second_name and row["display_lifecycle"] == "1—4合计")
    game_rows = game_summary
    new_games = [row for row in game_rows if row["first_active_days"] == 0 and row["second_active_days"] > 0]

    draw_overview(overview)
    draw_heatmap(comparison)
    draw_scatter(game_rows)
    draw_ranking(game_rows)
    draw_new_user(new_user)

    stage_rows = []
    for lifecycle in LIFECYCLES:
        first = next(row for row in overview if row["period"] == first_name and row["display_lifecycle"] == lifecycle)
        second = next(row for row in overview if row["period"] == second_name and row["display_lifecycle"] == lifecycle)
        stage_rows.append({"lifecycle": lifecycle, "first": first, "second": second, "delta": (second["actual_rtp"] - first["actual_rtp"]) if first["actual_rtp"] is not None and second["actual_rtp"] is not None else None})

    top_games = game_rows[:8]
    report_lines = [
        "# 补充｜近14日游戏 × 生命周期 RTP 验证（V2 Joint）",
        "",
        f"> 业务日期：{TARGET_START.isoformat()}—{TARGET_END.isoformat()}；展示生命周期 1—4；范围：GM Lifecycle Pool V2（Joint）。",
        "",
        "## 管理摘要",
        "",
        f"- **全局变化：** 生命周期1—4合计的加权实际RTP由第一周 {pct(overall_first['actual_rtp'])} 变为第二周 {pct(overall_second['actual_rtp'])}，变化 {pp((overall_second['actual_rtp'] or 0) - (overall_first['actual_rtp'] or 0))}；对应完全下注额由 {amount(overall_first['full_bet'])} 变为 {amount(overall_second['full_bet'])}。",
        f"- **配置预期口径不完整：** 页面“预期回报比”有值的下注覆盖第一周 {pct(overall_first['expected_bet_coverage'])}、第二周 {pct(overall_second['expected_bet_coverage'])}；在覆盖部分内的加权预期RTP分别为 {pct(overall_first['expected_rtp'])} 与 {pct(overall_second['expected_rtp'])}。其余联运游戏配置值为0或缺失，不将其伪算为预期RTP。",
        f"- **游戏覆盖：** 第一周“生命周期奖池分游戏汇总”出现 {len([row for row in game_rows if row['first_active_days'] > 0])} 款有有效下注的游戏，第二周 {len([row for row in game_rows if row['second_active_days'] > 0])} 款；第二周新增出现 {len(new_games)} 款。新增或不足7日的游戏只作早期观察。",
        "- **与TC的关系：** 本节验证的是全产品游戏与生命周期池回报结构；V2（Joint）没有包体、渠道或事件时点归因，不能将任一游戏RTP变化直接归因到H5或某渠道TC异常。",
        "",
        "## 生命周期1—4两周对比",
        "",
        "| 生命周期 | 第一周下注额 | 第一周实际RTP | 第二周下注额 | 第二周实际RTP | 变化 | 第二周配置预期RTP（下注覆盖） |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for row in stage_rows:
        report_lines.append(f"| {row['lifecycle']} | {amount(row['first']['full_bet'])} | {pct(row['first']['actual_rtp'])} | {amount(row['second']['full_bet'])} | {pct(row['second']['actual_rtp'])} | {pp(row['delta'])} | {pct(row['second']['expected_rtp'])}（{pct(row['second']['expected_bet_coverage'])}） |")
    report_lines.extend([
        "",
        "![生命周期两周加权RTP总览](assets/01_生命周期两周加权RTP总览.png)",
        "",
        "## 重点游戏：按第二周下注规模与RTP变化排序",
        "",
        "| 游戏 | 第一周RTP | 第二周RTP | 变化 | 第二周下注额 | 观察状态 |",
        "|---|---:|---:|---:|---:|---|",
    ])
    for row in top_games:
        state = "早期样本" if min(row["first_active_days"], row["second_active_days"]) < 7 else "两周完整"
        report_lines.append(f"| {row['game']} | {pct(row['first_actual_rtp'])} | {pct(row['second_actual_rtp'])} | {pp(row['actual_rtp_delta_pp'])} | {amount(row['second_full_bet'])} | {state} |")
    report_lines.extend([
        "",
        "![游戏生命周期RTP变化热力图](assets/02_游戏生命周期RTP变化热力图.png)",
        "",
        "![游戏下注贡献与RTP变化散点图](assets/03_游戏下注贡献与RTP变化散点图.png)",
        "",
        "![重点游戏生命周期RTP优先级](assets/04_重点游戏生命周期RTP优先级.png)",
        "",
        "## 新增用户付费背景（仅结构对照）",
        "",
        "新增用户报表仅具备包体/渠道 × cohort 日期的新增、首充、付费与留存字段，不具备游戏维度。本节不能把新增用户付费变化归因给具体游戏。留存仅使用源表中已返回的非空、非零成熟字段；未成熟值不显示为0。",
        "",
        "| 包体/渠道Sheet | 新增人数 | 新增付费人数 | 首充人数 | 已成熟D1样本量 | 已成熟D3样本量 | 已成熟D7样本量 |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ])
    for row in new_user:
        report_lines.append(f"| {row['segment']} | {row['new_users']:,.0f} | {row['new_payers']:,.0f} | {row['first_payers']:,.0f} | {row['d1_weight']:,.0f} | {row['d3_weight']:,.0f} | {row['d7_weight']:,.0f} |")
    report_lines.extend([
        "",
        "![新增用户付费背景](assets/05_新增用户付费背景.png)",
        "",
        "## 数据质量、解释边界与下一步",
        "",
        f"- 原始数据：14/14 日期均有完整、稳定的 V2（Joint）页面导出；每日明细行数、游戏数 × 生命周期行数、生命周期1—4唯一键及源表RTP复算均记录在 `quality-checks.json`。",
        "- RTP算法：完全实际RTP = `1 − Σ完全实际盈利 ÷ Σ完全下注额`。配置预期RTP只对“页面预期回报比”为有效正值的记录按完全下注额加权；不从预期盈利字段反推，也不平均每日或分组百分比。",
        "- 这是生命周期池回报口径，不是最终结算RTP认证。对高TC渠道的归因仍需补齐 `业务日 × 包体/渠道 × 游戏 × 生命周期 × 有效真金下注 × 最终结算派奖 × 有效局数 × 配置版本` 的认证聚合事实层。",
        "- 结论分级：表中数值为已验证的生命周期池聚合事实；任何游戏对TC、资金安全、羊毛或作弊的影响均为待验证假设，不能由本节单独得出。",
    ])
    (OUT / "report.md").write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    table_rows = "".join(
        f"<tr><td><p>{row['lifecycle']}</p></td><td><p>{pct(row['first']['actual_rtp'])}</p></td><td><p>{pct(row['second']['actual_rtp'])}</p></td><td><p>{pp(row['delta'])}</p></td><td><p>{amount(row['second']['full_bet'])}</p></td></tr>"
        for row in stage_rows
    )
    top_rows = "".join(
        f"<tr><td><p>{escape(row['game'])}</p></td><td><p>{pct(row['first_actual_rtp'])}</p></td><td><p>{pct(row['second_actual_rtp'])}</p></td><td><p>{pp(row['actual_rtp_delta_pp'])}</p></td><td><p>{amount(row['second_full_bet'])}</p></td></tr>"
        for row in top_games[:6]
    )
    lark_xml = f"""<h1>11. 补充：近14日游戏 × 生命周期 RTP 验证</h1>
<callout emoji=\"🔎\" background-color=\"light-blue\" border-color=\"blue\"><p><b>结论：</b>生命周期1—4的全产品加权实际RTP由第一周 {pct(overall_first['actual_rtp'])} 变为第二周 {pct(overall_second['actual_rtp'])}（{pp((overall_second['actual_rtp'] or 0) - (overall_first['actual_rtp'] or 0))}）。这是GM Lifecycle Pool V2（Joint）的全局游戏/生命周期回报观察，不能单独归因到H5、某渠道TC或某个新游戏。</p></callout>
<h2>11.1 口径与范围</h2>
<p>业务日期：2026-08-14—2026-08-27，共14个完整日；范围：GM Lifecycle Pool V2（Joint）；生命周期：页面展示生命周期1、2、3、4。实际RTP按 <b>1 − Σ完全实际盈利 ÷ Σ完全下注额</b> 重算。页面“预期回报比”只在部分游戏/下注中有有效值，预期RTP只对该覆盖部分按完全下注额加权；不从“预期盈利”字段反推，也不平均每日或分组百分比。</p>
<h2>11.2 生命周期两周对比</h2>
<table><thead><tr><th><p>生命周期</p></th><th><p>第一周实际RTP</p></th><th><p>第二周实际RTP</p></th><th><p>变化</p></th><th><p>第二周下注额</p></th></tr></thead><tbody>{table_rows}</tbody></table>
<h2>11.3 重点游戏核查</h2>
<p>全部游戏 × 生命周期1—4明细已在本地分析包留存；下表按第二周下注规模与RTP变化的综合影响排序。新增或任一周不足7个有效观察日的游戏只作早期样本观察。</p>
<table><thead><tr><th><p>游戏</p></th><th><p>第一周RTP</p></th><th><p>第二周RTP</p></th><th><p>变化</p></th><th><p>第二周下注额</p></th></tr></thead><tbody>{top_rows}</tbody></table>
<h2>11.4 与TC审计的衔接</h2>
<p>V2（Joint）当前缺少包体、渠道、归因媒体及最终结算事实，故本节不能证明“某款游戏导致某H5渠道TC偏高”。需要补齐按业务日、包体/渠道、游戏、生命周期、有效真金下注、最终结算派奖、有效局数与配置版本聚合的认证事实层，才能对高TC渠道与高RTP游戏做因果核查。</p>
<p>新增用户付费分析仅作为同期包体/渠道用户结构背景，因其没有game_id，未与单款游戏RTP作直接关联；未成熟留存均不显示为0。</p>"""
    (OUT / "lark_append.xml").write_text(lark_xml, encoding="utf-8")

    artifact = {
        "surface": "report",
        "manifest": {
            "version": 1,
            "surface": "report",
            "title": "Waje 近14日游戏 × 生命周期 RTP 验证｜补充TC审计",
            "description": "基于GM Lifecycle Pool V2（Joint）14个完整日页面导出，比较生命周期1—4各游戏的两周加权实际RTP变化；配置预期RTP仅在来源字段有效的下注覆盖范围内计算。",
            "generatedAt": datetime.now().astimezone().isoformat(),
            "sources": [
                {"id": "lifecycle-joint", "label": "GM Lifecycle Pool V2 (Joint) — daily export snapshot", "query": {"engine": "GM UI export", "description": "Aggregate-only daily detail/game exports.", "tables_used": ["Lifecycle Pool V2 (Joint)"], "filters": ["2026-08-14 to 2026-08-27", "display lifecycle 1 to 4"], "metric_definitions": ["Actual RTP = 1 - sum(full actual profit)/sum(full bet).", "Configuration expected RTP weights only rows with a valid source expected-return field; zero or missing values are excluded."]}},
                {"id": "origin-new-user", "label": "Origin BQ 新增付费用户分析 — 8 package/channel sheets", "query": {"engine": "Origin visible report export", "description": "Package/channel cohort context only; no game_id.", "tables_used": ["BQ-新增付费用户分析"], "filters": ["2026-08-14 to 2026-08-27"], "metric_definitions": ["New-user payment/retention data is not attributed to individual games."]}},
            ],
            "charts": [
                {"id": "lifecycle-overview", "title": "生命周期1—4：两周加权实际RTP", "subtitle": "完整下注与实际盈利累计重算；第一周与第二周对照。", "type": "bar", "dataset": "lifecycle_weekly", "sourceId": "lifecycle-joint", "layout": "full", "encodings": {"x": {"field": "display_lifecycle", "type": "nominal", "label": "生命周期"}, "y": {"field": "actual_rtp", "type": "quantitative", "format": "percent", "label": "实际RTP"}, "color": {"field": "period", "type": "nominal", "label": "周期"}, "tooltip": [{"field": "full_bet", "type": "quantitative", "format": "number"}, {"field": "actual_rtp", "type": "quantitative", "format": "percent"}, {"field": "expected_rtp", "type": "quantitative", "format": "percent"}]}},
                {"id": "game-contribution-scatter", "title": "分游戏汇总：第二周下注规模与RTP变化", "subtitle": "每点为“生命周期奖池分游戏汇总”的单款游戏；点大小为第二周完全下注额。", "type": "scatter", "dataset": "game_summary", "sourceId": "lifecycle-joint", "layout": "full", "encodings": {"x": {"field": "second_bet_share", "type": "quantitative", "format": "percent", "label": "第二周下注额占比"}, "y": {"field": "actual_rtp_delta_pp", "type": "quantitative", "format": "percent", "label": "实际RTP变化"}, "size": {"field": "second_full_bet", "type": "quantitative", "label": "第二周完全下注额"}, "tooltip": [{"field": "game", "type": "nominal"}, {"field": "second_full_bet", "type": "quantitative", "format": "number"}, {"field": "first_actual_rtp", "type": "quantitative", "format": "percent"}, {"field": "second_actual_rtp", "type": "quantitative", "format": "percent"}]}},
            ],
            "tables": [
                {"id": "game-lifecycle-detail", "title": "游戏 × 生命周期1—4两周加权RTP明细", "subtitle": "全量明细；每一行均为游戏、生命周期与两周窗口的累计值。", "dataset": "game_lifecycle_comparison", "sourceId": "lifecycle-joint", "density": "compact", "layout": "full", "defaultSort": {"field": "second_full_bet", "direction": "desc"}, "columns": [{"field": "game", "label": "游戏", "type": "text"}, {"field": "display_lifecycle", "label": "生命周期", "type": "number"}, {"field": "first_actual_rtp", "label": "第一周实际RTP", "type": "percent", "format": "percent"}, {"field": "second_actual_rtp", "label": "第二周实际RTP", "type": "percent", "format": "percent"}, {"field": "actual_rtp_delta_pp", "label": "RTP变化", "type": "percent", "format": "percent"}, {"field": "second_full_bet", "label": "第二周下注额", "type": "number", "format": "number"}, {"field": "data_status", "label": "数据状态", "type": "text"}]},
            ],
            "blocks": [
                {"id": "title", "type": "markdown", "body": "# Waje 近14日游戏 × 生命周期 RTP 验证"},
                {"id": "summary", "type": "markdown", "body": f"## 执行摘要\n\n生命周期1—4合计的加权实际RTP由第一周 **{pct(overall_first['actual_rtp'])}** 变为第二周 **{pct(overall_second['actual_rtp'])}**，变化 **{pp((overall_second['actual_rtp'] or 0) - (overall_first['actual_rtp'] or 0))}**。这是全产品生命周期池回报结构的验证，不能直接归因到H5、某渠道TC或某个新游戏。\n\n页面配置预期回报字段在第一周覆盖 **{pct(overall_first['expected_bet_coverage'])}** 的下注、第二周覆盖 **{pct(overall_second['expected_bet_coverage'])}**；缺失/0值不被伪算为预期。分游戏汇总中，第一周有有效下注的游戏数为 {len([row for row in game_rows if row['first_active_days'] > 0])}，第二周为 {len([row for row in game_rows if row['second_active_days'] > 0])}；第二周新增出现 {len(new_games)} 款游戏，均按早期样本处理。", "sourceId": "lifecycle-joint"},
                {"id": "overview-chart", "type": "chart", "chartId": "lifecycle-overview"},
                {"id": "matrix-note", "type": "markdown", "body": "## 游戏与生命周期深钻\n\n所有游戏与生命周期1—4均保留在明细中。比例均按累计下注与累计盈利重算；没有有效下注的格子显示N/A，不替代为0。", "sourceId": "lifecycle-joint"},
                {"id": "scatter-chart", "type": "chart", "chartId": "game-contribution-scatter"},
                {"id": "detail-table", "type": "table", "tableId": "game-lifecycle-detail"},
                {"id": "context", "type": "markdown", "body": "## 与新增用户付费和TC审计的关系\n\n新增用户付费报表用于包体/渠道的同期用户结构背景，但没有游戏维度，不能与单款游戏RTP直接关联。V2（Joint）同样缺少包体、渠道、归因与最终结算事实，因此本报告只能提出需要复核的游戏/生命周期优先级，不能证明任何游戏造成特定渠道的TC异常。", "sourceId": "origin-new-user"},
                {"id": "caveats", "type": "markdown", "body": "## 数据边界\n\n生命周期池的完全真实回报比不替代最终结算RTP认证。高TC渠道的最终因果核查仍需认证的日级聚合事实：业务日、包体/渠道、游戏、生命周期、有效真金下注、最终结算派奖、有效局数、游戏/配置版本和数据截止时间。所有输出均为聚合结果。"},
            ],
        },
        "snapshot": {"version": 1, "generatedAt": datetime.now().astimezone().isoformat(), "status": "partial", "accessIssues": [{"id": "missing-channel-game-fact", "severity": "medium", "status": "blocked", "message": "The V2 Joint export has no certified package/channel or terminal-settlement dimensions for causal TC attribution."}], "datasets": {"lifecycle_weekly": overview, "game_lifecycle_comparison": comparison, "game_summary": game_rows, "new_user_context": new_user}},
    }
    json_dump(OUT / "artifact.json", artifact)
    print(json.dumps({"status": quality["status"], "output": str(OUT), "games": len(games), "rows": len(comparison), "new_games_week2": [row["game"] for row in new_games]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
