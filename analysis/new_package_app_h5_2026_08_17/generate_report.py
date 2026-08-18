#!/usr/bin/env python3
"""Generate the standalone Waje new-package App/H5 cohort analysis report.

The source workbook is read-only. All calculations and chart labels are built
from the same normalized in-memory cohort records to keep the HTML auditable.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from html import escape
from math import pi, sqrt
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SOURCE = Path("/Users/robin/Desktop/waje data/新包分析2026.8.11_new.xlsx")
OUTPUT = Path(
    "/Users/robin/Documents/wajetan_analyst/output/html/"
    "新包APP与H5渠道包体留存付费深度分析-2026-08-17.html"
)
REPORT_DATE = date(2026, 8, 17)

SHEET_MAP: dict[str, tuple[str, str, str]] = {
    "WajeSpecial-facebook": ("App", "Waje Special Android", "Facebook"),
    "WajeSpecial-googleadwords_int": ("App", "Waje Special Android", "Google Ads"),
    "WajeSpecial-Google商店": ("App", "Waje Special Android", "Google Play"),
    "PAWAJEIOS-AppStore商店": ("App", "PA Waje iOS", "App Store"),
    "PAWAJEBETH5": ("H5", "WajeBet H5", "未标注/包体汇总"),
    "pww": ("PWW", "Waje H5 PWW", "未标注/包体汇总"),
    "wajeH5-fb": ("H5", "Waje H5", "Facebook"),
    "wajeH5ga-googlewors_int": ("H5", "Waje H5", "Google Ads"),
}

RETENTION_FIELDS = {
    "D1": "次留",
    "D3": "3日留",
    "D7": "7日留",
    "D15": "15日留",
    "D30": "30日留",
    "D60": "60日留",
}

APP_COLOR = "#2E83C5"
H5_COLOR = "#4DAA80"
PWW_COLOR = "#D4A72C"
YELLOW = "#E6B63E"
RED = "#D86767"
TEXT = "#17324D"
MUTED = "#61778A"
GRID = "#D9E7F0"

PLATFORM_COLORS = {"App": APP_COLOR, "H5": H5_COLOR, "PWW": PWW_COLOR}
PLATFORM_LABELS = {"App": "App", "H5": "标准 H5", "PWW": "PWW", "全部": "全部"}


def n(value: Any) -> float | None:
    """Return a numeric source value or None, preserving blank source cells."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_date(value: Any, epoch: datetime) -> tuple[date | None, bool]:
    """Read native/text/Excel-serial dates and flag repaired serial values."""
    if isinstance(value, datetime):
        return value.date(), False
    if isinstance(value, date):
        return value, False
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date(), False
        except ValueError:
            pass
    try:
        serial = float(text)
        if 20000 < serial < 100000:
            converted = from_excel(serial, epoch)
            return (converted.date() if isinstance(converted, datetime) else converted), True
    except (TypeError, ValueError):
        pass
    return None, False


def pct(value: float | None, decimals: int = 1) -> str:
    return "N/A" if value is None else f"{value * 100:.{decimals}f}%"


def signed_pp(value: float | None, decimals: int = 1) -> str:
    return "N/A" if value is None else f"{value * 100:+.{decimals}f}pp"


def signed_pct(value: float | None, decimals: int = 1) -> str:
    return "N/A" if value is None else f"{value * 100:+.{decimals}f}%"


def integer(value: float | int | None) -> str:
    return "N/A" if value is None else f"{value:,.0f}"


def short_int(value: float | int | None) -> str:
    if value is None:
        return "N/A"
    if abs(value) >= 1_000_000:
        return f"{value / 1_000_000:.2f}M"
    if abs(value) >= 1_000:
        return f"{value / 1_000:.1f}k"
    return f"{value:.0f}"


def date_label(value: date) -> str:
    return f"{value.month}/{value.day}"


def safe(value: Any) -> str:
    return escape(str(value), quote=True)


def fmt_range(start: date, end: date) -> str:
    return f"{start:%Y-%m-%d} 至 {end:%Y-%m-%d}"


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def color_for_delta(value: float | None, positive_good: bool = True) -> str:
    if value is None:
        return "neutral"
    good = value >= 0 if positive_good else value <= 0
    return "positive" if good else "negative"


def load_records() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    serial_repairs = 0
    invalid_dates: list[tuple[str, int, Any]] = []
    sheet_quality: dict[str, dict[str, Any]] = {}

    for worksheet in workbook.worksheets:
        if worksheet.title not in SHEET_MAP:
            raise ValueError(f"未配置工作表映射：{worksheet.title}")
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        platform, package, channel = SHEET_MAP[worksheet.title]
        count = 0
        dates: list[date] = []
        blanks = 0
        for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            raw = values[0] if values else None
            if raw is None or not str(raw).strip():
                blanks += 1
                continue
            cohort_date, repaired = parse_date(raw, workbook.epoch)
            if cohort_date is None:
                invalid_dates.append((worksheet.title, row_index, raw))
                continue
            serial_repairs += int(repaired)
            record = dict(zip(headers, values))
            record.update(
                source_sheet=worksheet.title,
                cohort_date=cohort_date,
                platform=platform,
                package=package,
                channel=channel,
            )
            records.append(record)
            count += 1
            dates.append(cohort_date)
        sheet_quality[worksheet.title] = {
            "records": count,
            "blanks": blanks,
            "min": min(dates),
            "max": max(dates),
            "missing_days": (max(dates) - min(dates)).days + 1 - len(set(dates)),
        }

    return records, {
        "serial_repairs": serial_repairs,
        "invalid_dates": invalid_dates,
        "sheet_quality": sheet_quality,
        "max_date": max(record["cohort_date"] for record in records),
        "server_values": sorted({str(record.get("区服") or "") for record in records}),
    }


def filtered(records: Iterable[dict[str, Any]], start: date, end: date) -> list[dict[str, Any]]:
    return [record for record in records if start <= record["cohort_date"] <= end]


def compute_metrics(records: list[dict[str, Any]]) -> dict[str, Any]:
    new_users = sum(n(record.get("新增人数")) or 0 for record in records)
    result: dict[str, Any] = {"records": len(records), "new_users": new_users}

    for name, field in RETENTION_FIELDS.items():
        valid = [record for record in records if n(record.get(field)) is not None]
        denominator = sum(n(record.get("新增人数")) or 0 for record in valid)
        numerator = sum((n(record.get(field)) or 0) * (n(record.get("新增人数")) or 0) for record in valid)
        result[name.lower()] = numerator / denominator if denominator else None
        result[f"{name.lower()}_denominator"] = denominator

    for prefix, rate_field, count_field in (
        ("payment", "新增付费率", "新增付费人数"),
        ("first_charge", "首充付费率", "首充付费人数"),
    ):
        payer_count = sum(n(record.get(count_field)) or 0 for record in records)
        inferred_denominators = []
        for record in records:
            rate = n(record.get(rate_field))
            count = n(record.get(count_field))
            if rate is not None and rate > 0 and count is not None:
                inferred_denominators.append(count / rate)
        inferred_denominator = sum(inferred_denominators)
        result[f"{prefix}_count"] = payer_count
        result[f"{prefix}_per_new"] = payer_count / new_users if new_users else None
        result[f"{prefix}_source_rate"] = payer_count / inferred_denominator if inferred_denominator else None
        result[f"{prefix}_coverage"] = inferred_denominator / new_users if new_users else None
        result[f"{prefix}_inferred_denominator"] = inferred_denominator
    return result


def grouped_metrics(records: list[dict[str, Any]], key: str, start: date, end: date) -> dict[str, dict[str, Any]]:
    subset = filtered(records, start, end)
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in subset:
        groups[str(record[key])].append(record)
    return {group: compute_metrics(group_records) for group, group_records in groups.items()}


def svg_chart(title: str, description: str, width: int, height: int, body: str, classes: str = "") -> str:
    return (
        f'<figure class="chart {classes}"><figcaption><strong>{safe(title)}</strong>'
        f'<span>{safe(description)}</span></figcaption>'
        f'<svg viewBox="0 0 {width} {height}" role="img" aria-label="{safe(title)}：{safe(description)}" '
        f'preserveAspectRatio="xMidYMid meet">{body}</svg></figure>'
    )


def svg_text(x: float, y: float, text: str, cls: str = "label", anchor: str = "start") -> str:
    return f'<text x="{x:.1f}" y="{y:.1f}" text-anchor="{anchor}" class="{cls}">{safe(text)}</text>'


def retention_path_chart(platform_metrics: dict[str, dict[str, Any]], window_label: str) -> str:
    stages = [("D1", "d1"), ("D3", "d3"), ("D7", "d7"), ("D15", "d15")]
    width, height = 740, 292
    left, right, top, bottom = 58, 24, 24, 48
    max_y = max((platform_metrics[p][field] or 0 for p in platform_metrics for _, field in stages), default=0.6)
    max_y = max(0.6, min(0.7, (int(max_y * 10 + 1) / 10)))
    plot_w, plot_h = width - left - right, height - top - bottom
    body = [f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" class="plot-frame"/>']
    for tick in (0, 0.2, 0.4, 0.6):
        if tick > max_y + 0.001:
            continue
        y = top + plot_h * (1 - tick / max_y)
        body.append(f'<line x1="{left}" x2="{left + plot_w}" y1="{y:.1f}" y2="{y:.1f}" class="gridline"/>')
        body.append(svg_text(left - 10, y + 4, pct(tick, 0), "axis", "end"))
    for index, (label, _) in enumerate(stages):
        x = left + plot_w * index / (len(stages) - 1)
        body.append(svg_text(x, height - 18, label, "axis", "middle"))
    for platform, color in PLATFORM_COLORS.items():
        points = []
        for index, (_, field) in enumerate(stages):
            value = platform_metrics[platform][field] or 0
            x = left + plot_w * index / (len(stages) - 1)
            y = top + plot_h * (1 - value / max_y)
            points.append((x, y, value))
        path = " ".join((f"M {points[0][0]:.1f} {points[0][1]:.1f}",) + tuple(f"L {x:.1f} {y:.1f}" for x, y, _ in points[1:]))
        body.append(f'<path d="{path}" fill="none" stroke="{color}" class="series-line"/>')
        for x, y, value in points:
            body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="5.5" fill="{color}"/>')
            body.append(svg_text(x, y - 11, pct(value), "value", "middle"))
    body.append(f'<rect x="{left}" y="{height - 42}" width="12" height="12" rx="3" fill="{APP_COLOR}"/>')
    body.append(svg_text(left + 18, height - 32, "App", "legend"))
    body.append(f'<rect x="{left + 78}" y="{height - 42}" width="12" height="12" rx="3" fill="{H5_COLOR}"/>')
    body.append(svg_text(left + 96, height - 32, "标准 H5", "legend"))
    body.append(f'<rect x="{left + 146}" y="{height - 42}" width="12" height="12" rx="3" fill="{PWW_COLOR}"/>')
    body.append(svg_text(left + 164, height - 32, "PWW", "legend"))
    return svg_chart("App、标准 H5 与 PWW 的前 15 日留存路径", f"同一批成熟 D15 cohort，{window_label}；留存按新增人数加权。标准 H5 不含 PWW；源表未提供 D14 留存，使用 D15 作为第二周后观测点。", width, height, "".join(body))


def decay_chart(platform_metrics: dict[str, dict[str, Any]], window_label: str) -> str:
    stages = (("D1→D3", "d1", "d3"), ("D3→D7", "d3", "d7"), ("D7→D15", "d7", "d15"))
    values: dict[str, list[float]] = {}
    for platform in ("App", "H5", "PWW"):
        metrics = platform_metrics[platform]
        values[platform] = [
            1 - metrics[next_field] / metrics[base_field]
            for _, base_field, next_field in stages
        ]
    width, height = 760, 300
    left, right, top, bottom = 58, 24, 26, 64
    plot_w, plot_h = width - left - right, height - top - bottom
    raw_max = max(value for platform_values in values.values() for value in platform_values)
    max_y = max(0.6, (int(raw_max * 10) + 1) / 10)
    body = [f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" class="plot-frame"/>']
    for fraction in (0, 0.25, 0.5, 0.75, 1):
        y = top + plot_h * (1 - fraction)
        body.append(f'<line x1="{left}" x2="{left + plot_w}" y1="{y:.1f}" y2="{y:.1f}" class="gridline"/>')
        body.append(svg_text(left - 10, y + 4, pct(max_y * fraction, 0), "axis", "end"))
    group_w = plot_w / len(stages)
    bar_w = min(34, group_w * 0.18)
    gap = 7
    for stage_index, (stage_label, _, _) in enumerate(stages):
        center = left + group_w * (stage_index + 0.5)
        total_w = bar_w * 3 + gap * 2
        for platform_index, platform in enumerate(("App", "H5", "PWW")):
            value = values[platform][stage_index]
            x = center - total_w / 2 + platform_index * (bar_w + gap)
            bar_h = plot_h * value / max_y
            y = top + plot_h - bar_h
            body.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{bar_h:.1f}" rx="4" fill="{PLATFORM_COLORS[platform]}"/>')
            body.append(svg_text(x + bar_w / 2, y - 7, pct(value), "value", "middle"))
        body.append(svg_text(center, height - 32, stage_label, "axis", "middle"))
    legend_x = left
    for platform in ("App", "H5", "PWW"):
        body.append(f'<rect x="{legend_x}" y="{height - 53}" width="12" height="12" rx="3" fill="{PLATFORM_COLORS[platform]}"/>')
        body.append(svg_text(legend_x + 18, height - 43, PLATFORM_LABELS[platform], "legend"))
        legend_x += 96 if platform == "H5" else 78
    return svg_chart(
        "阶段衰减率可视化",
        f"衰减 = 1 − 后一阶段留存 ÷ 前一阶段留存；数值越高，阶段流失越重。{window_label}。",
        width,
        height,
        "".join(body),
    )


def decay_table(platform_metrics: dict[str, dict[str, Any]], window_label: str) -> str:
    """Show sequential retention ratios and loss, always using one mature D15 cohort."""
    stages = (("D3 / D1", "d1", "d3"), ("D7 / D3", "d3", "d7"), ("D15 / D7", "d7", "d15"))
    rows = []
    for platform in ("App", "H5", "PWW", "全部"):
        metrics = platform_metrics[platform]
        cells = []
        for _, base_field, next_field in stages:
            base, nxt = metrics.get(base_field), metrics.get(next_field)
            ratio = nxt / base if base not in (None, 0) and nxt is not None else None
            loss = 1 - ratio if ratio is not None else None
            cells.append(
                f'<td class="num"><strong>{pct(ratio)}</strong><br><span class="muted">衰减 {pct(loss)}</span></td>'
            )
        rows.append(f"<tr><th>{PLATFORM_LABELS[platform]}</th>{''.join(cells)}</tr>")
    return (
        '<div class="decay-box"><h3>分阶段衰减量化</h3>'
        + decay_chart(platform_metrics, window_label)
        + '<div class="table-wrap"><table class="decay-table"><thead><tr><th>形态</th>'
        + '<th class="num">D3 / D1</th><th class="num">D7 / D3</th><th class="num">D15 / D7</th>'
        + '</tr></thead><tbody>' + ''.join(rows) + '</tbody></table></div>'
        + '<p class="table-hint">表格可左右滑动查看完整指标。</p>'
        + f'<p class="note">“保留”表示后一阶段留存 ÷ 前一阶段留存；“衰减”= 1 − 保留。基于同一批成熟 D15 cohort：{safe(window_label)}。源表没有 D14 留存字段，故最后一段为 D15 / D7。</p></div>'
    )


def trend_chart(records: list[dict[str, Any]], start: date, end: date) -> str:
    subset = filtered(records, start, end)
    by_day: dict[date, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for record in subset:
        by_day[record["cohort_date"]][record["platform"]] += n(record.get("新增人数")) or 0
    days = [start + timedelta(days=index) for index in range((end - start).days + 1)]
    width, height = 740, 270
    left, right, top, bottom = 62, 22, 24, 50
    max_y = max((by_day[day][platform] for day in days for platform in ("App", "H5", "PWW")), default=1)
    max_y = (int(max_y / 1000) + 1) * 1000
    plot_w, plot_h = width - left - right, height - top - bottom
    body = [f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" class="plot-frame"/>']
    for fraction in (0, 0.25, 0.5, 0.75, 1):
        y = top + plot_h * (1 - fraction)
        value = max_y * fraction
        body.append(f'<line x1="{left}" x2="{left + plot_w}" y1="{y:.1f}" y2="{y:.1f}" class="gridline"/>')
        body.append(svg_text(left - 10, y + 4, short_int(value), "axis", "end"))
    tick_indices = range(len(days)) if len(days) <= 8 else (0, len(days) // 2, len(days) - 1)
    for index in tick_indices:
        x = left + plot_w * index / max(1, len(days) - 1)
        body.append(svg_text(x, height - 20, date_label(days[index]), "axis", "middle"))
    for platform, color in PLATFORM_COLORS.items():
        points = []
        for index, day in enumerate(days):
            x = left + plot_w * index / max(1, len(days) - 1)
            y = top + plot_h * (1 - by_day[day][platform] / max_y)
            points.append((x, y))
        path = " ".join((f"M {points[0][0]:.1f} {points[0][1]:.1f}",) + tuple(f"L {x:.1f} {y:.1f}" for x, y in points[1:]))
        body.append(f'<path d="{path}" fill="none" stroke="{color}" class="series-line"/>')
        for x, y in points:
            body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3.5" fill="{color}"/>')
    body.append(f'<rect x="{left}" y="{height - 43}" width="12" height="12" rx="3" fill="{APP_COLOR}"/>')
    body.append(svg_text(left + 18, height - 33, "App", "legend"))
    body.append(f'<rect x="{left + 78}" y="{height - 43}" width="12" height="12" rx="3" fill="{H5_COLOR}"/>')
    body.append(svg_text(left + 96, height - 33, "标准 H5", "legend"))
    body.append(f'<rect x="{left + 146}" y="{height - 43}" width="12" height="12" rx="3" fill="{PWW_COLOR}"/>')
    body.append(svg_text(left + 164, height - 33, "PWW", "legend"))
    return svg_chart("最新 7 日新增趋势", f"{fmt_range(start, end)}；仅展示新增，D7 尚未成熟。", width, height, "".join(body))


def bubble_chart(items: list[dict[str, Any]]) -> str:
    width, height = 760, 430
    left, right, top, bottom = 72, 34, 28, 62
    plot_w, plot_h = width - left - right, height - top - bottom
    max_x = max(item["d7"] or 0 for item in items)
    max_y = max(item["payment_per_new"] or 0 for item in items)
    x_max = max(0.28, (int(max_x * 100 / 5) + 1) * 0.05)
    y_max = max(0.32, (int(max_y * 100 / 5) + 1) * 0.05)
    median_x = median([item["d7"] or 0 for item in items])
    median_y = median([item["payment_per_new"] or 0 for item in items])
    body = [f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" class="plot-frame"/>']
    for fraction in (0, 0.25, 0.5, 0.75, 1):
        x = left + plot_w * fraction
        y = top + plot_h * (1 - fraction)
        body.append(f'<line x1="{x:.1f}" x2="{x:.1f}" y1="{top}" y2="{top + plot_h}" class="gridline"/>')
        body.append(f'<line x1="{left}" x2="{left + plot_w}" y1="{y:.1f}" y2="{y:.1f}" class="gridline"/>')
        body.append(svg_text(x, height - 28, pct(x_max * fraction, 0), "axis", "middle"))
        body.append(svg_text(left - 12, y + 4, pct(y_max * fraction, 0), "axis", "end"))
    mx = left + plot_w * median_x / x_max
    my = top + plot_h * (1 - median_y / y_max)
    body.append(f'<line x1="{mx:.1f}" x2="{mx:.1f}" y1="{top}" y2="{top + plot_h}" class="median-line"/>')
    body.append(f'<line x1="{left}" x2="{left + plot_w}" y1="{my:.1f}" y2="{my:.1f}" class="median-line"/>')
    max_new = max(item["new_users"] for item in items)
    positions: list[tuple[float, float, str]] = []
    for item in items:
        x = left + plot_w * (item["d7"] or 0) / x_max
        y = top + plot_h * (1 - (item["payment_per_new"] or 0) / y_max)
        radius = 9 + 23 * sqrt(item["new_users"] / max_new)
        color = PLATFORM_COLORS[item["platform"]]
        body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{radius:.1f}" fill="{color}" fill-opacity="0.78" stroke="#FFFFFF" stroke-width="2"/>')
        positions.append((x, y, item["short_label"]))
    for index, (x, y, label) in enumerate(positions):
        offset = -14 if index % 2 == 0 else 22
        body.append(svg_text(x, y + offset, label, "bubble-label", "middle"))
    body.append(svg_text(left + plot_w / 2, height - 4, "D7 留存（%）", "axis-title", "middle"))
    body.append(f'<text x="16" y="{top + plot_h / 2}" transform="rotate(-90 16 {top + plot_h / 2})" class="axis-title" text-anchor="middle">新增付费人数 / 新增（%）</text>')
    body.append(f'<rect x="{left}" y="{height - 52}" width="12" height="12" rx="3" fill="{APP_COLOR}"/>')
    body.append(svg_text(left + 18, height - 42, "App", "legend"))
    body.append(f'<rect x="{left + 78}" y="{height - 52}" width="12" height="12" rx="3" fill="{H5_COLOR}"/>')
    body.append(svg_text(left + 96, height - 42, "标准 H5", "legend"))
    body.append(f'<rect x="{left + 146}" y="{height - 52}" width="12" height="12" rx="3" fill="{PWW_COLOR}"/>')
    body.append(svg_text(left + 164, height - 42, "PWW", "legend"))
    return svg_chart(
        "渠道质量矩阵：留存与付费人数率",
        "横轴 D7、纵轴新增付费人数/新增；圆面积代表新增规模，虚线为 8 个渠道中位数。",
        width,
        height,
        "".join(body),
    )


def pareto_chart(items: list[dict[str, Any]]) -> str:
    items = sorted(items, key=lambda item: item["d7_retained"], reverse=True)
    width, height = 760, 360
    left, right, top, bottom = 54, 34, 26, 86
    plot_w, plot_h = width - left - right, height - top - bottom
    max_value = max(item["d7_retained"] for item in items)
    total = sum(item["d7_retained"] for item in items)
    slot = plot_w / len(items)
    body = [f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" class="plot-frame"/>']
    for fraction in (0, 0.5, 1):
        y = top + plot_h * (1 - fraction)
        body.append(f'<line x1="{left}" x2="{left + plot_w}" y1="{y:.1f}" y2="{y:.1f}" class="gridline"/>')
        body.append(svg_text(left - 10, y + 4, short_int(max_value * fraction), "axis", "end"))
        body.append(svg_text(left + plot_w + 9, y + 4, pct(fraction, 0), "axis", "start"))
    cumulative = 0.0
    points = []
    for index, item in enumerate(items):
        x = left + index * slot + slot * 0.16
        bar_w = slot * 0.68
        height_bar = plot_h * item["d7_retained"] / max_value
        y = top + plot_h - height_bar
        color = PLATFORM_COLORS[item["platform"]]
        body.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{height_bar:.1f}" rx="4" fill="{color}"/>')
        cumulative += item["d7_retained"]
        px = x + bar_w / 2
        py = top + plot_h * (1 - cumulative / total)
        points.append((px, py))
        body.append(f'<text x="{px:.1f}" y="{height - 45}" class="axis tiny" text-anchor="end" transform="rotate(-38 {px:.1f} {height - 45})">{safe(item["short_label"])}</text>')
    if points:
        path = " ".join((f"M {points[0][0]:.1f} {points[0][1]:.1f}",) + tuple(f"L {x:.1f} {y:.1f}" for x, y in points[1:]))
        body.append(f'<path d="{path}" fill="none" stroke="{MUTED}" class="pareto-line"/>')
        for x, y in points:
            body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="{MUTED}"/>')
    body.append(svg_text(left, height - 8, "柱：成熟 D7 留存人数", "legend"))
    body.append(svg_text(left + 170, height - 8, "线：累计占比", "legend"))
    return svg_chart("成熟 D7 留存人数贡献帕累托", "当前成熟 D7 cohort；用于识别影响整体留存最多的包体/渠道。", width, height, "".join(body))


def grouped_rate_chart(items: list[dict[str, Any]]) -> str:
    width, height = 760, 420
    left, right, top, bottom = 208, 36, 24, 54
    plot_w, plot_h = width - left - right, height - top - bottom
    max_value = 1.0
    row_h = plot_h / len(items)
    series = [
        ("人数/新增", "payment_per_new", APP_COLOR),
        ("来源显示率", "payment_source_rate", YELLOW),
        ("分母覆盖率", "payment_coverage", H5_COLOR),
    ]
    body = [f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" class="plot-frame"/>']
    for fraction in (0, 0.25, 0.5, 0.75, 1):
        x = left + plot_w * fraction
        body.append(f'<line x1="{x:.1f}" x2="{x:.1f}" y1="{top}" y2="{top + plot_h}" class="gridline"/>')
        body.append(svg_text(x, height - 20, pct(fraction, 0), "axis", "middle"))
    for index, item in enumerate(items):
        y_base = top + index * row_h
        body.append(svg_text(left - 12, y_base + row_h * 0.6, item.get("chart_label", item["label"]), "axis", "end"))
        for s_index, (_, key, color) in enumerate(series):
            value = item[key] or 0
            y = y_base + 6 + s_index * (row_h - 12) / len(series)
            bar_h = max(3.5, (row_h - 16) / len(series) - 2)
            body.append(f'<rect x="{left}" y="{y:.1f}" width="{plot_w * value:.1f}" height="{bar_h:.1f}" rx="2" fill="{color}"/>')
    legend_x = left
    for name, _, color in series:
        body.append(f'<rect x="{legend_x}" y="{height - 46}" width="11" height="11" rx="2" fill="{color}"/>')
        body.append(svg_text(legend_x + 16, height - 36, name, "legend"))
        legend_x += 103
    return svg_chart("新增付费口径并列对比", "三项口径不可合并为单一“付费率”；分母覆盖率越低，显示率可比性越弱。", width, height, "".join(body))


def make_heatmap(items: list[dict[str, Any]]) -> str:
    values = [item[field] or 0 for item in items for field in ("d1", "d3", "d7")]
    low, high = min(values), max(values)
    def style(value: float | None) -> str:
        if value is None:
            return "background:#F2F5F7;color:#738392;"
        ratio = (value - low) / (high - low) if high > low else 0.5
        alpha = 0.18 + 0.68 * ratio
        return f"background:rgba(50,137,199,{alpha:.3f});color:{'#FFFFFF' if ratio > .58 else TEXT};"
    rows = []
    for item in items:
        dot_class = "app-dot" if item["platform"] == "App" else "pww-dot" if item["platform"] == "PWW" else "h5-dot"
        rows.append(
            "<tr>"
            f"<th><span class=\"tiny-dot {dot_class}\"></span>{safe(item['label'])}</th>"
            f"<td style=\"{style(item['d1'])}\">{pct(item['d1'])}</td>"
            f"<td style=\"{style(item['d3'])}\">{pct(item['d3'])}</td>"
            f"<td style=\"{style(item['d7'])}\">{pct(item['d7'])}</td>"
            f"<td class=\"{color_for_delta(item['d7_delta'])}\">{signed_pp(item['d7_delta'])}</td>"
            "</tr>"
        )
    return (
        '<div class="table-wrap"><table class="heatmap"><thead><tr><th>包体 / 渠道</th><th>D1</th><th>D3</th><th>D7</th><th>D7环比</th></tr></thead>'
        f"<tbody>{''.join(rows)}</tbody></table></div><p class=\"table-hint\">表格可左右滑动查看完整指标。</p>"
    )


def channel_table(items: list[dict[str, Any]]) -> str:
    rows = []
    for item in items:
        tag_class = "app-tag" if item["platform"] == "App" else "pww-tag" if item["platform"] == "PWW" else "h5-tag"
        rows.append(
            "<tr>"
            f"<td><span class=\"platform-tag {tag_class}\">{PLATFORM_LABELS[item['platform']]}</span></td>"
            f"<td><strong>{safe(item['package'])}</strong><br><span class=\"muted\">{safe(item['channel'])}</span></td>"
            f"<td class=\"num\">{integer(item['new_users'])}<br><span class=\"muted\">{pct(item['share'])}</span></td>"
            f"<td class=\"num\">{pct(item['d1'])}</td><td class=\"num\">{pct(item['d3'])}</td>"
            f"<td class=\"num emph\">{pct(item['d7'])}<br><span class=\"{color_for_delta(item['d7_delta'])}\">{signed_pp(item['d7_delta'])}</span></td>"
            f"<td class=\"num\">{pct(item['payment_per_new'])}</td>"
            f"<td class=\"num\">{pct(item['first_charge_per_new'])}</td>"
            f"<td class=\"num\">{pct(item['payment_coverage'])}</td>"
            "</tr>"
        )
    return (
        '<div class="table-wrap"><table class="detail-table"><thead><tr>'
        '<th>形态</th><th>包体 / 渠道</th><th class="num">新增 / 占比</th><th class="num">D1</th><th class="num">D3</th>'
        '<th class="num">D7 / 环比</th><th class="num">付费人数率</th><th class="num">首充人数率*</th><th class="num">付费分母覆盖</th>'
        '</tr></thead><tbody>' + ''.join(rows) + '</tbody></table></div><p class="table-hint">表格可左右滑动查看完整指标。</p>'
    )


def long_retention_table(long_stats: dict[str, dict[str, Any]], windows: dict[str, tuple[date, date]]) -> str:
    rows = []
    for platform in ("App", "H5", "PWW", "全部"):
        cells = []
        for stage in ("D15", "D30", "D60"):
            stats = long_stats[stage][platform]
            cells.append(f"<td class=\"num\">{pct(stats[stage.lower()])}<br><span class=\"muted\">n={integer(stats['new_users'])}</span></td>")
        rows.append(f"<tr><th>{PLATFORM_LABELS[platform]}</th>{''.join(cells)}</tr>")
    header_detail = " · ".join(f"{stage}: {fmt_range(*windows[stage])}" for stage in ("D15", "D30", "D60"))
    return (
        '<div class="table-wrap"><table class="long-table"><thead><tr><th>形态</th><th>D15</th><th>D30</th><th>D60</th></tr></thead>'
        f"<tbody>{''.join(rows)}</tbody></table></div><p class=\"table-hint\">表格可左右滑动查看完整指标。</p><p class=\"note\">各阶段使用独立的最近 28 天成熟 cohort：{safe(header_detail)}。因此横向阅读用于生命周期观察，不与当前 D7 作同样本环比。</p>"
    )


def delta_cell(value: float | None, positive_good: bool = True) -> str:
    return f'<span class="{color_for_delta(value, positive_good)}">{signed_pp(value)}</span>'


def build_report(records: list[dict[str, Any]], quality: dict[str, Any]) -> str:
    cutoff = quality["max_date"]
    d7_end = cutoff - timedelta(days=7)
    current_start = d7_end - timedelta(days=27)
    current_end = d7_end
    previous_end = current_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=27)
    current = filtered(records, current_start, current_end)
    previous = filtered(records, previous_start, previous_end)
    overall_current = compute_metrics(current)
    overall_previous = compute_metrics(previous)

    platform_current = grouped_metrics(records, "platform", current_start, current_end)
    platform_previous = grouped_metrics(records, "platform", previous_start, previous_end)
    sheet_current = grouped_metrics(records, "source_sheet", current_start, current_end)
    sheet_previous = grouped_metrics(records, "source_sheet", previous_start, previous_end)
    d15_end = cutoff - timedelta(days=15)
    d15_start = d15_end - timedelta(days=27)
    d15_records = filtered(records, d15_start, d15_end)
    platform_d15 = {
        "App": compute_metrics([record for record in d15_records if record["platform"] == "App"]),
        "H5": compute_metrics([record for record in d15_records if record["platform"] == "H5"]),
        "PWW": compute_metrics([record for record in d15_records if record["platform"] == "PWW"]),
        "全部": compute_metrics(d15_records),
    }

    items: list[dict[str, Any]] = []
    short_labels = {
        "WajeSpecial-facebook": "安卓-FB",
        "WajeSpecial-googleadwords_int": "安卓-GA",
        "WajeSpecial-Google商店": "安卓-Play",
        "PAWAJEIOS-AppStore商店": "iOS-Store",
        "PAWAJEBETH5": "Bet H5",
        "pww": "PWW",
        "wajeH5-fb": "H5-FB",
        "wajeH5ga-googlewors_int": "H5-GA",
    }
    for sheet, (platform, package, channel) in SHEET_MAP.items():
        current_stats = sheet_current[sheet]
        prior_stats = sheet_previous[sheet]
        package_label = "Waje H5 PWW（APP 安装型 H5 游戏包）" if sheet == "pww" else package
        items.append(
            {
                "sheet": sheet,
                "platform": platform,
                "package": package_label,
                "channel": channel,
                "label": f"{package_label} · {channel}",
                "chart_label": "PWW（APP安装型H5包）" if sheet == "pww" else f"{package} · {channel}",
                "short_label": short_labels[sheet],
                "is_pww": sheet == "pww",
                **current_stats,
                "share": current_stats["new_users"] / overall_current["new_users"],
                "d1_delta": (current_stats["d1"] or 0) - (prior_stats["d1"] or 0),
                "d3_delta": (current_stats["d3"] or 0) - (prior_stats["d3"] or 0),
                "d7_delta": (current_stats["d7"] or 0) - (prior_stats["d7"] or 0),
                "new_delta": current_stats["new_users"] / prior_stats["new_users"] - 1 if prior_stats["new_users"] else None,
                "d7_retained": current_stats["new_users"] * (current_stats["d7"] or 0),
            }
        )
    items.sort(key=lambda item: item["new_users"], reverse=True)

    h5_fb = next(item for item in items if item["sheet"] == "wajeH5-fb")
    h5_ga = next(item for item in items if item["sheet"] == "wajeH5ga-googlewors_int")
    bet_h5 = next(item for item in items if item["sheet"] == "PAWAJEBETH5")
    android_ga = next(item for item in items if item["sheet"] == "WajeSpecial-googleadwords_int")
    android_play = next(item for item in items if item["sheet"] == "WajeSpecial-Google商店")
    ios = next(item for item in items if item["sheet"] == "PAWAJEIOS-AppStore商店")
    pww = next(item for item in items if item["sheet"] == "pww")

    long_stats: dict[str, dict[str, Any]] = {}
    long_windows: dict[str, tuple[date, date]] = {}
    for stage, lag in (("D15", 15), ("D30", 30), ("D60", 60)):
        end = cutoff - timedelta(days=lag)
        start = end - timedelta(days=27)
        long_windows[stage] = (start, end)
        subset = filtered(records, start, end)
        long_stats[stage] = {
            "App": compute_metrics([record for record in subset if record["platform"] == "App"]),
            "H5": compute_metrics([record for record in subset if record["platform"] == "H5"]),
            "PWW": compute_metrics([record for record in subset if record["platform"] == "PWW"]),
            "全部": compute_metrics(subset),
        }

    leading_start = cutoff - timedelta(days=6)
    leading_end = cutoff

    first_charge_gt_payment = sum(
        1
        for record in records
        if (n(record.get("首充付费人数")) or 0) > (n(record.get("新增付费人数")) or 0)
    )
    h5_fb_opportunity = h5_fb["new_users"] * max(0, (h5_ga["d7"] or 0) - (h5_fb["d7"] or 0))
    h5_total_share = platform_current["H5"]["new_users"] / overall_current["new_users"]
    pww_total_share = platform_current["PWW"]["new_users"] / overall_current["new_users"]
    d7_gap = (platform_current["App"]["d7"] or 0) - (platform_current["H5"]["d7"] or 0)

    def metric_card(label: str, value: str, detail: str, tone: str = "blue") -> str:
        return f'<article class="metric-card {tone}"><p>{safe(label)}</p><strong>{value}</strong><span>{detail}</span></article>'

    app_new_delta = platform_current["App"]["new_users"] / platform_previous["App"]["new_users"] - 1
    h5_new_delta = platform_current["H5"]["new_users"] / platform_previous["H5"]["new_users"] - 1
    pww_new_delta = platform_current["PWW"]["new_users"] / platform_previous["PWW"]["new_users"] - 1
    total_new_delta = overall_current["new_users"] / overall_previous["new_users"] - 1
    total_d7_delta = (overall_current["d7"] or 0) - (overall_previous["d7"] or 0)

    platform_rows = []
    for platform in ("App", "H5", "PWW"):
        current_stats = platform_current[platform]
        prior_stats = platform_previous[platform]
        tag_class = "app-tag" if platform == "App" else "pww-tag" if platform == "PWW" else "h5-tag"
        platform_rows.append(
            "<tr>"
            f"<th><span class=\"platform-tag {tag_class}\">{PLATFORM_LABELS[platform]}</span></th>"
            f"<td class=\"num\">{integer(current_stats['new_users'])}<br>{signed_pct(current_stats['new_users']/prior_stats['new_users']-1)}</td>"
            f"<td class=\"num\">{pct(current_stats['d1'])}<br>{delta_cell((current_stats['d1'] or 0)-(prior_stats['d1'] or 0))}</td>"
            f"<td class=\"num\">{pct(current_stats['d3'])}<br>{delta_cell((current_stats['d3'] or 0)-(prior_stats['d3'] or 0))}</td>"
            f"<td class=\"num emph\">{pct(current_stats['d7'])}<br>{delta_cell((current_stats['d7'] or 0)-(prior_stats['d7'] or 0))}</td>"
            f"<td class=\"num\">{pct(current_stats['payment_per_new'])}</td>"
            f"<td class=\"num\">{pct(current_stats['payment_coverage'])}</td>"
            "</tr>"
        )

    platform_table = (
        '<div class="table-wrap"><table class="platform-table"><thead><tr><th>形态</th><th class="num">新增 / 环比</th><th class="num">D1 / 环比</th>'
        '<th class="num">D3 / 环比</th><th class="num">D7 / 环比</th><th class="num">付费人数率</th><th class="num">付费分母覆盖</th>'
        '</tr></thead><tbody>' + ''.join(platform_rows) + '</tbody></table></div><p class="table-hint">表格可左右滑动查看完整指标。</p>'
    )

    quality_rows = []
    for sheet in SHEET_MAP:
        detail = quality["sheet_quality"][sheet]
        quality_rows.append(
            f"<tr><td>{safe(sheet)}</td><td class=\"num\">{integer(detail['records'])}</td><td>{detail['min']:%Y-%m-%d}</td>"
            f"<td>{detail['max']:%Y-%m-%d}</td><td class=\"num\">{integer(detail['missing_days'])}</td></tr>"
        )

    # The HTML deliberately contains all required styles and no external dependency.
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="color-scheme" content="light">
<title>新包 App、H5 与 PWW 渠道包体留存付费深度分析｜2026-08-17</title>
<style>
  :root {{
    --ink:#17324D; --muted:#61778A; --blue:#2E83C5; --blue-2:#81BEE8; --blue-pale:#EAF5FD;
    --green:#4DAA80; --green-pale:#E6F5EC; --yellow:#E6B63E; --yellow-pale:#FFF5CD;
    --red:#D86767; --red-pale:#FDEEEE; --line:#D8E6EF; --paper:#FFFFFF; --bg:#F5FAFD;
  }}
  * {{ box-sizing:border-box; }}
  html {{ scroll-behavior:smooth; background:var(--bg); }}
  body {{ margin:0; color:var(--ink); background:linear-gradient(180deg,#EFF8FE 0,#F8FCFE 26%,#F4FAF8 100%); font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; line-height:1.58; }}
  .report {{ width:min(1200px, calc(100% - 36px)); margin:24px auto 60px; }}
  .hero {{ overflow:hidden; position:relative; border-radius:24px; padding:46px 52px 38px; background:linear-gradient(128deg,#123E6B,#236DA9 58%,#5FA8C9); color:#fff; box-shadow:0 18px 50px rgba(31,91,139,.20); }}
  .hero::after {{ content:""; position:absolute; width:430px; height:430px; border-radius:50%; right:-135px; top:-220px; background:rgba(199,239,240,.18); }}
  .hero > * {{ position:relative; z-index:1; }}
  .eyebrow {{ display:inline-block; margin:0 0 12px; padding:4px 12px; border-radius:99px; font-size:12px; letter-spacing:.08em; color:#E7F6FF; background:rgba(255,255,255,.15); }}
  h1 {{ max-width:850px; margin:0; font-size:clamp(31px,4.7vw,51px); line-height:1.15; letter-spacing:-.035em; }}
  .subtitle {{ max-width:765px; margin:16px 0 0; color:#E6F5FC; font-size:17px; }}
  .hero-meta {{ display:flex; flex-wrap:wrap; gap:10px 22px; margin-top:25px; font-size:13px; color:#DFF3FF; }}
  .hero-meta strong {{ color:#fff; }}
  .toc {{ display:flex; flex-wrap:wrap; gap:8px; margin:20px 0 8px; padding:0; list-style:none; }}
  .toc a {{ display:block; padding:7px 12px; border:1px solid var(--line); border-radius:99px; color:#35617C; background:#fff; text-decoration:none; font-size:13px; }}
  .toc a:hover {{ color:var(--blue); border-color:var(--blue-2); }}
  section {{ margin-top:28px; padding:31px 34px; border:1px solid rgba(216,230,239,.9); border-radius:20px; background:rgba(255,255,255,.92); box-shadow:0 8px 26px rgba(39,97,136,.045); }}
  section.tint {{ background:linear-gradient(140deg,#F6FBFF,#FBFFFD); }}
  .section-top {{ display:flex; gap:18px; align-items:flex-start; justify-content:space-between; margin-bottom:20px; }}
  h2 {{ margin:0; font-size:24px; line-height:1.28; letter-spacing:-.02em; }}
  h2 .num {{ display:inline-flex; justify-content:center; align-items:center; width:28px; height:28px; margin-right:8px; border-radius:50%; color:#fff; background:var(--blue); font-size:14px; vertical-align:2px; }}
  h3 {{ margin:28px 0 12px; font-size:17px; }}
  h4 {{ margin:0 0 8px; font-size:15px; }}
  p {{ margin:8px 0; }}
  .muted,.note {{ color:var(--muted); }}
  .note {{ font-size:12px; }}
  .period {{ flex:0 0 auto; margin:0; padding:5px 10px; border-radius:99px; color:#52758B; background:var(--blue-pale); font-size:12px; white-space:nowrap; }}
  .kpi-grid {{ display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:12px; }}
  .metric-card {{ min-height:126px; padding:17px 16px 14px; border-radius:16px; border:1px solid transparent; }}
  .metric-card p {{ margin:0 0 8px; color:var(--muted); font-size:12px; font-weight:600; }}
  .metric-card strong {{ display:block; font-size:27px; line-height:1.05; letter-spacing:-.04em; }}
  .metric-card span {{ display:block; margin-top:8px; font-size:12px; color:#507087; }}
  .metric-card.blue {{ background:var(--blue-pale); border-color:#D4EBFA; }}
  .metric-card.green {{ background:var(--green-pale); border-color:#D4EFDF; }}
  .metric-card.yellow {{ background:var(--yellow-pale); border-color:#F7E6A2; }}
  .metric-card.red {{ background:var(--red-pale); border-color:#F5D5D5; }}
  .highlight-grid {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:14px; margin-top:18px; }}
  .highlight {{ padding:16px 17px; border-left:4px solid var(--blue); border-radius:11px; background:#F5FAFD; }}
  .highlight.green {{ border-left-color:var(--green); background:#F5FCF8; }}
  .highlight.yellow {{ border-left-color:var(--yellow); background:#FFFBEA; }}
  .highlight p {{ margin:0; font-size:14px; }}
  .highlight strong {{ color:#164D77; }}
  .cols-2 {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:24px; align-items:start; }}
  .cols-3 {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:16px; }}
  .chart {{ margin:0; padding:0; min-width:0; }}
  .chart figcaption {{ display:flex; flex-wrap:wrap; gap:5px 10px; justify-content:space-between; margin:0 0 8px; color:var(--muted); font-size:12px; }}
  .chart figcaption strong {{ color:var(--ink); font-size:14px; }}
  .chart svg {{ display:block; width:100%; height:auto; overflow:visible; }}
  .plot-frame {{ fill:#FBFEFF; stroke:var(--line); stroke-width:1; }}
  .gridline {{ stroke:var(--line); stroke-width:1; }}
  .median-line {{ stroke:#A2B8C7; stroke-width:1; stroke-dasharray:4 4; }}
  .series-line {{ stroke-width:3.2; stroke-linecap:round; stroke-linejoin:round; }}
  .pareto-line {{ stroke-width:2.6; stroke-linecap:round; stroke-linejoin:round; }}
  svg text {{ font-family:inherit; fill:var(--ink); }}
  svg .axis {{ font-size:11px; fill:var(--muted); }} svg .axis-title {{ font-size:12px; fill:var(--muted); font-weight:600; }}
  svg .label,svg .legend {{ font-size:12px; fill:var(--muted); }} svg .value {{ font-size:11px; fill:var(--ink); font-weight:700; }}
  svg .bubble-label {{ font-size:10.5px; fill:var(--ink); font-weight:700; paint-order:stroke; stroke:#fff; stroke-width:3px; stroke-linejoin:round; }} svg .tiny {{ font-size:10px; }}
  .rule {{ height:1px; margin:24px 0; background:var(--line); }}
  table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:13px; }}
  th,td {{ padding:11px 10px; border-bottom:1px solid var(--line); vertical-align:middle; }}
  thead th {{ color:#4D6D83; background:#F2F8FC; font-size:12px; font-weight:700; text-align:left; }}
  tbody tr:last-child th,tbody tr:last-child td {{ border-bottom:0; }}
  .num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  .emph {{ color:#124F7B; font-weight:700; }}
  .table-wrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:13px; }}
  .table-wrap table {{ min-width:760px; }}
  .table-hint {{ display:none; }}
  .platform-tag {{ display:inline-block; min-width:35px; padding:3px 8px; border-radius:99px; font-size:11px; font-weight:700; text-align:center; }}
  .app-tag {{ color:#1D659C; background:#DCEFFC; }} .h5-tag {{ color:#27805C; background:#DDF3E6; }} .pww-tag {{ color:#876700; background:#FFF1BE; }}
  .tiny-dot {{ display:inline-block; width:8px; height:8px; margin-right:7px; border-radius:50%; }} .app-dot {{ background:var(--blue); }} .h5-dot {{ background:var(--green); }} .pww-dot {{ background:#D4A72C; }}
  .positive {{ color:#1F8A61; font-weight:700; }} .negative {{ color:#C55757; font-weight:700; }} .neutral {{ color:var(--muted); }}
  .heatmap td {{ font-weight:700; text-align:right; font-variant-numeric:tabular-nums; }} .heatmap th {{ white-space:nowrap; }}
  .callout {{ margin:16px 0 0; padding:13px 15px; border-radius:12px; color:#37586E; background:#F4FAFD; font-size:13px; }}
  .callout.warning {{ border-left:4px solid var(--yellow); background:#FFFBE9; }} .callout.alert {{ border-left:4px solid var(--red); background:#FFF5F5; }}
  .formula {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:10px; margin-top:15px; }}
  .formula > div {{ padding:12px; border-radius:12px; background:#F6FAFC; font-size:12px; }} .formula strong {{ display:block; margin-bottom:4px; color:#28587A; }}
  .action-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:14px; }}
  .action {{ padding:18px; border:1px solid var(--line); border-radius:15px; background:#FCFEFF; }}
  .action .priority {{ display:inline-block; margin-bottom:8px; padding:3px 8px; border-radius:99px; font-size:11px; font-weight:800; }}
  .p0 {{ color:#A44646; background:#FDECEC; }} .p1 {{ color:#946E00; background:#FFF5D5; }} .p2 {{ color:#2B6C8B; background:#E7F5FC; }}
  .action ul {{ margin:8px 0 0; padding-left:18px; color:#48677C; font-size:13px; }} .action li+li {{ margin-top:4px; }}
  .quality-summary {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:11px; }}
  .quality-summary > div {{ padding:12px 13px; border-radius:12px; background:#F4FAFD; }} .quality-summary strong {{ display:block; font-size:21px; }} .quality-summary span {{ color:var(--muted); font-size:12px; }}
  .footer {{ margin:24px 0 0; padding:18px 8px; color:#6B8291; font-size:12px; text-align:center; }}
  .source-chip {{ display:inline-block; margin:3px 4px 0 0; padding:2px 7px; border-radius:99px; color:#517389; background:#EEF6FA; font-size:11px; }}
  @media (max-width:950px) {{ .kpi-grid {{ grid-template-columns:repeat(3,minmax(0,1fr)); }} .cols-2,.action-grid {{ grid-template-columns:1fr; }} .cols-3 {{ grid-template-columns:1fr; }} }}
  @media (max-width:640px) {{ .report {{ width:min(100% - 20px,1200px); margin-top:10px; }} .hero {{ padding:31px 24px; border-radius:17px; }} section {{ padding:24px 17px; border-radius:16px; }} .kpi-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .highlight-grid,.quality-summary,.formula {{ grid-template-columns:1fr; }} .section-top {{ display:block; }} .period {{ display:inline-block; margin-top:8px; white-space:normal; }} h2 {{ font-size:21px; }} .toc {{ display:none; }} .table-hint {{ display:block; margin:6px 0 0; color:var(--muted); font-size:11px; }} }}
  @media print {{ @page {{ size:A4; margin:10mm; }} body {{ background:#fff; font-size:11px; }} .report {{ width:100%; margin:0; }} .hero {{ padding:28px 32px; border-radius:0; box-shadow:none; }} .toc {{ display:none; }} section {{ margin-top:13px; padding:19px 22px; border-radius:0; box-shadow:none; break-inside:auto; }} .section-top,h2,h3 {{ break-after:avoid-page; }} .chart,.metric-card,.highlight,.action,.table-wrap,.formula,.quality-summary {{ break-inside:avoid-page; }} .metric-card {{ min-height:92px; padding:11px; }} .metric-card strong {{ font-size:22px; }} .chart figcaption {{ font-size:10px; }} .action {{ padding:12px; }} .table-wrap {{ overflow:visible; }} .table-wrap table {{ min-width:0; font-size:10px; }} th,td {{ padding:6px; }} }}
</style>
</head>
<body>
<main class="report">
  <header class="hero">
    <p class="eyebrow">WAJE · 新包 cohort 深度分析</p>
    <h1>App、标准 H5 与 PWW：渠道、包体与新用户价值的真实差异</h1>
    <p class="subtitle">以成熟 cohort 为唯一统计基础，拆解新增规模、首周留存、付费人数转化及来源口径覆盖。报告中的“版本”指包体/产品形态，不等同于客户端发布版本。</p>
    <div class="hero-meta"><span>数据截至 <strong>{cutoff:%Y-%m-%d}</strong></span><span>8 个包体/渠道序列 · <strong>{integer(len(records))}</strong> 条 cohort 日记录</span><span>报告生成 <strong>{REPORT_DATE:%Y-%m-%d}</strong></span></div>
  </header>

  <nav class="toc" aria-label="报告目录"><a href="#summary">核心结论</a><a href="#platform">App / H5 / PWW</a><a href="#channel">渠道与包体</a><a href="#lifecycle">生命周期</a><a href="#payment">付费口径</a><a href="#actions">优化动作</a><a href="#quality">数据质量</a></nav>

  <section id="summary" class="tint">
    <div class="section-top"><div><h2><span class="num">1</span>核心结论</h2><p class="muted">当前成熟 D7 cohort：{fmt_range(current_start,current_end)}；对比上一窗口：{fmt_range(previous_start,previous_end)}。</p></div><p class="period">D1 / D3 / D7 使用同一批成熟 D7 cohort</p></div>
    <div class="kpi-grid">
      {metric_card("成熟 D7 窗口新增", integer(overall_current['new_users']), f"环比 {signed_pct(total_new_delta)} · 上期 {integer(overall_previous['new_users'])}", "blue")}
      {metric_card("整体 D1", pct(overall_current['d1']), f"环比 {signed_pp((overall_current['d1'] or 0)-(overall_previous['d1'] or 0))}", "blue")}
      {metric_card("整体 D3", pct(overall_current['d3']), f"环比 {signed_pp((overall_current['d3'] or 0)-(overall_previous['d3'] or 0))}", "green")}
      {metric_card("整体 D7", pct(overall_current['d7']), f"环比 {signed_pp(total_d7_delta)} · 基本持平", "green")}
      {metric_card("付费人数 / 新增", pct(overall_current['payment_per_new']), f"来源显示率 {pct(overall_current['payment_source_rate'])}", "yellow")}
      {metric_card("付费分母覆盖率", pct(overall_current['payment_coverage']), "显示率不可直接当作新增付费率", "yellow")}
    </div>
    <div class="highlight-grid">
      <article class="highlight"><p><strong>整体稳定掩盖形态断层。</strong> App / 标准 H5 / PWW 的 D7 分别为 <strong>{pct(platform_current['App']['d7'])} / {pct(platform_current['H5']['d7'])} / {pct(platform_current['PWW']['d7'])}</strong>；标准 H5 与 PWW 分别贡献 {pct(h5_total_share)}、{pct(pww_total_share)} 的新增。App 与标准 H5 相差 <strong>{pct(d7_gap)}</strong>。</p></article>
      <article class="highlight green"><p><strong>H5 Google 是改善样本。</strong> D7 {pct(h5_ga['d7'])}、环比 {signed_pp(h5_ga['d7_delta'])}，但新增环比 {signed_pct(h5_ga['new_delta'])}，需要同时看量与质。</p></article>
      <article class="highlight yellow"><p><strong>最大优先级在 H5 Facebook。</strong> 新增 {integer(h5_fb['new_users'])}（总新增 {pct(h5_fb['share'])}），D7 仅 {pct(h5_fb['d7'])}，付费人数率 {pct(h5_fb['payment_per_new'])}。</p></article>
    </div>
  </section>

  <section id="platform">
    <div class="section-top"><div><h2><span class="num">2</span>App、标准 H5 与 PWW：规模、留存与付费的形态差异</h2><p class="muted">PWW 为 App 安装、H5 运行的混合形态，独立于 App 和标准浏览器 H5 汇总。</p></div><p class="period">当前 vs 上一成熟 D7 窗口</p></div>
    {platform_table}
    <div class="cols-2" style="margin-top:24px">
      {retention_path_chart(platform_d15, fmt_range(d15_start,d15_end))}
      {trend_chart(records, leading_start, leading_end)}
    </div>
    {decay_table(platform_d15, fmt_range(d15_start,d15_end))}
    <div class="callout"><strong>读数：</strong>App 新增环比 {signed_pct(app_new_delta)}、D7 {signed_pp((platform_current['App']['d7'] or 0)-(platform_previous['App']['d7'] or 0))}；标准 H5 新增环比 {signed_pct(h5_new_delta)}、D7 {signed_pp((platform_current['H5']['d7'] or 0)-(platform_previous['H5']['d7'] or 0))}；PWW 新增环比 {signed_pct(pww_new_delta)}、D7 {signed_pp((platform_current['PWW']['d7'] or 0)-(platform_previous['PWW']['d7'] or 0))}。三类差异需要联查渠道结构、入口性能和首局路径，不能直接归因于产品形态。</div>
    <p class="note"><strong>PWW 归类说明：</strong>PWW 是 <strong>APP 形式安装的 H5 游戏包</strong>。本节将其作为独立第三类，不计入 App 或标准 H5；“全部”指标仍包含 PWW，避免与普通浏览器 H5 混淆。</p>
  </section>

  <section id="channel" class="tint">
    <div class="section-top"><div><h2><span class="num">3</span>包体（产品版本）× 渠道：交叉细钻</h2><p class="muted">包体是当前源表可用的“版本”口径；未包含真实 app/web build 字段。</p></div><p class="period">8 个包体/渠道序列</p></div>
    <div class="cols-2">
      {bubble_chart(items)}
      {pareto_chart(items)}
    </div>
    <h3>首周留存热力图</h3>
    {make_heatmap(items)}
    <p class="note">D7 环比均对比上一成熟 D7 窗口。色深仅表示当前窗口的相对留存水平，不代表统计显著性。</p>
    <h3>完整交叉汇总</h3>
    {channel_table(items)}
    <p class="note">* 首充人数率为来源“首充付费人数/新增人数”的展示结果，仅作独立观察；因其与新增付费人数不满足集合包含关系，不能当作漏斗下一步。</p>
  </section>

  <section id="lifecycle">
    <div class="section-top"><div><h2><span class="num">4</span>生命周期：先盯首周，再看长周期</h2><p class="muted">D1、D3、D7 同 cohort；D15 以后严格按各自成熟窗口单独观察。</p></div><p class="period">未成熟 cohort 一律 N/A</p></div>
    <div class="cols-3">
      <article class="highlight"><h4>高量低留存：H5 Facebook</h4><p>D1 {pct(h5_fb['d1'])} → D3 {pct(h5_fb['d3'])} → D7 {pct(h5_fb['d7'])}。若只从 D7 尺度看，达到当前 H5 Google 水平的理论差额约 <strong>{integer(h5_fb_opportunity)}</strong> 名 D7 留存用户；这是优先级估算，不是预测。</p></article>
      <article class="highlight green"><h4>改善信号：WajeBet H5</h4><p>D7 {pct(bet_h5['d7'])}，环比 <strong>{signed_pp(bet_h5['d7_delta'])}</strong>；D1、D3 同步改善。应结合活动、入口和首局路径验证改善是否可复现。</p></article>
      <article class="highlight yellow"><h4>小样本波动：PWW（APP 安装型 H5 游戏包）</h4><p>安装形态为 App、游戏运行形态为 H5。新增 {integer(pww['new_users'])}，D7 {pct(pww['d7'])}、环比 {signed_pp(pww['d7_delta'])}。先补全渠道映射并观察连续两个窗口，再决定扩大投入。</p></article>
    </div>
    <h3>长周期成熟留存快照</h3>
    {long_retention_table(long_stats,long_windows)}
  </section>

  <section id="payment" class="tint">
    <div class="section-top"><div><h2><span class="num">5</span>付费诊断：先治理分母，再解释转化</h2><p class="muted">来源显示率、人数/新增率与分母覆盖率分别呈现，避免将不同统计人群混成一项“付费率”。</p></div><p class="period">当前成熟 D7 窗口</p></div>
    {grouped_rate_chart(items)}
    <div class="formula">
      <div><strong>人数口径</strong><code>Σ 新增付费人数 ÷ Σ 新增人数</code>。用于当前新用户样本内的可加总比较。</div>
      <div><strong>来源显示率</strong><code>Σ 新增付费人数 ÷ Σ 反推有效分母</code>。反推分母为每个 cohort 的 <code>新增付费人数 ÷ 来源新增付费率</code>。</div>
      <div><strong>分母覆盖率</strong><code>Σ 反推有效分母 ÷ Σ 新增人数</code>。它衡量来源显示率的统计分母与本报告新增 cohort 的重合程度，不是付费率。</div>
    </div>
    <div class="callout warning"><strong>当前窗口的算法示例：</strong>Σ新增付费人数 = <strong>{integer(overall_current['payment_count'])}</strong>；Σ反推有效分母 = <strong>{integer(overall_current['payment_inferred_denominator'])}</strong>；Σ新增 = <strong>{integer(overall_current['new_users'])}</strong>。因此，来源显示率 = {integer(overall_current['payment_count'])} ÷ {integer(overall_current['payment_inferred_denominator'])} = <strong>{pct(overall_current['payment_source_rate'])}</strong>；分母覆盖率 = {integer(overall_current['payment_inferred_denominator'])} ÷ {integer(overall_current['new_users'])} = <strong>{pct(overall_current['payment_coverage'])}</strong>。覆盖率 72.0% 表示：可反推的来源统计分母约覆盖新增 cohort 的 72%，不能理解为“72% 用户付费”。</div>
    <p class="note">反推仅使用来源显示率大于 0 的 cohort；若“人数=0 且显示率=0”，无法从该行反推其原始分母，应另行以底层字段核对。覆盖率低于 100% 可能来自统计人群、状态筛选、去重键或时间归属不同，不直接等同于数据丢失；高于 100% 同样需要排查范围扩大或重复计数。</p>
    <div class="callout alert"><strong>口径风险已确认：</strong>在 {integer(first_charge_gt_payment)}/{integer(len(records))} 个 cohort 日中，来源“首充付费人数”高于“新增付费人数”。因此报告不制作“新增 → 付费 → 首充”漏斗，也不把首充率解释为新增付费子集转化。需由数据开发明确两项的用户范围、去重键和时间归属。</div>
    <div class="cols-3" style="margin-top:16px">
      <article class="highlight"><h4>iOS 是高质量基准</h4><p>iOS App Store D7 {pct(ios['d7'])}，付费人数率 {pct(ios['payment_per_new'])}，但新增仅 {pct(ios['share'])}。可提炼首局/首付路径，不可直接外推媒体效果。</p></article>
      <article class="highlight green"><h4>WajeBet H5 付费值得复核</h4><p>人数率 {pct(bet_h5['payment_per_new'])}，来源显示率 {pct(bet_h5['payment_source_rate'])}，覆盖率 {pct(bet_h5['payment_coverage'])}。改善留存的同时须验证分母是否一致。</p></article>
      <article class="highlight yellow"><h4>H5 Google 的覆盖不足</h4><p>人数率 {pct(h5_ga['payment_per_new'])}，显示率 {pct(h5_ga['payment_source_rate'])}，覆盖率仅 {pct(h5_ga['payment_coverage'])}。展示率较高不等于新用户转化同步改善。</p></article>
    </div>
  </section>

  <section id="actions">
    <div class="section-top"><div><h2><span class="num">6</span>问题总结与可执行优化方案</h2><p class="muted">结论优先指向可验证的产品、投放和数据治理动作，不将相关性写成因果。</p></div><p class="period">建议按 7 / 14 / 28 天复盘</p></div>
    <div class="action-grid">
      <article class="action"><span class="priority p0">P0 · H5 Facebook 新手首周</span><h4>优先定位“高量低留存”的断点</h4><p class="muted">证据：新增 {integer(h5_fb['new_users'])}，D1/D3/D7 = {pct(h5_fb['d1'])}/{pct(h5_fb['d3'])}/{pct(h5_fb['d7'])}；D7 较 H5 Google 低 {pct((h5_ga['d7'] or 0)-(h5_fb['d7'] or 0))}。</p><ul><li>按渠道、落地页、地区、设备档位、网络、H5 build 串联：曝光→加载成功→注册完成→首局开始→首局完成→首付。</li><li>优先排查加载失败、白屏、游戏资源与首局中断；投放侧排查 campaign 与素材承诺偏差。</li><li>验证：加载成功率、首局完成率、D1/D3/D7、首付人数率；护栏：异常率、成本与作弊风险。</li></ul></article>
      <article class="action"><span class="priority p0">P0 · 付费报表口径</span><h4>统一新增付费与首充的分子、分母、去重</h4><p class="muted">证据：整体人数率 {pct(overall_current['payment_per_new'])}，来源显示率 {pct(overall_current['payment_source_rate'])}，有效分母覆盖 {pct(overall_current['payment_coverage'])}；首充人数几乎全量高于新增付费人数。</p><ul><li>冻结指标定义：观察对象、注册/支付时间归属、成功状态、退款处理、用户去重键。</li><li>在起源和 BI 同时输出“人数/新增”“来源显示率”“覆盖率”，禁止同名“付费率”混用。</li><li>验收：同日期、同渠道、同包体下明细可回溯；差异超过 1% 自动告警。</li></ul></article>
      <article class="action"><span class="priority p1">P1 · Android Google 质量下滑</span><h4>验证高规模渠道的留存下滑来源</h4><p class="muted">证据：Google Ads D7 {pct(android_ga['d7'])}（{signed_pp(android_ga['d7_delta'])}）；Google Play D7 {pct(android_play['d7'])}（{signed_pp(android_play['d7_delta'])}）。</p><ul><li>将媒体计划、归因 campaign、安装包版本、首局游戏和异常码接入同一 cohort 视图。</li><li>做版本/活动前后 14 天描述性对比；无对照组不输出因果结论。</li><li>验证：D1、首局完成、D3、D7和异常率；护栏：新增规模、CPA、支付成功率。</li></ul></article>
      <article class="action"><span class="priority p1">P1 · H5 改善样本复用</span><h4>验证 WajeBet H5 与 H5 Google 的改善是否可复制</h4><p class="muted">证据：WajeBet H5 D7 {pct(bet_h5['d7'])}（{signed_pp(bet_h5['d7_delta'])}）；H5 Google D7 {pct(h5_ga['d7'])}（{signed_pp(h5_ga['d7_delta'])}），但量级下降 {signed_pct(h5_ga['new_delta'])}。</p><ul><li>拆解改善发生在哪个新手阶段、媒体计划和页面/游戏组合；补齐 H5 build 与发布时间。</li><li>优先用小流量灰度，将被验证的入口或首局改动复制到 H5 Facebook。</li><li>复盘：7 天看 D1/首局，14 天看 D3，28 天看成熟 D7；保持渠道结构可比。</li></ul></article>
      <article class="action"><span class="priority p2">P2 · PWW 稳定性</span><h4>先补归因，再决定资源投入</h4><p class="muted">证据：当前窗口新增 {integer(pww['new_users'])}，低于主渠道量级，D7 变动 {signed_pp(pww['d7_delta'])}。</p><ul><li>补齐媒体、campaign、包体与真实版本；统一原始日期格式和日更新检查。</li><li>以连续两个成熟 D7 窗口、且样本达到预设阈值后再评估投放扩量。</li></ul></article>
      <article class="action"><span class="priority p2">P2 · 补齐真实版本与性能维度</span><h4>让下一轮分析能定位产品原因</h4><p class="muted">当前源表只有包体/渠道，不含 build、设备、网络、页面性能或新手关键事件。</p><ul><li>新增字段：app_version、web_build、release_id、device_tier、network_type、load/result/error_code。</li><li>新增事件：landing_view、load_success、register_complete、game_start、game_end、payment_initiated、payment_success。</li><li>目标：每个版本可按渠道、端、设备和首局路径复核 D1/D3/D7 与付费。</li></ul></article>
    </div>
  </section>

  <section id="quality" class="tint">
    <div class="section-top"><div><h2><span class="num">7</span>数据质量、映射与统计边界</h2><p class="muted">将“数据为零”“数据未成熟”“数据缺失”和“口径不一致”明确区分。</p></div><p class="period">源文件只读；不修改原始工作簿</p></div>
    <div class="quality-summary">
      <div><strong>{integer(len(records))}</strong><span>有效 cohort 日记录</span></div>
      <div><strong>{integer(len(SHEET_MAP))}</strong><span>源工作表 / 包体渠道序列</span></div>
      <div><strong>{integer(quality['serial_repairs'])}</strong><span>PWW Excel 序列日期已标准化</span></div>
      <div><strong>{integer(sum(detail['missing_days'] for detail in quality['sheet_quality'].values()))}</strong><span>各表日期范围内缺失日</span></div>
    </div>
    <div class="callout warning"><strong>版本边界：</strong>当前文件没有真实的 <code>app_version</code>、<code>web_version</code>、<code>build</code> 或上线事件字段；本报告的“版本”仅指包体/产品形态。不得据此认定某个软件发布版本导致留存或付费变化。</div>
    <h3>源表连续性与映射</h3>
    <div class="table-wrap"><table class="quality-table"><thead><tr><th>源工作表</th><th class="num">有效记录</th><th>最早 cohort</th><th>最新 cohort</th><th class="num">范围内缺失日</th></tr></thead><tbody>{''.join(quality_rows)}</tbody></table></div><p class="table-hint">表格可左右滑动查看完整指标。</p>
    <h3>统一口径</h3>
    <div class="formula"><div><strong>成熟留存</strong>Dk 仅统计 cohort_date ≤ 数据截至日 − k 的记录；未成熟值为 N/A，不参与分母。</div><div><strong>加权留存</strong>Σ（cohort 留存率 × cohort 新增）÷ Σ cohort 新增。避免简单平均放大小样本。</div><div><strong>可追溯性</strong>每一条汇总均保留 source_sheet、cohort_date、包体、渠道和原始指标字段。</div></div>
    <p class="note">PWW 的安装形态为 App、游戏运行形态为 H5；本报告将其作为独立第三类，不计入 App 或标准 H5，整体指标仍包含 PWW。区服字段在全部源记录中均为 “Waje Special”，没有可解释差异，未作为图表切分维度。C‑T、TC 比、TX 率和人均 TX 缺少统一底层分子分母，未进入本报告的留存/付费结论。</p>
  </section>

  <footer class="footer">数据源：<span class="source-chip">新包分析2026.8.11_new.xlsx</span><span class="source-chip">8 张 cohort 日表</span><span class="source-chip">截至 {cutoff:%Y-%m-%d}</span><br>本报告为描述性分析；渠道、版本、活动和产品体验之间的因果关系须通过补充字段与对照实验验证。</footer>
</main>
</body>
</html>"""


def main() -> None:
    records, quality = load_records()
    if len(records) != 2120:
        raise AssertionError(f"有效记录应为 2120，实际为 {len(records)}")
    if quality["max_date"] != date(2026, 8, 10):
        raise AssertionError(f"最大日期应为 2026-08-10，实际为 {quality['max_date']}")
    current = filtered(records, date(2026, 7, 7), date(2026, 8, 3))
    previous = filtered(records, date(2026, 6, 9), date(2026, 7, 6))
    if len(current) != 224 or len(previous) != 224:
        raise AssertionError(f"成熟 D7 窗口记录数错误：{len(current)}/{len(previous)}")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(build_report(records, quality), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
